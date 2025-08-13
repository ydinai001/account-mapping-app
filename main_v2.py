#!/usr/bin/env python3
"""
Multi-Project Account Mapping Desktop Application v2

A tool to automatically map account descriptions between multiple P&L projects
in Excel workbooks with pattern recognition, manual editing capabilities, 
and monthly P&L statement generation.

Version 2 Features:
- Multi-project support (Excel workbooks with multiple sheets)
- Project switching and isolation
- Project-specific settings and data storage
- Enhanced UI with project selection menu
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import os
import platform
import re
import json
import time
from typing import Dict, List, Optional, Tuple
from collections import OrderedDict

# Import our project management classes
from project_manager import ProjectManager, Project

# Performance profiling decorator
def profile_performance(method_name):
    def decorator(func):
        def wrapper(*args, **kwargs):
            start_time = time.time()
            result = func(*args, **kwargs)
            end_time = time.time()
            duration = end_time - start_time
            # Performance tracking disabled
            return result
        return wrapper
    return decorator


class MultiProjectAccountMappingApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Account Mapping Tool v2 - Multi-Project")
        
        # Set window size (reduced by 50% for better screen compatibility)
        self.root.minsize(993, 662)  # 50% smaller for better usability on smaller screens
        
        # Center the window on the screen
        self.center_window(993, 662)
        
        # Project management
        self.project_manager = ProjectManager()
        
        # UI Variables
        self.project_var = tk.StringVar()
        self.source_range_var = tk.StringVar()
        self.rolling_range_var = tk.StringVar()
        self.status_var = tk.StringVar()
        self.step4_status_var = tk.StringVar()
        self.filter_var = tk.StringVar()
        self.sort_var = tk.StringVar()
        
        # Zoom functionality
        self.zoom_levels = [0.75, 0.85, 1.0, 1.15, 1.3, 1.5]
        self.current_zoom_index = 2  # Default to 1.0 (100%)
        self.current_zoom_level = 1.0
        
        # Checkbox management for bulk editing
        self.checkbox_states = {}
        
        # Flag to prevent trace callbacks during project loading
        self._loading_project_data = False
        
        # Flag to prevent event handling during initialization
        self._initializing = True
        
        # Popup window state
        self.popup_window = None
        self.popup_tree = None
        self.is_popped_out = False
        
        # Step 3 popup state
        self.step3_popup_window = None
        self.step3_popup_tree = None
        self.step3_popup_generate_button = None
        self.step3_is_popped_out = False
        
        # Step 3 state tracking
        self.mappings_modified_after_generation = False
        self.projects_with_session_generated_data = set()  # Track projects with data generated in current session
        self.projects_with_modified_mappings = set()  # Track projects where mappings were manually edited
        
        # Filter timer for delayed filtering
        self.filter_timer = None
        
        # Performance optimization: DataFrame and computation caching
        self.dataframe_cache = {}  # Cache for loaded Excel files
        self.file_timestamps = {}  # Track file modification times
        self.mapping_signatures = {}  # Store mapping signatures to detect changes
        self.fuzzy_match_cache = {}  # Cache for fuzzy string matching results
        self.last_generated_mappings = {}  # Store last generated mappings per project
        self.rolling_accounts_cache = {}  # Cache rolling accounts per project to speed up edit dialogs
        self.target_month_cache = {}  # Cache target month column to speed up statement regeneration
        self.source_amounts_cache = {}  # Cache source amounts to avoid repeated extraction
        # Range data cache removed - always extract fresh data to detect new accounts
        
        # Initialize UI state
        self.status_var.set("Welcome! Please upload a Source P&L workbook to begin.")
        
        # Create GUI
        self.create_widgets()
        
        # Configure styles
        self.setup_treeview_styles()
        self.apply_header_alignment()
        self.setup_keyboard_shortcuts()
        
        # Load existing projects if any
        self.refresh_project_menu()
        self.update_ui_state()
        
        # Apply initial zoom
        self.apply_zoom()
        
        # Load project-specific data if a project was previously selected
        current_project = self.project_manager.get_current_project()
        if current_project:
            self.load_project_data()
        else:
            self._loading_project_data = True
            try:
                self.load_range_settings()
            finally:
                self._loading_project_data = False
        
        # Delay clearing initialization flag until UI is fully rendered
        # This prevents events from firing during the initial UI rendering phase
        self.root.after(100, self._complete_initialization)
    
    def _complete_initialization(self):
        """Complete the initialization process after UI is fully rendered"""
        self._initializing = False
    
    def center_window(self, width, height):
        """Center the window on the screen"""
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        
        self.root.geometry(f"{width}x{height}+{x}+{y}")
    
    # ========== Performance Optimization Methods ==========
    
    def _get_file_timestamp(self, filepath):
        """Get file modification timestamp."""
        try:

            pass
            return os.path.getmtime(filepath)
        except:

            pass
            return None
    
    def _is_file_cached(self, filepath, sheet_name=None):
        """Check if file is cached and up to date."""
        cache_key = f"{filepath}:{sheet_name}" if sheet_name else filepath
        
        if cache_key not in self.dataframe_cache:

        
            pass
            return False
        
        current_timestamp = self._get_file_timestamp(filepath)
        cached_timestamp = self.file_timestamps.get(filepath)
        
        return current_timestamp == cached_timestamp
    
    def _load_excel_with_cache(self, filepath, sheet_name=None):
        """Load Excel file with caching to avoid repeated reads.
        Uses openpyxl with data_only=True to read calculated values from formulas."""
        cache_key = f"{filepath}:{sheet_name}" if sheet_name else filepath
        
        # Check if we have a valid cache entry
        if self._is_file_cached(filepath, sheet_name):
            self.status_var.set(f"Using cached data for {os.path.basename(filepath)}")
            return self.dataframe_cache[cache_key]
        
        # Load the file
        self.status_var.set(f"Loading {os.path.basename(filepath)}...")
        try:
            # Use openpyxl directly with data_only=True to read calculated values instead of formulas
            from openpyxl import load_workbook
            
            # Load workbook with data_only=True to get calculated values
            wb = load_workbook(filepath, data_only=True)
            
            # If sheet_name is specified, use it; otherwise use the first sheet
            if sheet_name:

                pass
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    raise ValueError(f"Sheet '{sheet_name}' not found in workbook")
            else:
                ws = wb.active
            
            # Convert worksheet to pandas DataFrame
            # Get all values from the worksheet
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(list(row))
            
            # Create DataFrame from the data
            if data:
                df = pd.DataFrame(data)
                
                # Use first row as headers if it looks like headers
                if len(data) > 1:
                    # Check if first row contains mostly strings (likely headers)
                    first_row = data[0]
                    if first_row and sum(1 for cell in first_row if isinstance(cell, str) and cell) > len(first_row) * 0.3:
                        df.columns = [str(col) if col is not None else f"Column_{i}" for i, col in enumerate(first_row)]
                        df = df.iloc[1:].reset_index(drop=True)
            else:
                df = pd.DataFrame()
            
            # Update cache
            self.dataframe_cache[cache_key] = df
            self.file_timestamps[filepath] = self._get_file_timestamp(filepath)
            
            return df
        except Exception as e:
            raise e
    
    def _clear_cache_for_file(self, filepath):
        """Clear cache entries for a specific file."""
        # Remove all cache entries that start with this filepath
        keys_to_remove = [k for k in self.dataframe_cache.keys() if k.startswith(filepath)]
        for key in keys_to_remove:
            del self.dataframe_cache[key]
        
        # Also clear target month cache if this is a source file
        current_project = self.project_manager.get_current_project()
        if current_project and filepath == current_project.source_file_path:
            cache_key = f"{current_project.name}:{filepath}"
            if cache_key in self.target_month_cache:
                del self.target_month_cache[cache_key]
        
        # Remove timestamp
        if filepath in self.file_timestamps:
            del self.file_timestamps[filepath]
    
    def _compute_mapping_signature(self, mappings):
        """Compute a signature for mappings to detect changes."""
        if not mappings:

            pass
            return ""
        
        try:
            # Create a string representation of mappings
            mapping_str = ""
            for source_account, mapping_info in mappings.items():
                # Handle both dictionary and string formats (for backward compatibility)
                if isinstance(mapping_info, dict):
                    mapped_account = mapping_info.get('rolling_account', '')
                else:
                    mapped_account = str(mapping_info)
                mapping_str += f"{source_account}|{mapped_account}|"
            
            # Return a hash of the mapping string
            return str(hash(mapping_str))
        except Exception:
            # If signature computation fails, return a fallback value
            return str(len(mappings) if mappings else 0)
    
    def _have_mappings_changed(self, project_name, new_mappings):
        """Check if mappings have changed since last generation."""
        new_signature = self._compute_mapping_signature(new_mappings)
        old_signature = self.mapping_signatures.get(project_name, "")
        
        return new_signature != old_signature
    
    def _update_mapping_signature(self, project_name, mappings):
        """Update the stored mapping signature."""
        self.mapping_signatures[project_name] = self._compute_mapping_signature(mappings)
    
    def _get_fuzzy_match_key(self, str1, str2):
        """Generate a cache key for fuzzy string matching."""
        # Sort strings to ensure consistent key regardless of order
        return tuple(sorted([str1.lower(), str2.lower()]))
    
    def _get_cached_fuzzy_score(self, str1, str2):
        """Get cached fuzzy matching score if available."""
        key = self._get_fuzzy_match_key(str1, str2)
        return self.fuzzy_match_cache.get(key)
    
    def _cache_fuzzy_score(self, str1, str2, score):
        """Cache a fuzzy matching score."""
        key = self._get_fuzzy_match_key(str1, str2)
        self.fuzzy_match_cache[key] = score
    
    def create_widgets(self):
        """Create all GUI widgets"""
        # Copyright footer at bottom (pack first to reserve space)
        footer_frame = ttk.Frame(self.root)
        footer_frame.pack(fill=tk.X, side=tk.BOTTOM, padx=10, pady=(0, 10))
        
        copyright_label = ttk.Label(footer_frame, text="© 2025 California Private Capital Group", 
                                   font=("Arial", 9), foreground="gray")
        copyright_label.pack(side=tk.RIGHT)
        
        # Create canvas and scrollbar for main content
        canvas_container = ttk.Frame(self.root)
        canvas_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=(10, 0))
        
        # Create canvas for scrolling
        self.main_canvas = tk.Canvas(canvas_container, highlightthickness=0)
        self.main_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Create vertical scrollbar
        v_scrollbar = ttk.Scrollbar(canvas_container, orient=tk.VERTICAL, command=self.main_canvas.yview)
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Configure canvas to use scrollbar
        self.main_canvas.configure(yscrollcommand=v_scrollbar.set)
        
        # Create scrollable frame inside canvas
        self.scrollable_frame = ttk.Frame(self.main_canvas)
        self.scrollable_frame_id = self.main_canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        
        # Main container inside scrollable frame
        main_frame = ttk.Frame(self.scrollable_frame)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Create header with project info and controls
        self.create_header(main_frame)
        
        # Create main content area with fixed grid layout (no resizable dividers)
        content_frame = ttk.Frame(main_frame)
        content_frame.pack(fill=tk.BOTH, expand=True, pady=(10, 0))
        
        # Configure grid weights - Step 2 gets most space, others get fixed space
        content_frame.columnconfigure(0, weight=1)
        content_frame.rowconfigure(0, weight=0, minsize=150)  # Step 1: Fixed height with minimum
        content_frame.rowconfigure(1, weight=1, minsize=400)  # Step 2: Main area with minimum height for tree
        content_frame.rowconfigure(2, weight=0, minsize=80)   # Step 3: Fixed height with minimum
        content_frame.rowconfigure(3, weight=0, minsize=200)  # Step 3: Fixed height with minimum
        
        # Step 1: File upload section (top, fixed height)
        self.create_step1_upload_section(content_frame, 0)
        
        # Step 2: Mapping section (main area, expandable)
        self.create_step2_mappings_section(content_frame, 1)
        
        # Step 3: Export section (fixed height)
        self.create_step3_export_section(content_frame, 2)
        
        # Step 3: Monthly statements section (fixed height)
        self.create_step4_monthly_section(content_frame, 3)
        
        # Configure scroll region after content is created
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.update_scroll_region()
        )
        
        # Ensure canvas width matches container
        self.main_canvas.bind(
            "<Configure>",
            lambda e: self.configure_canvas_width()
        )
        
        # Bind mouse wheel to canvas for scrolling
        self.bind_mouse_wheel()
    
    def bind_mouse_wheel(self):
        """Bind mouse wheel events for scrolling"""
        # Bind to canvas only, not globally
        # Different bindings for different platforms
        if self.root.tk.call('tk', 'windowingsystem') == 'aqua':  # macOS
            self.main_canvas.bind("<MouseWheel>", self.on_mouse_wheel)
            self.main_canvas.bind("<Shift-MouseWheel>", self.on_shift_mouse_wheel)
        else:  # Windows and Linux
            self.main_canvas.bind("<MouseWheel>", self.on_mouse_wheel)
            self.main_canvas.bind("<Button-4>", self.on_mouse_wheel_up)
            self.main_canvas.bind("<Button-5>", self.on_mouse_wheel_down)
        
        # Also bind to scrollable frame
        if self.root.tk.call('tk', 'windowingsystem') == 'aqua':  # macOS
            self.scrollable_frame.bind("<MouseWheel>", self.on_mouse_wheel)
        else:  # Windows and Linux
            self.scrollable_frame.bind("<MouseWheel>", self.on_mouse_wheel)
            self.scrollable_frame.bind("<Button-4>", self.on_mouse_wheel_up)
            self.scrollable_frame.bind("<Button-5>", self.on_mouse_wheel_down)
    
    def on_mouse_wheel(self, event):
        """Handle mouse wheel scrolling"""
        # Check if we're over a widget that should handle its own scrolling
        widget = self.root.winfo_containing(event.x_root, event.y_root)
        
        # Check if the widget under mouse is a Treeview or Text widget (they handle their own scrolling)
        if widget:
            widget_class = widget.winfo_class()
            # Check widget and its parents for scrollable widgets
            check_widget = widget
            while check_widget:
                if isinstance(check_widget, (ttk.Treeview, tk.Text, tk.Listbox)):
                    # Let the widget handle its own scrolling
                    return
                try:
                    check_widget = check_widget.master
                except:
                    break
        
        # Only scroll main canvas if we're not over a scrollable widget
        if self.main_canvas.winfo_exists():
            if self.root.tk.call('tk', 'windowingsystem') == 'aqua':  # macOS
                self.main_canvas.yview_scroll(int(-1 * event.delta), "units")
            else:  # Windows
                self.main_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
    
    def on_shift_mouse_wheel(self, event):
        """Handle horizontal scrolling with shift+wheel (macOS)"""
        # For horizontal scrolling if needed in future
        pass
    
    def on_mouse_wheel_up(self, event):
        """Handle mouse wheel up for Linux"""
        # Check if we're over a scrollable widget
        widget = self.root.winfo_containing(event.x_root, event.y_root)
        if widget:
            check_widget = widget
            while check_widget:
                if isinstance(check_widget, (ttk.Treeview, tk.Text, tk.Listbox)):
                    return
                try:
                    check_widget = check_widget.master
                except:
                    break
        
        if self.main_canvas.winfo_exists():
            self.main_canvas.yview_scroll(-1, "units")
    
    def on_mouse_wheel_down(self, event):
        """Handle mouse wheel down for Linux"""
        # Check if we're over a scrollable widget
        widget = self.root.winfo_containing(event.x_root, event.y_root)
        if widget:
            check_widget = widget
            while check_widget:
                if isinstance(check_widget, (ttk.Treeview, tk.Text, tk.Listbox)):
                    return
                try:
                    check_widget = check_widget.master
                except:
                    break
        
        if self.main_canvas.winfo_exists():
            self.main_canvas.yview_scroll(1, "units")
    
    def update_scroll_region(self):
        """Update the scroll region to encompass all content"""
        self.main_canvas.configure(scrollregion=self.main_canvas.bbox("all"))
    
    def configure_canvas_width(self):
        """Configure canvas width to match container"""
        canvas_width = self.main_canvas.winfo_width()
        self.main_canvas.itemconfig(self.scrollable_frame_id, width=canvas_width)
    
    def create_header(self, parent):
        """Create header with project title and controls"""
        header_frame = ttk.Frame(parent)
        header_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Left side - project title
        title_frame = ttk.Frame(header_frame)
        title_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        self.project_title_label = ttk.Label(
            title_frame, 
            text="No Project Selected", 
            font=("Arial", 16, "bold")
        )
        self.project_title_label.pack(side=tk.LEFT)
        
        # Target month display
        self.target_month_label = ttk.Label(
            title_frame,
            text="",
            font=("Arial", 14, "bold"),
            foreground="darkblue"
        )
        self.target_month_label.pack(side=tk.LEFT, padx=(20, 0))
        
        # Right side - project menu and controls
        controls_frame = ttk.Frame(header_frame)
        controls_frame.pack(side=tk.RIGHT)
        
        # Project selection menu
        ttk.Label(controls_frame, text="Project:").pack(side=tk.LEFT, padx=(0, 5))
        
        self.project_menu = ttk.Combobox(
            controls_frame, 
            textvariable=self.project_var,
            width=20,
            state="readonly"
        )
        self.project_menu.pack(side=tk.LEFT, padx=(0, 10))
        self.project_menu.bind("<<ComboboxSelected>>", self.on_project_selected)
        
        # Reset/Clear buttons
        self.clear_project_button = ttk.Button(
            controls_frame,
            text="Clear Project",
            command=self.clear_current_project,
            width=12
        )
        self.clear_project_button.pack(side=tk.LEFT, padx=(5, 5))
        
        self.clear_all_button = ttk.Button(
            controls_frame,
            text="Start Fresh",
            command=self.clear_all_projects,
            width=12
        )
        self.clear_all_button.pack(side=tk.LEFT, padx=(5, 5))
        
        # Backup functionality buttons
        backup_frame = ttk.Frame(controls_frame)
        backup_frame.pack(side=tk.LEFT, padx=(10, 0))
        
        self.backup_button = ttk.Button(
            backup_frame,
            text="Create Backup",
            command=self.create_backup,
            width=12
        )
        self.backup_button.pack(side=tk.TOP, pady=(0, 2))
        
        self.load_backup_button = ttk.Button(
            backup_frame,
            text="Load Backup",
            command=self.show_backup_menu,
            width=12
        )
        self.load_backup_button.pack(side=tk.TOP)
        
    
    def create_step1_upload_section(self, parent, row):
        """Create Step 1: Upload and Project Selection"""
        step1_frame = ttk.LabelFrame(parent, text="Step 1: Upload Source P&L Workbook & Select Project", padding="5")
        step1_frame.grid(row=row, column=0, sticky="nsew", pady=(0, 5))
        step1_frame.columnconfigure(1, weight=1)
        
        current_row = 0
        
        # Source P&L workbook upload
        ttk.Label(step1_frame, text="Source P&L Workbook:").grid(row=current_row, column=0, sticky="w", padx=(0, 10))
        
        self.source_file_var = tk.StringVar()
        source_entry = ttk.Entry(step1_frame, textvariable=self.source_file_var, state="readonly")
        source_entry.grid(row=current_row, column=1, sticky="we", padx=(0, 10))
        
        ttk.Button(step1_frame, text="Browse...", command=self.select_source_workbook).grid(row=current_row, column=2, columnspan=2, padx=(0, 10), sticky="w")
        
        current_row += 1
        
        # Project list status
        self.projects_status_label = ttk.Label(step1_frame, text="No projects loaded", foreground="gray")
        self.projects_status_label.grid(row=current_row, column=0, columnspan=4, sticky="w", pady=(5, 0))
        
        current_row += 1
        
        # Separator
        separator1 = ttk.Separator(step1_frame, orient="horizontal")
        separator1.grid(row=current_row, column=0, columnspan=4, sticky="we", pady=10)
        
        current_row += 1
        
        # Rolling P&L workbook upload (disabled until project selected)
        ttk.Label(step1_frame, text="Rolling P&L Workbook:").grid(row=current_row, column=0, sticky="w", padx=(0, 10))
        
        self.rolling_file_var = tk.StringVar()
        self.rolling_entry = ttk.Entry(step1_frame, textvariable=self.rolling_file_var, state="readonly")
        self.rolling_entry.grid(row=current_row, column=1, sticky="we", padx=(0, 10))
        
        self.rolling_browse_button = ttk.Button(step1_frame, text="Browse...", command=self.select_rolling_workbook, state="disabled")
        self.rolling_browse_button.grid(row=current_row, column=2, columnspan=2, padx=(0, 10), sticky="w")
        
        current_row += 1
        
        # Rolling sheet selection dropdown
        self.rolling_sheet_frame = ttk.Frame(step1_frame)
        self.rolling_sheet_frame.grid(row=current_row, column=0, columnspan=4, sticky="we", pady=(5, 0))
        self.rolling_sheet_frame.columnconfigure(1, weight=1)
        
        self.rolling_sheet_label = ttk.Label(self.rolling_sheet_frame, text="Rolling Sheet:", state="disabled")
        self.rolling_sheet_label.grid(row=0, column=0, padx=(0, 5), sticky="w")
        
        # Simple dropdown for sheet selection
        self.rolling_sheet_var = tk.StringVar()
        self.rolling_sheet_combo = ttk.Combobox(self.rolling_sheet_frame, textvariable=self.rolling_sheet_var, 
                                               state="disabled", width=30)
        self.rolling_sheet_combo.grid(row=0, column=1, padx=(0, 5), sticky="w")
        self.rolling_sheet_combo.bind('<<ComboboxSelected>>', self.on_rolling_sheet_selected)
        
        # Note: Rolling sheet selection is not memorized (removed per user request)
        
        current_row += 1
        
        # Separator
        separator2 = ttk.Separator(step1_frame, orient="horizontal")
        separator2.grid(row=current_row, column=0, columnspan=4, sticky="we", pady=10)
        
        current_row += 1
        
        # Range specifications (disabled until project selected)
        ranges_frame = ttk.Frame(step1_frame)
        ranges_frame.grid(row=current_row, column=0, columnspan=4, sticky="w")
        
        # Source range
        self.source_range_label = ttk.Label(ranges_frame, text="Source Range:", state="disabled")
        self.source_range_label.pack(side=tk.LEFT, padx=(0, 5))
        
        self.source_range_entry = ttk.Entry(ranges_frame, textvariable=self.source_range_var, width=10, state="disabled")
        self.source_range_entry.pack(side=tk.LEFT, padx=(0, 10))
        self.source_range_var.trace_add("write", self.save_range_settings)
        
        self.source_preview_button = ttk.Button(ranges_frame, text="Preview", command=lambda: self.preview_range("source"), state="disabled")
        self.source_preview_button.pack(side=tk.LEFT, padx=(0, 20))
        
        # Rolling range
        self.rolling_range_label = ttk.Label(ranges_frame, text="Rolling Range:", state="disabled")
        self.rolling_range_label.pack(side=tk.LEFT, padx=(0, 5))
        
        self.rolling_range_entry = ttk.Entry(ranges_frame, textvariable=self.rolling_range_var, width=10, state="disabled")
        self.rolling_range_entry.pack(side=tk.LEFT, padx=(0, 10))
        self.rolling_range_var.trace_add("write", self.save_range_settings)
        
        self.rolling_preview_button = ttk.Button(ranges_frame, text="Preview", command=lambda: self.preview_range("rolling"), state="disabled")
        self.rolling_preview_button.pack(side=tk.LEFT)
        
        current_row += 1
        
        # Generate mappings button (disabled until ranges set)
        button_frame = ttk.Frame(step1_frame)
        button_frame.grid(row=current_row, column=0, columnspan=4, pady=(10, 0))
        
        self.generate_button = ttk.Button(
            button_frame, 
            text="Generate Mappings", 
            command=self.generate_mappings,
            state="disabled"
        )
        self.generate_button.pack()
        
        # Load mapping file option
        mapping_frame = ttk.Frame(step1_frame)
        mapping_frame.grid(row=current_row + 1, column=0, columnspan=4, pady=(10, 0))
        
        ttk.Label(mapping_frame, text="Optional - Load Mapping File:").pack(side=tk.LEFT, padx=(0, 10))
        
        self.mapping_file_var = tk.StringVar()
        self.mapping_entry = ttk.Entry(mapping_frame, textvariable=self.mapping_file_var, state="readonly", width=40)
        self.mapping_entry.pack(side=tk.LEFT, padx=(0, 10))
        
        self.mapping_browse_button = ttk.Button(mapping_frame, text="Browse...", command=self.select_mapping_file, state="disabled")
        self.mapping_browse_button.pack(side=tk.LEFT)
        
    
    def create_step2_mappings_section(self, parent, row):
        """Create Step 2: Review & Edit Mappings - Adapted from main.py"""
        step2_frame = ttk.LabelFrame(parent, text="Step 2: Review & Edit Mappings", padding="5")
        step2_frame.grid(row=row, column=0, sticky="nsew", pady=(5, 5), padx=5)
        step2_frame.columnconfigure(0, weight=1)
        step2_frame.columnconfigure(1, weight=0)  # Scrollbar column should not expand 
        step2_frame.rowconfigure(3, weight=1)  # Make the tree row expandable
        step2_frame.rowconfigure(0, weight=0)  # Instructions row fixed
        step2_frame.rowconfigure(1, weight=0)  # Filter row fixed  
        step2_frame.rowconfigure(2, weight=0)  # Buttons row fixed
        
        # Instructions for editing - make it more visible
        instructions_frame = ttk.Frame(step2_frame)
        instructions_frame.grid(row=0, column=0, columnspan=2, sticky="we", pady=(0, 5))
        
        instruction_label = ttk.Label(instructions_frame, text="📝 Double-click or press Enter to edit mapping • Space bar to toggle checkboxes • Use checkboxes for bulk editing", 
                 font=("Arial", 11, "bold"), foreground="white", background="darkblue")
        instruction_label.pack(side=tk.LEFT, padx=5, pady=2, fill=tk.BOTH, expand=True)
        
        # Pop-out button
        self.popup_button = ttk.Button(instructions_frame, text="🔲 Pop Out", command=self.pop_out_mapping_window, width=12)
        self.popup_button.pack(side=tk.RIGHT, padx=5, pady=2)
        
        # Filter and sort controls
        filter_frame = ttk.Frame(step2_frame)
        filter_frame.grid(row=1, column=0, columnspan=2, sticky="we", pady=(5, 5))
        
        # Filter controls
        ttk.Label(filter_frame, text="🔍 Filter:").pack(side=tk.LEFT, padx=(0, 5))
        self.filter_var = tk.StringVar()
        self.filter_var.trace_add('write', self.apply_filter)
        self.filter_entry = ttk.Entry(filter_frame, textvariable=self.filter_var, width=30)
        self.filter_entry.pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Button(filter_frame, text="Clear", 
                  command=self.clear_filter).pack(side=tk.LEFT, padx=(0, 10))
        
        # Sort controls
        ttk.Label(filter_frame, text="📊 Sort by:").pack(side=tk.LEFT, padx=(10, 5))
        self.sort_var = tk.StringVar(value="Original Order")
        sort_combo = ttk.Combobox(filter_frame, textvariable=self.sort_var, 
                                 values=["Original Order", "Account Description A-Z", "Account Description Z-A", 
                                        "Mapped Account A-Z", "Mapped Account Z-A", "Confidence Level"], 
                                 width=20, state="readonly")
        sort_combo.pack(side=tk.LEFT, padx=(0, 5))
        sort_combo.bind('<<ComboboxSelected>>', self.apply_sort)
        
        # Selection counter (keep only this for user feedback)
        selection_frame = ttk.Frame(step2_frame)
        selection_frame.grid(row=2, column=0, columnspan=2, sticky="w", pady=(0, 10))
        
        self.selection_label = ttk.Label(selection_frame, text="0 items selected", font=("Arial", 9), foreground="gray")
        self.selection_label.pack(side=tk.LEFT)
        
        # Create treeview for mappings with checkbox column and source data - 20 visible rows
        columns = ("Select", "Account Description", "Source Amount", "Mapped Account", "Confidence")
        self.mapping_tree = ttk.Treeview(step2_frame, columns=columns, show="headings", height=20)
        
        # Configure larger font for the treeview
        style = ttk.Style()
        style.configure("Large.Treeview", font=("Arial", 14))
        self.mapping_tree.configure(style="Large.Treeview")
        
        # Configure columns with left alignment
        self.mapping_tree.heading("Select", text="☐", anchor=tk.W)
        self.mapping_tree.heading("Account Description", text="Account Description", anchor=tk.W)
        self.mapping_tree.heading("Source Amount", text="Source Amount", anchor=tk.W)
        self.mapping_tree.heading("Mapped Account", text="Mapped Account", anchor=tk.W)
        self.mapping_tree.heading("Confidence", text="Confidence", anchor=tk.W)
        
        self.mapping_tree.column("Select", width=80, anchor="center")
        self.mapping_tree.column("Account Description", width=500)
        self.mapping_tree.column("Source Amount", width=150, anchor="w")
        self.mapping_tree.column("Mapped Account", width=400)
        self.mapping_tree.column("Confidence", width=120)
        
        # Configure tags for bold formatting of headings
        self.mapping_tree.tag_configure("heading", font=("Arial", 16, "bold"))
        self.mapping_tree.tag_configure("normal", font=("Arial", 14))
        
        # Scrollbar for treeview
        scrollbar = ttk.Scrollbar(step2_frame, orient="vertical", command=self.mapping_tree.yview)
        self.mapping_tree.configure(yscrollcommand=scrollbar.set)
        
        self.mapping_tree.grid(row=3, column=0, sticky="nsew", padx=(0, 5), pady=(5, 5))
        scrollbar.grid(row=3, column=1, sticky="ns", pady=(5, 5))
        
        # Add placeholder text when no mappings are loaded
        self.add_placeholder_text()
        
        # Bind events
        self.mapping_tree.bind("<Double-1>", self.edit_mapping)  # Double-click for single edit
        self.mapping_tree.bind("<Return>", self.edit_mapping)  # Enter key for single edit
        self.mapping_tree.bind("<Button-1>", self.on_tree_click)  # Single click for checkbox toggle
        self.mapping_tree.bind("<Up>", self.on_arrow_key_navigation)  # Up arrow key navigation
        self.mapping_tree.bind("<Down>", self.on_arrow_key_navigation)  # Down arrow key navigation
        self.mapping_tree.bind("<space>", self.on_space_key_toggle)  # Space bar to toggle checkbox
        self.mapping_tree.bind("<Key-space>", self.on_space_key_toggle)  # Alternative space binding
        self.mapping_tree.bind("<FocusIn>", lambda e: None)  # Enable keyboard focus
        
        # Platform-specific right-click binding
        if platform.system() == "Darwin":  # macOS
            self.mapping_tree.bind("<Button-2>", self.show_context_menu)
            self.mapping_tree.bind("<Control-Button-1>", self.show_context_menu)
        else:  # Windows/Linux
            self.mapping_tree.bind("<Button-3>", self.show_context_menu)
        
        # Create context menu
        self.context_menu = tk.Menu(self.root, tearoff=0)
        self.context_menu.add_command(label="Edit Selected Categories", command=self.bulk_edit_mappings)
        self.context_menu.add_separator()
        self.context_menu.add_command(label="Select All", command=self.select_all_items)
        self.context_menu.add_command(label="Deselect All", command=self.deselect_all_items)
    
    def create_step3_export_section(self, parent, row):
        """Create Step 3: Export Mappings"""
        step3_frame = ttk.LabelFrame(parent, text="", padding="3")
        step3_frame.grid(row=row, column=0, sticky="nsew", pady=(0, 5))
        step3_frame.columnconfigure(0, weight=1)
        
        # Button frame
        button_frame = ttk.Frame(step3_frame)
        button_frame.grid(row=0, column=0, pady=(5, 0))
        
        self.save_mappings_button = ttk.Button(
            button_frame, 
            text="Save User Mappings", 
            command=self.save_user_mappings, 
            state="disabled"
        )
        self.save_mappings_button.grid(row=0, column=0, padx=(0, 5))
        
        self.reset_mappings_button = ttk.Button(
            button_frame, 
            text="Reset Mappings", 
            command=self.reset_mappings, 
            state="disabled"
        )
        self.reset_mappings_button.grid(row=0, column=1)
        
        # Status label
        status_label = ttk.Label(step3_frame, textvariable=self.status_var)
        status_label.grid(row=1, column=0, pady=(10, 0))
    
    def create_step4_monthly_section(self, parent, row):
        """Create Step 3: Generate Monthly P&L Statement - Adapted from main.py"""
        step4_frame = ttk.LabelFrame(parent, text="Step 3: Generate Monthly P&L Statement", padding="3")
        step4_frame.grid(row=row, column=0, sticky="nsew", pady=(0, 3))
        step4_frame.columnconfigure(0, weight=1)
        step4_frame.rowconfigure(2, weight=1)
        
        # Instructions
        instructions_frame = ttk.Frame(step4_frame)
        instructions_frame.grid(row=0, column=0, columnspan=2, sticky="we", pady=(0, 10))
        
        ttk.Label(instructions_frame, text="📊 Extract monthly amounts from column I, aggregate by mappings, and generate rolling P&L statement", 
                 font=("Arial", 11, "bold"), foreground="white", background="darkgreen").pack(side=tk.LEFT, padx=5, pady=2)
        
        # Pop-out button for Step 3 preview
        self.step3_popup_button = ttk.Button(instructions_frame, text="🔲 Pop Out", command=self.pop_out_step3_window, width=12)
        self.step3_popup_button.pack(side=tk.RIGHT, padx=5, pady=2)
        
        # Generate monthly statement button - increased width for text visibility
        self.generate_monthly_button = ttk.Button(step4_frame, text="📊 Generate Monthly Statement", 
                                                 command=self.generate_monthly_statement, state="disabled", width=35)
        self.generate_monthly_button.grid(row=1, column=0, pady=(0, 10), padx=10)
        
        # Preview table frame
        preview_frame = ttk.LabelFrame(step4_frame, text="Preview - Rolling P&L Statement", padding="5")
        preview_frame.grid(row=2, column=0, columnspan=2, sticky="wens", pady=(0, 10))
        preview_frame.columnconfigure(0, weight=1)
        preview_frame.rowconfigure(0, weight=1)
        
        # Create treeview for preview - 4 columns with historical data
        preview_columns = ("Account Description", "Previous Month -2", "Previous Month -1", "Target Month")
        self.preview_tree = ttk.Treeview(preview_frame, columns=preview_columns, show="headings", height=10)
        
        # Configure preview columns - will be updated dynamically with actual month names
        self.preview_tree.heading("Account Description", text="Account Description", anchor=tk.W)
        self.preview_tree.heading("Previous Month -2", text="Previous Month -2", anchor=tk.W)
        self.preview_tree.heading("Previous Month -1", text="Previous Month -1", anchor=tk.W) 
        self.preview_tree.heading("Target Month", text="Target Month", anchor=tk.W) 
        
        self.preview_tree.column("Account Description", width=400)
        self.preview_tree.column("Previous Month -2", width=150, anchor="e")
        self.preview_tree.column("Previous Month -1", width=150, anchor="e")
        self.preview_tree.column("Target Month", width=150, anchor="e")
        
        # Scrollbar for preview
        preview_scrollbar = ttk.Scrollbar(preview_frame, orient="vertical", command=self.preview_tree.yview)
        self.preview_tree.configure(yscrollcommand=preview_scrollbar.set)
        
        self.preview_tree.grid(row=0, column=0, sticky="wens")
        preview_scrollbar.grid(row=0, column=1, sticky="ns")
        
        # Action buttons frame
        action_frame = ttk.Frame(step4_frame)
        action_frame.grid(row=3, column=0, columnspan=2, pady=(10, 0))
        
        # Finalize and export button for current project
        self.finalize_button = ttk.Button(action_frame, text="✅ Finalize & Export Final Excel for Project", 
                                         command=self.finalize_and_export, state="disabled")
        self.finalize_button.pack(side=tk.LEFT, padx=(0, 10))
        
        # Finalize and export button for all projects
        self.finalize_all_button = ttk.Button(action_frame, text="✅ Finalize & Export Final Excel for All Projects", 
                                             command=self.finalize_and_export_all_projects, state="disabled")
        self.finalize_all_button.pack(side=tk.LEFT)
        
        
        # Step 3 status label
        self.step4_status_var = tk.StringVar()
        self.step4_status_label = ttk.Label(step4_frame, textvariable=self.step4_status_var)
        self.step4_status_label.grid(row=4, column=0, columnspan=2, pady=(10, 0))
    
    # Project Management Methods
    def select_source_workbook(self):
        """Select and process source P&L workbook"""
        filename = filedialog.askopenfilename(
            title="Select Source P&L Workbook",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if filename:
            self.source_file_var.set(filename)
            
            # Check if projects already exist
            existing_projects = self.project_manager.get_project_names()
            
            if existing_projects:
                # Ask user if they want to rescan and potentially lose existing data
                result = messagebox.askyesno(
                    "Projects Already Exist",
                    f"Found {len(existing_projects)} existing projects:\n\n" + 
                    "\n".join([f"• {name}" for name in existing_projects]) +
                    "\n\nScanning the workbook will replace existing projects and may cause data loss.\n\n" +
                    "Do you want to proceed with scanning?"
                )
                
                if result:
                    self.projects_status_label.config(text="Scanning workbook for projects...", foreground="blue")
                    self.scan_and_create_projects()
                else:
                    self.projects_status_label.config(text="Source workbook selected - existing projects preserved", foreground="darkgreen")
            else:
                # No existing projects, safe to scan automatically
                self.projects_status_label.config(text="Scanning workbook for projects...", foreground="blue")
                self.scan_and_create_projects()
    
    def scan_and_create_projects(self, silent_mode=False):
        """Scan source workbook and create projects"""
        if getattr(self, '_initializing', False):

            pass
            return
            
        source_path = self.source_file_var.get()
        if not source_path:
            messagebox.showerror("Error", "Please select a Source P&L workbook first.")
            return
        
        try:
            # Scan workbook for projects
            num_projects = self.project_manager.create_projects_from_workbook(source_path)
            
            if num_projects == 0:
                messagebox.showwarning("Warning", "No projects found in the workbook.")
                return
            
            # Initialize new projects with default ranges
            self.initialize_projects_with_defaults()
            
            # Refresh project menu
            self.refresh_project_menu()
            
            # Check if both source and rolling files are ready for auto-processing
            self.check_and_start_auto_process()
            
            # Update UI state to reflect the updated project
            self.update_ui_state()
            
            # Update target month display
            self.update_target_month_display()
            
            # Save the updated project state
            self.project_manager.save_settings()
            
            # Update status
            project_names = self.project_manager.get_project_names()
            self.projects_status_label.config(
                text=f"Found {num_projects} projects: {', '.join(project_names)}",
                foreground="darkgreen"
            )
            
            # Save settings
            self.project_manager.save_settings()
            
            if not silent_mode:
                messagebox.showinfo(
                    "Projects Loaded", 
                    f"Successfully loaded {num_projects} projects:\n\n" + 
                    "\n".join([f"• {name}" for name in project_names]) +
                    "\n\nPlease select a project to continue."
                )
            
        except Exception as e:
            messagebox.showerror("Error", f"Error scanning workbook: {str(e)}")
    
    def initialize_projects_with_defaults(self):
        """Initialize newly created projects with default range settings"""
        default_source_range = self.get_default_source_range()
        default_rolling_range = self.get_default_rolling_range()
        
        # Apply defaults ONLY to projects that don't have ranges set
        # This preserves any restored range memory
        for project_name in self.project_manager.get_project_names():
            project = self.project_manager.projects.get(project_name)
            if project:
                # Only set defaults if ranges are truly empty (not from memory)
                if not project.source_range:
                    project.source_range = default_source_range
                else:
                    pass  # Debug output removed
                    
                if not project.rolling_range:
                    project.rolling_range = default_rolling_range
                else:
                    pass  # Debug output removed
        self.project_manager.save_settings()
    
    def refresh_project_menu(self):
        """Refresh the project selection menu"""
        project_names = self.project_manager.get_project_names()
        self.project_menu['values'] = project_names
        
        # Set current project if available
        current_project = self.project_manager.get_current_project()
        if current_project:
            self.project_var.set(current_project.name)
            self.update_project_title(current_project.name)
        else:
            self.project_var.set("")
            self.update_project_title("No Project Selected")
    
    def on_project_selected(self, event=None):
        """Handle project selection"""
        # Skip if we're still initializing to prevent redundant calls
        if getattr(self, '_initializing', False):

            pass
            return
            
        # Save current project's UI state before switching
        current_project = self.project_manager.get_current_project()
        if current_project:
            self.save_ui_state(current_project)
        
        # Clear caches when switching projects
        self.rolling_accounts_cache.clear()
        self.target_month_cache.clear()
        self.source_amounts_cache.clear()
        # Range data cache removed - no need to clear
        
        project_name = self.project_var.get()
        if project_name and self.project_manager.select_project(project_name):
            self.update_project_title(project_name)
            self.load_project_data()
            self.update_ui_state()
            self.project_manager.save_settings()
    
    def update_project_title(self, project_name):
        """Update the project title display"""
        if project_name and project_name != "No Project Selected":
            self.project_title_label.config(text=f"Project: {project_name}")
            self.root.title(f"Account Mapping Tool v2 - {project_name}")
        else:
            self.project_title_label.config(text="No Project Selected")
            self.root.title("Account Mapping Tool v2 - Multi-Project")
    
    @profile_performance("load_project_data")
    def load_project_data(self):
        """Load data for the currently selected project"""
        current_project = self.project_manager.get_current_project()
        if not current_project:

            pass
            return
        
        self._loading_project_data = True
        
        try:
            # Load project-specific settings with proper defaults for project isolation
            if current_project.source_range:
                self.source_range_var.set(current_project.source_range)
            else:
                # Load global default for projects without saved ranges
                self.source_range_var.set(self.get_default_source_range())
                
            if current_project.rolling_range:
                self.rolling_range_var.set(current_project.rolling_range)
            else:
                # Load global default for projects without saved ranges
                self.rolling_range_var.set(self.get_default_rolling_range())
        finally:
            # Re-enable trace callbacks
            self._loading_project_data = False
        
        # Load UI state (filter, sort, zoom, etc.)
        self.load_ui_state(current_project)
        
        # Load source workbook path if available
        if self.project_manager.source_workbook_path:
            self.source_file_var.set(self.project_manager.source_workbook_path)
        elif current_project.source_file_path:
            self.source_file_var.set(current_project.source_file_path)
        
        # Note: New account detection is now done in load_project_data before populating the tree
        # This ensures new accounts are properly added and saved
        
        # Load rolling workbook path if available
        if self.project_manager.rolling_workbook_path:
            self.rolling_file_var.set(self.project_manager.rolling_workbook_path)
            
            # Populate rolling sheet dropdown when switching projects
            sheet_names = self.project_manager.get_rolling_sheets(self.project_manager.rolling_workbook_path)
            if sheet_names:
                self.rolling_sheet_combo['values'] = sheet_names
            else:
                self.rolling_sheet_combo['values'] = []
        else:
            # Clear rolling file and sheet dropdown if no rolling workbook
            self.rolling_file_var.set("")
            self.rolling_sheet_combo['values'] = []
        
        # Load rolling sheet selection
        if current_project.rolling_sheet:
            self.rolling_sheet_var.set(current_project.rolling_sheet)
            # Force the combobox to update its display
            try:
                self.rolling_sheet_combo.set(current_project.rolling_sheet)
            except Exception as e:
                pass  # Debug output removed
            if current_project.rolling_sheet in self.rolling_sheet_combo['values']:
                current_value = self.rolling_sheet_combo.get()
            else:
                pass
        else:
            # Clear rolling sheet selection if no sheet is saved
            self.rolling_sheet_var.set("")
            self.rolling_sheet_combo.set("")
        
        # Load mapping file if available
        if current_project.mapping_file_path:
            self.mapping_file_var.set(current_project.mapping_file_path)
        
        # Clear any cached monthly data to ensure fresh source amounts for this project
        if hasattr(current_project, 'monthly_data'):
            current_project.monthly_data = {}
        
        # Keep target month to preserve historical data when switching projects
        # Target month will be properly detected during statement generation
        
        # Populate mappings if they exist
        if current_project.mappings:
            # Check for new accounts and add them before populating the tree
            # This ensures new accounts like 8540 are included
            if current_project.source_range:
                self.check_and_add_new_accounts(silent_mode=True)
            
            self.populate_mapping_tree(current_project.mappings)
            
            # Handle mapping signature initialization for cached data
            has_cached_monthly_data = (hasattr(current_project, 'monthly_data') and 
                                     current_project.monthly_data)
            
            if has_cached_monthly_data:
                # If project has cached monthly data and was not generated in this session, 
                # mark signature as potentially stale but don't clear it completely
                # This preserves the ability to detect real changes
                if current_project.name not in self.projects_with_session_generated_data:
                    # Store current signature but mark that cached data exists
                    self._update_mapping_signature(current_project.name, current_project.mappings)
                    # Add flag to indicate this project has cached data
                    if not hasattr(self, 'projects_with_cached_data'):
                        self.projects_with_cached_data = set()
                    self.projects_with_cached_data.add(current_project.name)
                else:
                    # Data was generated in current session, signature is valid
                    self._update_mapping_signature(current_project.name, current_project.mappings)
            else:
                # If no cached data, initialize signature normally
                self._update_mapping_signature(current_project.name, current_project.mappings)
        else:
            # Clear mapping tree if no mappings exist for this project
            self.populate_mapping_tree({})
        
        # Load Step 3 data if available
        self.load_step4_data(current_project)
        
        # Update target month display
        self.update_target_month_display()
        
        # Force UI update to ensure all changes are visible
        self.root.update_idletasks()
        
        # Attempt automatic workflow execution
        self.attempt_automatic_workflow()
        
        # Summary of what was loaded
        
        # Set comprehensive status message for user
        status_parts = []
        if current_project.rolling_sheet:
            status_parts.append(f"rolling sheet '{current_project.rolling_sheet}'")
        if current_project.source_range:
            status_parts.append(f"source range '{current_project.source_range}'")
        if current_project.rolling_range:
            status_parts.append(f"rolling range '{current_project.rolling_range}'")
        if current_project.mappings:
            status_parts.append(f"{len(current_project.mappings)} account mappings")
        
        if status_parts:
            status_message = f"✅ Project '{current_project.name}' loaded with stored data: {', '.join(status_parts)}"
            self.status_var.set(status_message)
        else:
            self.status_var.set(f"Project '{current_project.name}' selected. Please upload files and set ranges to begin.")
    
    def update_ui_state(self):
        """Update UI elements based on current state"""
        has_projects = self.project_manager.has_projects()
        current_project = self.project_manager.get_current_project()
        has_current_project = current_project is not None
        
        # Enable/disable rolling P&L controls
        state = "normal" if has_current_project else "disabled"
        self.rolling_browse_button.config(state=state)
        self.rolling_sheet_label.config(state=state)
        self.rolling_sheet_combo.config(state="readonly" if has_current_project else "disabled")
        
        # Enable/disable range controls
        self.source_range_label.config(state=state)
        self.source_range_entry.config(state=state)
        self.source_preview_button.config(state=state)
        self.rolling_range_label.config(state=state)
        self.rolling_range_entry.config(state=state)
        self.rolling_preview_button.config(state=state)
        
        # Enable/disable mapping file controls
        self.mapping_browse_button.config(state=state)
        
        
        # Enable/disable generate button based on ranges
        can_generate = (has_current_project and 
                       self.source_range_var.get().strip() and 
                       self.rolling_range_var.get().strip())
        self.generate_button.config(state="normal" if can_generate else "disabled")
        
        # Enable/disable Step 3 buttons based on mappings availability
        has_mappings = (has_current_project and 
                       hasattr(current_project, 'mappings') and 
                       current_project.mappings)
        
        if hasattr(self, 'save_mappings_button'):
            self.save_mappings_button.config(state="normal" if has_mappings else "disabled")
        if hasattr(self, 'reset_mappings_button'):
            self.reset_mappings_button.config(state="normal" if has_mappings else "disabled")
        
        # Enable/disable Step 3 buttons based on mappings and rolling sheet availability
        has_rolling_sheet = (has_current_project and 
                           hasattr(current_project, 'rolling_sheet') and 
                           current_project.rolling_sheet)
        can_generate_monthly = has_mappings and has_rolling_sheet
        
        if hasattr(self, 'generate_monthly_button'):
            self.generate_monthly_button.config(state="normal" if can_generate_monthly else "disabled")
        # Enable finalize button if monthly statement has been generated or Step 3 completed
        can_finalize = False
        if has_current_project:
            # Check if Step 3 data exists (monthly statement generated)
            has_monthly_data = (hasattr(current_project, 'aggregated_data') and 
                              current_project.aggregated_data)
            # Check if Step 3 was previously completed
            step4_completed = hasattr(current_project, 'step4_completed') and current_project.step4_completed
            can_finalize = has_monthly_data or step4_completed
            
        if hasattr(self, 'finalize_button'):
            # Update button text with current project name
            project_name = current_project.name if current_project else "Project"
            self.finalize_button.config(
                text=f"✅ Finalize & Export Final Excel for {project_name}",
                state="normal" if can_finalize else "disabled"
            )
        
        # Manage "All Projects" button - enabled if any project has generated data
        can_finalize_all = False
        if has_projects:

            pass
            for project in self.project_manager.projects.values():
                if hasattr(project, 'aggregated_data') and project.aggregated_data:
                    can_finalize_all = True
                    break
        
        if hasattr(self, 'finalize_all_button'):
            self.finalize_all_button.config(state="normal" if can_finalize_all else "disabled")
        
        # Update status
        if not has_projects:
            self.status_var.set("Please upload a Source P&L workbook to begin.")
        elif not has_current_project:
            self.status_var.set("Please select a project to continue.")
        elif has_mappings:
            self.status_var.set(f"Working on project: {current_project.name} ({len(current_project.mappings)} mappings loaded)")
        else:
            self.status_var.set(f"Working on project: {current_project.name}")
    
    def select_rolling_workbook(self):
        """Select rolling P&L workbook"""
        filename = filedialog.askopenfilename(
            title="Select Rolling P&L Workbook",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if filename:
            self.rolling_file_var.set(filename)
            self.project_manager.set_rolling_workbook(filename)
            
            # Load sheet names for selection
            sheet_names = self.project_manager.get_rolling_sheets(filename)
            if sheet_names:
                # Populate the combobox with sheet names
                self.rolling_sheet_combo['values'] = sheet_names
                
                # Preserve existing rolling sheet selection if it exists in the new workbook
                current_project = self.project_manager.get_current_project()
                if current_project and current_project.rolling_sheet and current_project.rolling_sheet in sheet_names:
                    # Keep the existing selection if it's valid
                    self.rolling_sheet_var.set(current_project.rolling_sheet)
                else:
                    # Clear selection only if the previous selection is not valid
                    self.rolling_sheet_var.set("")
                    if current_project and current_project.rolling_sheet:
                        # Clear the rolling sheet in the project if it's no longer valid
                        current_project.rolling_sheet = None
            
            self.project_manager.save_settings()
            
            # Check if both source and rolling files are now ready for auto-processing
            self.check_and_start_auto_process()
    
    def on_rolling_sheet_selected(self, event=None):
        """Handle rolling sheet selection from combobox"""
        selected_sheet = self.rolling_sheet_var.get()
        if selected_sheet:
            # Save to current project
            current_project = self.project_manager.get_current_project()
            if current_project:
                # Clear cache if sheet changed
                if hasattr(current_project, 'rolling_sheet') and current_project.rolling_sheet != selected_sheet:
                    # Clear all cache entries for this project
                    keys_to_remove = [k for k in self.rolling_accounts_cache.keys() if k.startswith(f"{current_project.name}:")]
                    for key in keys_to_remove:
                        del self.rolling_accounts_cache[key]
                
                current_project.rolling_sheet = selected_sheet
                
                # Load memorized ranges for this sheet if available
                if selected_sheet in current_project.sheet_ranges:
                    sheet_range_data = current_project.sheet_ranges[selected_sheet]
                    if 'source' in sheet_range_data:
                        self.source_range_var.set(sheet_range_data['source'])
                    if 'rolling' in sheet_range_data:
                        self.rolling_range_var.set(sheet_range_data['rolling'])
                else:
                    # No sheet-specific ranges, but still load the project-level source range
                    if current_project.source_range:
                        self.source_range_var.set(current_project.source_range)
                
                self.project_manager.save_settings()
                self.update_ui_state()
    
    # Add placeholder methods for remaining functionality
    def select_mapping_file(self):
        """Select and load mapping file for current project"""
        filename = filedialog.askopenfilename(
            title="Select Mapping File",
            filetypes=[("JSON files", "*.json"), ("Mapping files", "*.mapping"), ("All files", "*.*")]
        )
        
        if filename:
            current_project = self.project_manager.get_current_project()
            if current_project:

                pass
                try:
                    # Load mapping file
                    mappings = self.load_mapping_file(filename)
                    
                    if mappings:
                        # Preserve current range settings
                        current_source_range = self.source_range_var.get()
                        current_rolling_range = self.rolling_range_var.get()
                        
                        # Store mappings in current project
                        current_project.mappings = mappings
                        current_project.mapping_file_path = filename
                        self.mapping_file_var.set(filename)
                        
                        # Ensure range settings are saved to project (override empty ranges)
                        if current_source_range:
                            current_project.source_range = current_source_range
                        if current_rolling_range:
                            current_project.rolling_range = current_rolling_range
                        
                        # If ranges are still empty, use current UI values
                        if not current_project.source_range:
                            current_project.source_range = self.source_range_var.get()
                        if not current_project.rolling_range:
                            current_project.rolling_range = self.rolling_range_var.get()
                        
                        # Update UI
                        self.populate_mapping_tree(mappings)
                        self.update_ui_state()  # Enable Step 3 buttons when mappings loaded
                        
                        # Update status
                        total_mappings = len(mappings)
                        high_confidence = len([m for m in mappings.values() if m.get('confidence', '') == 'High'])
                        medium_confidence = len([m for m in mappings.values() if m.get('confidence', '') == 'Medium'])
                        low_confidence = len([m for m in mappings.values() if m.get('confidence', '') == 'Low'])
                        
                        status_msg = f"Loaded {total_mappings} mappings: {high_confidence} High, {medium_confidence} Medium, {low_confidence} Low confidence"
                        self.status_var.set(status_msg)
                        
                        # Save settings
                        self.project_manager.save_settings()
                        
                        messagebox.showinfo("Success", f"Successfully loaded {total_mappings} mappings from file!")
                    else:
                        messagebox.showerror("Error", "No valid mappings found in file.")
                        
                except Exception as e:
                    messagebox.showerror("Error", f"Error loading mapping file: {str(e)}")
    
    def load_mapping_file(self, filename):
        """Load mappings from a file"""
        try:

            pass
            with open(filename, 'r') as f:
                data = json.load(f)
            
            # Handle different file formats
            if isinstance(data, dict):

                pass
                if 'mappings' in data:
                    # Project file format
                    mappings = data['mappings']
                elif all(isinstance(v, dict) for v in data.values()):
                    # Direct mappings format
                    mappings = data
                else:
                    # Legacy format - convert simple key-value pairs
                    mappings = {}
                    for source, rolling in data.items():
                        mappings[source] = {
                            'rolling_account': rolling,
                            'confidence': 'Manual',
                            'similarity': 100.0,
                            'user_edited': True
                        }
            else:
                raise ValueError("Invalid file format")
            
            # Validate mapping structure
            validated_mappings = {}
            for source_account, mapping_info in mappings.items():
                if isinstance(mapping_info, dict):
                    validated_mappings[source_account] = {
                        'rolling_account': mapping_info.get('rolling_account', ''),
                        'confidence': mapping_info.get('confidence', 'Manual'),
                        'similarity': mapping_info.get('similarity', 100.0),
                        'user_edited': mapping_info.get('user_edited', True)
                    }
                else:
                    # Legacy format - simple string mapping
                    validated_mappings[source_account] = {
                        'rolling_account': str(mapping_info),
                        'confidence': 'Manual',
                        'similarity': 100.0,
                        'user_edited': True
                    }
            
            return validated_mappings
            
        except json.JSONDecodeError:
            raise ValueError("File is not a valid JSON format")
        except Exception as e:
            raise ValueError(f"Error reading file: {str(e)}")
    
    def load_mappings_from_saved_file(self, silent_mode=False):
        """Load and apply mappings from the saved mapping file path"""
        current_project = self.project_manager.get_current_project()
        if not current_project or not current_project.mapping_file_path:

            pass
            return False
        
        # Don't load from mapping file if we already have mappings
        # This prevents overwriting updated mappings with old ones
        if current_project.mappings:
            print(f"  ℹ️ Skipping mapping file load - already have {len(current_project.mappings)} mappings")
            return False
        
        try:
            # Check if the mapping file exists
            if not os.path.exists(current_project.mapping_file_path):

                pass
                if not silent_mode:
                    pass  # Debug output removed
                return False
            
            # Load mappings from file
            mappings = self.load_mapping_file(current_project.mapping_file_path)
            
            if mappings:
                print(f"  📂 Loading {len(mappings)} mappings from {os.path.basename(current_project.mapping_file_path)}")
                # Store mappings in current project
                current_project.mappings = mappings
                
                # Update UI only if not in silent mode
                if not silent_mode:
                    self.populate_mapping_tree(mappings)
                
                # Update status
                filename = os.path.basename(current_project.mapping_file_path)
                if not silent_mode:
                    self.status_var.set(f"Automatically loaded mappings from: {filename}")
                else:
                    pass
                
                # Save updated project state
                self.project_manager.save_settings()
                
                return True
            
        except Exception as e:

            
            pass
            if not silent_mode:
                pass  # Debug output removed
            return False
        
        return False
    
    def attempt_automatic_workflow(self, silent_mode=False):
        """Attempt to automatically execute workflow steps based on available data"""
        current_project = self.project_manager.get_current_project()
        if not current_project:

            pass
            return
        
        try:
            import os
            
            # Track what was done automatically
            auto_actions = []
            
            has_source_file = bool(current_project.source_file_path and os.path.exists(current_project.source_file_path))
            has_rolling_file = bool(self.project_manager.rolling_workbook_path and os.path.exists(self.project_manager.rolling_workbook_path))
            has_rolling_sheet = bool(current_project.rolling_sheet)
            has_ranges = bool(current_project.source_range and current_project.rolling_range)
            
            if not (has_source_file and has_rolling_file and has_rolling_sheet and has_ranges):

            
                pass
                if not silent_mode:
                    pass
                return
            
            # Step 2: Prioritize loading saved mapping files
            has_mapping_file = bool(current_project.mapping_file_path)
            if not current_project.mapping_file_path:
                # Try to find mapping files by convention
                potential_mapping_files = [
                    f"{current_project.name}_mappings.json",
                    f"{current_project.name.replace(' ', '_')}_mappings.json",
                    f"{current_project.name}_mapping.json",
                    f"{current_project.name.replace(' ', '_')}_mapping.json"
                ]
                
                for potential_file in potential_mapping_files:
                    if os.path.exists(potential_file):
                        current_project.mapping_file_path = potential_file
                        self.mapping_file_var.set(potential_file)
                        has_mapping_file = True
                        break
                
                if not has_mapping_file:
                    pass  # Debug output removed
            mappings_loaded_from_file = False
            if current_project.mapping_file_path and os.path.exists(current_project.mapping_file_path):
                # This ensures the most up-to-date saved mappings are used
                if self.load_mappings_from_saved_file(silent_mode=silent_mode):
                    auto_actions.append("Loaded mappings from saved file")
                    mappings_loaded_from_file = True
                else:
                    pass  # Debug output removed
            elif current_project.mappings:
                pass
            else:
                pass  # Debug output removed
            if not current_project.mappings and has_ranges:
                self.generate_mappings(silent_mode=silent_mode)
                if current_project.mappings:
                    auto_actions.append("Generated account mappings")
                else:

                    pass
                    return
            elif current_project.mappings:

                pass
            
            # Step 3: Generate monthly statement if mappings exist
            # Check if mappings have actually changed to avoid unnecessary regeneration
            mappings_changed = self._have_mappings_changed(current_project.name, current_project.mappings)
            
            # Check if Step 3 data already exists and is valid
            has_step3_data = (hasattr(current_project, 'aggregated_data') and 
                            current_project.aggregated_data and
                            hasattr(current_project, 'preview_data') and 
                            current_project.preview_data)
            
            # Check if project has cached data from previous session
            has_cached_data = (hasattr(self, 'projects_with_cached_data') and 
                             current_project.name in self.projects_with_cached_data)
            
            # Check if mappings were manually modified
            has_manual_modifications = (hasattr(self, 'projects_with_modified_mappings') and 
                                      current_project.name in self.projects_with_modified_mappings)
            
            
            # Force regeneration only if:
            # 1. Mappings were loaded from file AND have changed since last generation
            # 2. OR no aggregated data exists (first time or data cleared)
            # 3. OR mappings were manually edited (tracked separately)
            should_generate = (current_project.mappings and 
                             ((mappings_loaded_from_file and mappings_changed) or 
                              not has_step3_data or
                              has_manual_modifications))
            
            if should_generate:
                # Update UI state first to enable buttons
                self.update_ui_state()
                
                # Check if the generate button would be enabled
                can_generate = (has_source_file and has_rolling_file and 
                              has_rolling_sheet and current_project.mappings)
                
                if can_generate:

                
                    pass
                    if mappings_loaded_from_file and mappings_changed:
                        self.status_var.set("Regenerating monthly statement (mappings changed)...")
                        self.generate_monthly_statement(silent_mode=silent_mode)
                        auto_actions.append("Generated monthly statement")
                    elif mappings_loaded_from_file:
                        self.status_var.set("Using cached monthly statement (mappings unchanged)")
                        return  # Skip regeneration if mappings haven't changed
                    else:
                        self.status_var.set("Generating monthly statement...")
                        self.generate_monthly_statement(silent_mode=silent_mode)
                        auto_actions.append("Generated monthly statement")
            else:

                pass
                if has_step3_data:

                    pass
                else:

                    pass
            
            # Step 4: Skip auto-export during bulk processing - it will be handled at the end
            # Only do auto-export for individual project processing (not during bulk processing)
            if silent_mode and not hasattr(self, '_bulk_processing_active'):
                # Check if all projects are ready for final export
                all_projects_ready = True
                projects_with_data = []
                for proj_name, proj in self.project_manager.projects.items():
                    if hasattr(proj, 'aggregated_data') and proj.aggregated_data:
                        projects_with_data.append(proj_name)
                    else:
                        all_projects_ready = False
                
                # If we have projects with data and auto-processing enabled, trigger final export
                if projects_with_data:

                    pass
                    
                    # Generate automatic filename with timestamp
                    from datetime import datetime
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    target_month = current_project.target_month if current_project.target_month else "Unknown"
                    auto_filename = f"Final_Export_{target_month}_{timestamp}.xlsx"
                    
                    # Use default directory (same as rolling workbook) for auto-export
                    import os
                    if self.project_manager.rolling_workbook_path:
                        export_dir = os.path.dirname(self.project_manager.rolling_workbook_path)
                        auto_filepath = os.path.join(export_dir, auto_filename)
                    else:
                        # Fallback to user's desktop
                        auto_filepath = os.path.join(os.path.expanduser("~/Desktop"), auto_filename)
                    
                    if self.auto_export_all_projects(auto_filepath, silent_mode=True):
                        auto_actions.append(f"Exported final Excel ({len(projects_with_data)} projects)")
                    else:
                        pass  # Debug output removed
            elif not hasattr(self, '_bulk_processing_active'):
                # Check for projects ready for manual export (only if not in bulk processing)
                projects_with_data = []
                for proj_name, proj in self.project_manager.projects.items():
                    if hasattr(proj, 'aggregated_data') and proj.aggregated_data:
                        projects_with_data.append(proj_name)
                if projects_with_data and not silent_mode:

                    pass
            
            # Report automatic actions
            if auto_actions:
                status_msg = "Automatic actions: " + ", ".join(auto_actions)
                self.status_var.set(status_msg)
                
        except Exception as e:
            import traceback
            traceback.print_exc()
    
    def preview_range(self, file_type):
        """Preview the selected range from the specified file"""
        current_project = self.project_manager.get_current_project()
        if not current_project:
            messagebox.showwarning("Warning", "Please select a project first.")
            return
        
        if file_type == "source":
            file_path = self.project_manager.source_workbook_path
            sheet_name = current_project.source_sheet
            range_str = self.source_range_var.get()
        else:
            file_path = self.project_manager.rolling_workbook_path
            sheet_name = current_project.rolling_sheet
            range_str = self.rolling_range_var.get()
        
        if not file_path:
            messagebox.showwarning("Warning", f"Please select the {file_type} file first.")
            return
            
        if not sheet_name:
            messagebox.showwarning("Warning", f"Please select a {file_type} sheet first.")
            return
        
        if not range_str.strip():
            messagebox.showwarning("Warning", f"Please specify a range for {file_type}.")
            return
        
        try:
            # Read Excel file with specific sheet (using cache)
            import pandas as pd
            df = self._load_excel_with_cache(file_path, sheet_name)
            
            # Parse range and extract data (always exclude amounts for preview)
            data = self.extract_range_data(df, range_str, include_amounts=False)
            
            if not data:
                messagebox.showwarning("Warning", "No data found in the specified range.")
                return
            
            # Get target month for the title
            target_month = self.ensure_consistent_target_month(current_project)
            if target_month:
                clean_target_month = self.clean_target_month_text(target_month)
                target_month_text = f" - {clean_target_month}"
            else:
                target_month_text = ""
            
            # Show preview dialog
            self.show_range_preview(data, f"{file_type.title()} Range Preview - {current_project.name}{target_month_text}", range_str)
            
        except Exception as e:
            messagebox.showerror("Error", f"Error previewing range: {str(e)}")
    
    def extract_range_data(self, df, range_str, include_amounts=False):
        """Extract data from DataFrame based on range string - Only account descriptions for preview
        
        Args:
            df: DataFrame to extract from
            range_str: Range specification (e.g., 'A1:A100')
            include_amounts: Ignored for preview - always returns only account descriptions
        
        Returns:
            List of strings (account descriptions only)
        """
        import pandas as pd
        range_str = range_str.strip().upper()
        
        # Get current project for context
        current_project = self.project_manager.get_current_project()
        
        def is_numeric_value(value):
            """Check if a value is purely numeric (amount data to exclude)"""
            if pd.isna(value):

                pass
                return False
            str_val = str(value).strip()
            # Remove common currency symbols and separators
            clean_val = str_val.replace('$', '').replace(',', '').replace('(', '').replace(')', '')
            try:
                float(clean_val)
                return True
            except ValueError:

                pass
                return False
        
        try:

        
            pass
            if ':' in range_str:
                # Handle specific ranges like A1:A100, B2:D50
                parts = range_str.split(':')
                if len(parts) != 2:

                    pass
                    return []
                    
                start_cell, end_cell = parts
                
                # Parse start cell
                start_col, start_row = self.parse_cell_reference(start_cell)
                
                # Parse end cell
                end_col, end_row = self.parse_cell_reference(end_cell)
                
                if start_col is None or end_col is None:

                
                    pass
                    return []
                
                # If only column specified (like A:A), use all rows
                if start_row is None:
                    start_row = 0
                if end_row is None:
                    end_row = len(df) - 1
                
                # Validate ranges
                if start_col >= len(df.columns) or end_col >= len(df.columns):

                    pass
                    return []
                
                if start_row >= len(df) or end_row >= len(df):
                    end_row = len(df) - 1
                
                # Extract the range - preserve row-based chronological order
                data = []
                # Iterate by rows first to maintain chronological order
                for row_idx in range(start_row, end_row + 1):
                    if row_idx < len(df):
                        # Check all columns in this row
                        for col_idx in range(start_col, end_col + 1):
                            if col_idx < len(df.columns):

                                pass
                                try:
                                    cell_value = df.iloc[row_idx, col_idx]
                                    if pd.notna(cell_value):
                                        cell_str = str(cell_value).strip()
                                        if cell_str and cell_str != 'nan' and not is_numeric_value(cell_value):
                                            # Include all non-numeric text (account descriptions and subtotals)
                                            # Subtotals will be displayed as headings in the mapping tree
                                            data.append(cell_str)
                                except Exception as e:
                                    pass  # Debug output removed
                unique_data = list(dict.fromkeys(data))
                
                # Return fresh data without caching
                return unique_data
            
            else:
                # Handle single column like A, B, C - preserve row order
                try:
                    col_idx = self.column_letter_to_number(range_str) - 1
                    if col_idx < len(df.columns):
                        data = []
                        for row_idx in range(len(df)):
                            cell_value = df.iloc[row_idx, col_idx]
                            if pd.notna(cell_value):
                                cell_str = str(cell_value).strip()
                                if cell_str and cell_str != 'nan' and not is_numeric_value(cell_value):
                                    # Include all non-numeric text (account descriptions and subtotals)
                                    # Subtotals will be displayed as headings in the mapping tree
                                    data.append(cell_str)
                        
                        # Remove duplicates while preserving order
                        unique_data = list(dict.fromkeys(data))
                        
                        # Return fresh data without caching
                        return unique_data
                    else:

                        pass
                        return []
                except Exception as e:

                    pass
                    return []
                
        except Exception as e:
            import traceback
            traceback.print_exc()
            return []
        
        return []
    
    def parse_cell_reference(self, cell_ref):
        """Parse cell reference like A1, B2, etc. Returns (col_index, row_index)"""
        import re
        
        match = re.match(r'^([A-Z]+)(\d*)$', cell_ref.strip().upper())
        if not match:

            pass
            return None, None
        
        col_letters = match.group(1)
        row_num = match.group(2)
        
        col_idx = self.column_letter_to_number(col_letters) - 1
        row_idx = int(row_num) - 1 if row_num else None
        
        return col_idx, row_idx
    
    def column_letter_to_number(self, letters):
        """Convert column letters to number (A=1, B=2, etc.)"""
        result = 0
        for char in letters:
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result
    
    def show_range_preview(self, data, title, range_str):
        """Show preview dialog with extracted data"""
        dialog = tk.Toplevel(self.root)
        dialog.title(title)
        
        # Check if data contains tuples (account, amount) or just strings
        has_amounts = data and isinstance(data[0], tuple)
        
        # Set dialog size
        dialog_width = 1000
        dialog_height = 700
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center the dialog relative to parent window
        dialog.update_idletasks()
        self.root.update_idletasks()  # Ensure parent window dimensions are current
        
        # Get parent window position and size
        parent_x = self.root.winfo_x()
        parent_y = self.root.winfo_y()
        parent_width = self.root.winfo_width()
        parent_height = self.root.winfo_height()
        
        # Calculate center position relative to parent
        x = parent_x + (parent_width - dialog_width) // 2
        y = parent_y + (parent_height - dialog_height) // 2
        
        # Ensure dialog stays on screen (with some margin)
        x = max(50, x)  # Keep at least 50 pixels from left edge
        y = max(50, y)  # Keep at least 50 pixels from top edge
        
        # Set the final geometry
        dialog.geometry(f"{dialog_width}x{dialog_height}+{x}+{y}")
        
        # Header
        header_frame = ttk.Frame(dialog, padding="10")
        header_frame.pack(fill=tk.X)
        
        # Extract target month from title if present
        target_month = None
        if " - " in title:
            parts = title.split(" - ")
            if len(parts) >= 3:  # Format: "Source Range Preview - ProjectName - TargetMonth"
                target_month = parts[-1]
        
        # Display target month prominently if available
        if target_month:
            ttk.Label(header_frame, text=f"Target Month: {target_month}", font=("Arial", 14, "bold"), foreground="darkblue").pack(pady=(0, 5))
        
        ttk.Label(header_frame, text=f"Range: {range_str}", font=("Arial", 12, "bold")).pack()
        if has_amounts:
            ttk.Label(header_frame, text=f"Found {len(data)} accounts with amounts:", font=("Arial", 10)).pack()
        else:
            ttk.Label(header_frame, text=f"Found {len(data)} account descriptions:", font=("Arial", 10)).pack()
        
        # Data display frame
        data_frame = ttk.Frame(dialog, padding="10")
        data_frame.pack(fill=tk.BOTH, expand=True)
        
        if has_amounts:
            # Create treeview for account and amount data
            columns = ("Account Description", "Amount")
            tree = ttk.Treeview(data_frame, columns=columns, show="headings", height=20)
            
            tree.heading("Account Description", text="Account Description")
            tree.heading("Amount", text="Amount")
            tree.column("Account Description", width=600)
            tree.column("Amount", width=150)
            
            # Populate with data
            for account, amount in data:
                tree.insert("", "end", values=(account, amount))
        else:
            # Simple listbox for account descriptions only
            listbox = tk.Listbox(data_frame, height=25, font=("Arial", 10))
            
            # Populate with data
            for item in data:
                listbox.insert(tk.END, item)
            
            tree = listbox
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(data_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        # Pack components
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Button frame
        button_frame = ttk.Frame(dialog, padding="10")
        button_frame.pack(fill=tk.X)
        
        ttk.Button(button_frame, text="Close", command=dialog.destroy).pack(side=tk.RIGHT)
    
    
    
    
    @profile_performance("generate_monthly_statement")
    def generate_monthly_statement(self, silent_mode=False):
        """Generate monthly P&L statement by extracting data from column Q (Jun 2025 Actual) and aggregating by mappings"""
        try:
            current_project = self.project_manager.get_current_project()
            if not current_project:

                pass
                if not silent_mode:
                    messagebox.showerror("Error", "No project selected.")
                return
                
            self.step4_status_var.set("Extracting monthly data from source P&L...")
            self.root.update()
            
            # Validate required data
            if not current_project.source_file_path or not current_project.mappings:

                pass
                if not silent_mode:
                    messagebox.showerror("Error", "Please complete Steps 1-2 first.")
                return
            
            # Load source Excel file with specified sheet (using cache)
            source_df = self._load_excel_with_cache(current_project.source_file_path, current_project.source_sheet)
            
            # Extract monthly amounts from the target month column
            self.step4_status_var.set("Processing data from source P&L...")
            self.root.update()
            
            monthly_data = self.extract_monthly_amounts(source_df, current_project)
            
            if not monthly_data:

            
                pass
                if not silent_mode:
                    messagebox.showwarning("Warning", "No monthly amounts found in the specified source range. Please check the source file and range settings.")
                self.step4_status_var.set("No data found in source range")
                return
            
            # Aggregate amounts by mapping categories
            self.step4_status_var.set("Aggregating amounts by mapping categories...")
            self.root.update()
            
            aggregated_data = self.aggregate_by_mappings(monthly_data, current_project)
            
            # Get column header from source P&L cell I6
            target_month = self.get_source_month_header(source_df)
            
            # Load rolling P&L and extract existing data for preview
            self.step4_status_var.set("Loading rolling P&L for preview...")
            self.root.update()
            
            rolling_df = self._load_excel_with_cache(self.project_manager.rolling_workbook_path, current_project.rolling_sheet)
            
            # Apply rolling range to respect the specified range
            if current_project.rolling_range:
                rolling_rows = self.parse_range_for_rows(current_project.rolling_range)
                # Convert to 0-based index for pandas
                start_row = rolling_rows[0] - 2  # Subtract 2 to convert from Excel to pandas index
                end_row = rolling_rows[1] - 1
                if start_row >= 0 and end_row < len(rolling_df):
                    rolling_df = rolling_df.iloc[start_row:end_row+1].reset_index(drop=True)
            
            preview_data = self.prepare_preview_data(rolling_df, target_month, aggregated_data)
            
            # Store data for export
            current_project.monthly_data = monthly_data
            current_project.aggregated_data = aggregated_data
            current_project.preview_data = preview_data
            current_project.target_month = target_month
            
            # Mark Step 3 as having generated data
            if hasattr(current_project, 'workflow_state'):
                current_project.workflow_state['has_generated_monthly'] = True
            
            # Save Step 3 data to persistence
            self.project_manager.save_settings()
            
            # Populate preview table
            self.populate_preview_table(preview_data)
            
            # Update Step 3 popup tree if it exists
            if hasattr(self, 'step3_popup_tree') and self.step3_popup_tree and self.step3_is_popped_out:
                self.update_step3_popup_headings()
                self.sync_preview_tree_data(self.preview_tree, self.step3_popup_tree)
            
            # Enable action buttons
            self.finalize_button.config(state="normal")
            
            # Update popup finalize button if it exists
            self.update_popup_finalize_button_state()
            
            # Reset Step 3 modification flag since statement is regenerated
            self.reset_step4_modification_flag()
            
            # Update mapping signature to mark this as the current state
            current_project = self.project_manager.get_current_project()
            if current_project:
                self._update_mapping_signature(current_project.name, current_project.mappings)
                # Mark this project as having session-generated data
                self.projects_with_session_generated_data.add(current_project.name)
                # Clear modification tracking since statements were regenerated
                if hasattr(self, 'projects_with_modified_mappings'):
                    self.projects_with_modified_mappings.discard(current_project.name)
            
            self.step4_status_var.set(f"Generated preview for {len(aggregated_data)} categories. Target month: {target_month}")
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            if not silent_mode:
                messagebox.showerror("Error", f"Error generating monthly statement: {str(e)}")
            self.step4_status_var.set("Error generating statement")
    
    def finalize_and_export(self):
        """Finalize the monthly statement and export the final Excel file"""
        try:
            current_project = self.project_manager.get_current_project()
            if not current_project or not hasattr(current_project, 'aggregated_data'):
                messagebox.showerror("Error", "Please generate monthly statement first.")
                return
            
                
            # Ask user for save location
            filename = filedialog.asksaveasfilename(
                title="Save Final Rolling P&L",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            
            if not filename:

            
                pass
                return
            
            self.step4_status_var.set("Writing data to rolling P&L...")
            self.root.update()
            
            # Load original rolling P&L without data_only to preserve formulas
            from openpyxl import load_workbook
            wb = load_workbook(self.project_manager.rolling_workbook_path, data_only=False)
            
            # Get the correct worksheet
            if current_project.rolling_sheet in wb.sheetnames:
                ws = wb[current_project.rolling_sheet]
            else:
                ws = wb.active
            
            # Find target column based on source month header
            target_col = self.find_matching_column(ws, current_project.target_month)
            
            if not target_col:
                # Show column selection dialog if no automatic match found
                target_col = self.show_column_selection_dialog(ws, current_project.target_month)
                if not target_col:
                    # User cancelled column selection
                    self.step4_status_var.set("Export cancelled")
                    return
            
            # Write aggregated data to rolling P&L
            self.write_data_to_rolling_pnl(ws, target_col, current_project)
            
            # Save the workbook
            wb.save(filename)
            
            # Mark Step 3 as completed and save export file path
            current_project.step4_completed = True
            current_project.last_export_file = filename
            if hasattr(current_project, 'workflow_state'):
                current_project.workflow_state['step4_complete'] = True
            
            # Save completion state to persistence
            self.project_manager.save_settings()
            
            self.step4_status_var.set(f"Final rolling P&L exported to {os.path.basename(filename)}")
            messagebox.showinfo("Success", f"Final rolling P&L saved successfully to:\n{filename}")
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("Error", f"Error exporting final file: {str(e)}")
    
    def finalize_and_export_all_projects(self):
        """Finalize and export all projects that have generated monthly statements"""
        try:
            # Get all projects that have generated data
            projects_with_data = []
            total_projects = len(self.project_manager.projects)
            
            for project in self.project_manager.projects.values():
                has_aggregated = hasattr(project, 'aggregated_data') and project.aggregated_data
                
                if has_aggregated:
                    projects_with_data.append(project)
            
            
            if not projects_with_data:
                messagebox.showwarning("Warning", "No projects have generated monthly statements to export.")
                return
            
            # Ask user for save location
            filename = filedialog.asksaveasfilename(
                title="Save Combined Rolling P&L for All Projects",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            
            if not filename:

            
                pass
                return
            
            self.step4_status_var.set(f"Exporting data for {len(projects_with_data)} projects...")
            self.root.update()
            
            # Load original rolling P&L without data_only to preserve formulas
            from openpyxl import load_workbook
            wb = load_workbook(self.project_manager.rolling_workbook_path, data_only=False)
            
            projects_exported = 0
            failed_projects = []
            
            # Process each project
            for project in projects_with_data:
                try:
                    # Get the correct worksheet for this project
                    if project.rolling_sheet in wb.sheetnames:
                        ws = wb[project.rolling_sheet]
                    else:
                        failed_projects.append(f"{project.name}: Rolling sheet '{project.rolling_sheet}' not found")
                        continue  # Skip if sheet doesn't exist
                    
                    # Find target column based on source month header
                    target_col = self.find_matching_column(ws, project.target_month)
                    
                    if target_col:
                        # Write aggregated data to rolling P&L
                        self.write_data_to_rolling_pnl(ws, target_col, project)
                        projects_exported += 1
                        
                        self.step4_status_var.set(f"Exported {projects_exported}/{len(projects_with_data)} projects...")
                        self.root.update()
                    else:
                        failed_projects.append(f"{project.name}: Target month column '{project.target_month}' not found")
                    
                except Exception as e:
                    failed_projects.append(f"{project.name}: {str(e)}")
                    continue
            
            # Save the workbook
            wb.save(filename)
            
            # Mark all exported projects as completed
            for project in projects_with_data:
                if hasattr(project, 'workflow_state'):
                    project.workflow_state['step4_complete'] = True
                project.step4_completed = True
                project.last_export_file = filename
            
            # Save all project states
            self.project_manager.save_settings()
            
            # Show results
            if projects_exported == len(projects_with_data) and not failed_projects:
                self.step4_status_var.set(f"All projects exported successfully ({projects_exported} projects)")
                messagebox.showinfo("Success", 
                                  f"Combined rolling P&L exported successfully!\n\n"
                                  f"Projects exported: {projects_exported}/{len(projects_with_data)}\n"
                                  f"File saved to: {os.path.basename(filename)}")
            else:
                self.step4_status_var.set(f"Export completed with issues ({projects_exported}/{len(projects_with_data)} projects)")
                
                # Create detailed message
                msg = f"Export completed!\n\n"
                msg += f"Successfully exported: {projects_exported}/{len(projects_with_data)} projects\n"
                msg += f"File saved to: {os.path.basename(filename)}\n"
                
                if failed_projects:
                    msg += f"\nFailed projects ({len(failed_projects)}):\n"
                    for failure in failed_projects:
                        msg += f"• {failure}\n"
                
                messagebox.showwarning("Export Completed with Issues", msg)
            
            # Update UI state to refresh buttons
            self.update_ui_state()
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("Error", f"Error exporting all projects: {str(e)}")
            # Update UI state even on error
            self.update_ui_state()
    
    def auto_export_all_projects(self, filename, silent_mode=False):
        """Automated version of finalize_and_export_all_projects without user dialogs"""
        try:
            # Get all projects that have generated data
            projects_with_data = []
            for project in self.project_manager.projects.values():
                if hasattr(project, 'aggregated_data') and project.aggregated_data:
                    projects_with_data.append(project)
            
            if not projects_with_data:

            
                pass
                return False
            
            if not silent_mode:
                self.step4_status_var.set(f"Exporting data for {len(projects_with_data)} projects...")
                self.root.update()
            
            # Load original rolling P&L without data_only to preserve formulas
            from openpyxl import load_workbook
            wb = load_workbook(self.project_manager.rolling_workbook_path, data_only=False)
            
            projects_exported = 0
            
            # Process each project
            for project in projects_with_data:
                try:
                    # Get the correct worksheet for this project
                    if project.rolling_sheet in wb.sheetnames:
                        ws = wb[project.rolling_sheet]
                    else:

                        pass
                        continue  # Skip if sheet doesn't exist
                    
                    # Find target column based on source month header
                    target_col = self.find_matching_column(ws, project.target_month)
                    
                    if target_col:
                        # Write aggregated data to rolling P&L
                        self.write_data_to_rolling_pnl(ws, target_col, project)
                        projects_exported += 1
                        
                        if not silent_mode:
                            self.step4_status_var.set(f"Exported {projects_exported}/{len(projects_with_data)} projects...")
                            self.root.update()
                    else:
                        pass  # Debug output removed
                    
                except Exception as e:

                    
                    pass
                    continue
            
            # Save the workbook
            wb.save(filename)
            for project in projects_with_data:
                if hasattr(project, 'workflow_state'):
                    project.workflow_state['step4_complete'] = True
                project.step4_completed = True
                project.last_export_file = filename
            
            # Save all project states
            self.project_manager.save_settings()
            
            final_message = f"All projects exported successfully ({projects_exported}/{len(projects_with_data)} projects)"
            
            if not silent_mode:
                self.step4_status_var.set(final_message)
                import os
                messagebox.showinfo("Auto-Export Complete", 
                                  f"Combined rolling P&L exported successfully!\n\n"
                                  f"Projects exported: {projects_exported}/{len(projects_with_data)}\n"
                                  f"File saved to: {os.path.basename(filename)}")
            
            return True
            
        except Exception as e:
            error_msg = f"Error in auto-export: {e}"
            import traceback
            traceback.print_exc()
            
            if not silent_mode:
                messagebox.showerror("Auto-Export Error", error_msg)
            
            return False
    
    def bulk_process_all_projects(self):
        """Automatically process all projects with saved mappings and generate monthly statements"""
        self._bulk_processing_active = True
        
        project_names = self.project_manager.get_project_names()
        if not project_names:

            pass
            if hasattr(self, '_bulk_processing_active'):
                delattr(self, '_bulk_processing_active')
            return
        
        total_projects = len(project_names)
        processed_count = 0
        failed_count = 0
        
        # Show progress dialog
        progress_dialog = self.show_progress_dialog(total_projects, project_names)
        
        # Store the original current project to restore later
        original_project = self.project_manager.get_current_project()
        original_project_name = original_project.name if original_project else None
        
        try:

        
            pass
            for index, project_name in enumerate(project_names):
                self.update_progress(index, total_projects, project_name, "Loading project data...")
                
                try:
                    # Select the project
                    if self.project_manager.select_project(project_name):
                        current_project = self.project_manager.get_current_project()
                        
                        # Update progress
                        self.update_progress(index, total_projects, project_name, "Loading saved mappings...")
                        
                        # Load project data (this triggers the automatic workflow)
                        self.load_project_data()
                        
                        # Update progress
                        self.update_progress(index, total_projects, project_name, "Processing mappings and statements...")
                        
                        # Explicitly attempt automatic workflow with silent mode for bulk processing
                        self.attempt_automatic_workflow(silent_mode=True)
                        
                        # Check if processing was successful
                        if current_project.mappings and current_project.monthly_data:
                            processed_count += 1
                            self.update_progress(index + 1, total_projects, project_name, "✅ Complete")
                        else:
                            # Update progress with warning
                            self.update_progress(index + 1, total_projects, project_name, "⚠️ Partial")
                            
                    else:
                        failed_count += 1
                        # Update progress with error
                        self.update_progress(index + 1, total_projects, project_name, "❌ Failed")
                        
                except Exception as e:
                    failed_count += 1
                    # Update progress with error
                    self.update_progress(index + 1, total_projects, project_name, f"❌ Error: {str(e)[:20]}...")
                    # Continue with next project
                    
            # Restore original project selection
            if original_project_name and original_project_name in project_names:
                self.project_manager.select_project(original_project_name)
                self.load_project_data()
            
            # Update status and provide summary
            total_projects = len(project_names)
            success_message = f"Bulk processing completed!\n\n"
            success_message += f"📊 Summary:\n"
            success_message += f"  • Total projects: {total_projects}\n"
            success_message += f"  • Successfully processed: {processed_count}\n"
            success_message += f"  • Failed/Partial: {failed_count}\n\n"
            
            if processed_count > 0:
                success_message += f"✅ {processed_count} projects are ready with mappings and monthly statements."
            else:
                success_message += "⚠️  No projects were fully processed. Check file paths and prerequisites."
            
            if hasattr(self, '_bulk_processing_active'):
                delattr(self, '_bulk_processing_active')
            
            # Now check if we should auto-export (after all projects are processed)
            if processed_count > 0:
                # Check if projects are ready for final export
                projects_with_data = []
                for proj_name, proj in self.project_manager.projects.items():
                    if hasattr(proj, 'aggregated_data') and proj.aggregated_data:
                        projects_with_data.append(proj_name)
                
                if projects_with_data:

                
                    pass
                    
                    # Generate automatic filename with timestamp
                    from datetime import datetime
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    current_project = self.project_manager.get_current_project()
                    target_month = current_project.target_month if current_project and current_project.target_month else "Unknown"
                    auto_filename = f"Final_Export_{target_month}_{timestamp}.xlsx"
                    
                    # Use default directory (same as rolling workbook) for auto-export
                    import os
                    if self.project_manager.rolling_workbook_path:
                        export_dir = os.path.dirname(self.project_manager.rolling_workbook_path)
                        auto_filepath = os.path.join(export_dir, auto_filename)
                    else:
                        # Fallback to user's desktop
                        auto_filepath = os.path.join(os.path.expanduser("~/Desktop"), auto_filename)
                    
                    if self.auto_export_all_projects(auto_filepath, silent_mode=True):
                        success_message += f"\n\n📁 Final Excel exported: {os.path.basename(auto_filepath)}"
                    else:
                        success_message += f"\n\n⚠️ Final Excel export failed"
            
            # Final progress update
            self.update_progress(total_projects, total_projects, "All Projects", 
                               f"🏁 Complete: {processed_count} successful, {failed_count} failed")
            
            # Wait a moment to show final status
            self.root.after(2000, self.close_progress_dialog)
            
            # Update UI to reflect the processing results
            self.update_ui_state()
            self.status_var.set(f"Bulk processing: {processed_count}/{total_projects} projects processed successfully")
            
        except Exception as e:

            
            pass
            if hasattr(self, '_bulk_processing_active'):
                delattr(self, '_bulk_processing_active')
            # Close progress dialog on error
            self.close_progress_dialog()
            # Restore original project if possible
            if original_project_name:

                pass
                try:
                    self.project_manager.select_project(original_project_name)
                    self.load_project_data()
                except:
                    pass
    
    def check_and_start_auto_process(self):
        """Check if both source and rolling files are available and start auto-processing"""
        # Check if both files are available
        source_path = self.source_file_var.get()
        rolling_path = self.project_manager.rolling_workbook_path
        
        has_source = bool(source_path and os.path.exists(source_path))
        has_rolling = bool(rolling_path and os.path.exists(rolling_path))
        
        
        if has_source and has_rolling:
            self.status_var.set("Both files ready - starting automatic processing...")
            self.root.update()
            
            # Start bulk processing with progress tracking
            self.bulk_process_all_projects()
        else:

            pass
            if not has_source:
                self.status_var.set("Please upload Source P&L file to continue automatic processing")
            elif not has_rolling:
                self.status_var.set("Please upload Rolling P&L file to continue automatic processing")
    
    def show_progress_dialog(self, total_projects, project_names):
        """Create and show progress dialog for bulk processing"""
        import tkinter as tk
        from tkinter import ttk
        
        # Create progress dialog
        self.progress_dialog = tk.Toplevel(self.root)
        self.progress_dialog.title("Processing Projects")
        self.progress_dialog.transient(self.root)
        self.progress_dialog.grab_set()
        
        # Center the dialog
        window_width = 500
        window_height = 200
        x = (self.progress_dialog.winfo_screenwidth() // 2) - (window_width // 2)
        y = (self.progress_dialog.winfo_screenheight() // 2) - (window_height // 2)
        self.progress_dialog.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # Prevent closing with X button
        self.progress_dialog.protocol("WM_DELETE_WINDOW", lambda: None)
        
        # Main frame
        main_frame = ttk.Frame(self.progress_dialog, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Title
        title_label = ttk.Label(main_frame, text="🚀 Processing All Projects", 
                               font=("Arial", 14, "bold"))
        title_label.pack(pady=(0, 10))
        
        # Current project label
        self.progress_current_label = ttk.Label(main_frame, text="Initializing...", 
                                               font=("Arial", 11))
        self.progress_current_label.pack(pady=(0, 5))
        
        # Progress bar
        self.progress_bar = ttk.Progressbar(main_frame, length=400, mode='determinate')
        self.progress_bar.pack(pady=(0, 10))
        
        # Progress text (X of Y projects)
        self.progress_text_label = ttk.Label(main_frame, text=f"0 of {total_projects} projects", 
                                           font=("Arial", 10))
        self.progress_text_label.pack(pady=(0, 5))
        
        # Status label
        self.progress_status_label = ttk.Label(main_frame, text="Starting bulk processing...", 
                                             font=("Arial", 9), foreground="gray")
        self.progress_status_label.pack()
        
        # Set up progress bar
        self.progress_bar['maximum'] = total_projects
        self.progress_bar['value'] = 0
        
        # Update dialog
        self.progress_dialog.update()
        
        return self.progress_dialog
    
    def update_progress(self, current_project_index, total_projects, current_project_name, status_text):
        """Update progress dialog with current status"""
        if hasattr(self, 'progress_dialog') and self.progress_dialog.winfo_exists():
            # Update progress bar
            self.progress_bar['value'] = current_project_index
            
            # Update current project
            self.progress_current_label.config(text=f"📋 Processing: {current_project_name}")
            
            # Update progress text
            self.progress_text_label.config(text=f"{current_project_index} of {total_projects} projects")
            
            # Update status
            self.progress_status_label.config(text=status_text)
            
            # Force update
            self.progress_dialog.update()
    
    def close_progress_dialog(self):
        """Close progress dialog"""
        if hasattr(self, 'progress_dialog') and self.progress_dialog.winfo_exists():
            self.progress_dialog.destroy()
    
    def show_column_selection_dialog(self, worksheet, target_month):
        """Show dialog to manually select target column when automatic matching fails"""
        import tkinter as tk
        from tkinter import ttk
        
        # Create dialog
        dialog = tk.Toplevel(self.root)
        dialog.title("Select Target Column")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Calculate center position immediately
        window_width = 600
        window_height = 400
        x = (dialog.winfo_screenwidth() // 2) - (window_width // 2)
        y = (dialog.winfo_screenheight() // 2) - (window_height // 2)
        
        # Set geometry with center position from the start
        dialog.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        selected_column = None
        
        # Main frame
        main_frame = ttk.Frame(dialog, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Instructions
        instructions = ttk.Label(main_frame, 
            text=f"Automatic column matching failed for target month: '{target_month}'\n"
                 f"Please select the target column from the Rolling P&L:", 
            font=("TkDefaultFont", 10))
        instructions.pack(pady=(0, 15))
        
        # Scrollable frame for columns
        canvas = tk.Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Column selection variable
        column_var = tk.IntVar()
        
        # Add radio buttons for each column
        for col in range(1, min(worksheet.max_column + 1, 50)):  # Limit to first 50 columns
            cell_value = worksheet.cell(row=1, column=col).value
            if cell_value:
                display_text = f"Column {col}: {str(cell_value).strip()}"
                ttk.Radiobutton(scrollable_frame, text=display_text, 
                               variable=column_var, value=col).pack(anchor="w", pady=2)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Buttons frame
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(15, 0))
        
        def on_ok():
            nonlocal selected_column
            selected_column = column_var.get() if column_var.get() > 0 else None
            dialog.destroy()
        
        def on_cancel():
            nonlocal selected_column
            selected_column = None
            dialog.destroy()
        
        ttk.Button(button_frame, text="OK", command=on_ok).pack(side=tk.RIGHT, padx=(10, 0))
        ttk.Button(button_frame, text="Cancel", command=on_cancel).pack(side=tk.RIGHT)
        
        # Wait for dialog to close
        dialog.wait_window()
        
        return selected_column
    
    def back_to_edit_mappings(self):
        """Return to Step 2 to edit mappings"""
        current_project = self.project_manager.get_current_project()
        if current_project:
            # Clear Step 3 data
            if hasattr(current_project, 'monthly_data'):
                delattr(current_project, 'monthly_data')
            if hasattr(current_project, 'aggregated_data'):
                delattr(current_project, 'aggregated_data')
            if hasattr(current_project, 'preview_data'):
                delattr(current_project, 'preview_data')
            if hasattr(current_project, 'target_month'):
                delattr(current_project, 'target_month')
        
        # Clear preview table
        for item in self.preview_tree.get_children():
            self.preview_tree.delete(item)
        
        # Disable Step 3 buttons
        self.finalize_button.config(state="disabled")
        
        self.step4_status_var.set("Returned to edit mappings. Please regenerate monthly statement after editing.")
        
        # Focus on mappings section (scroll to it)
        messagebox.showinfo("Edit Mappings", "You can now edit mappings in Step 2. Click 'Generate Monthly Statement' when ready.")
    
    # Step 3 Helper Methods
    def find_target_month_column(self, source_df, start_col, end_col):
        """Find the most recent month column within the entire spreadsheet (not limited to source range)"""
        from datetime import datetime
        import re
        
        # Check cache first
        current_project = self.project_manager.get_current_project()
        if current_project:
            cache_key = f"{current_project.name}:{current_project.source_file_path}"
            if cache_key in self.target_month_cache:
                cached_col, cached_header = self.target_month_cache[cache_key]
                # Verify the cached column still exists and has the same header
                if cached_col <= len(source_df.columns):

                    pass
                    return cached_col
        
        # Expand search to entire spreadsheet to find the most recent month
        # The source range is for account descriptions, but months could be anywhere
        search_start_col = 1
        search_end_col = len(source_df.columns)
        
        
        # Look for month headers in the entire spreadsheet
        month_columns = []
        
        for col_idx in range(search_start_col - 1, search_end_col):  # Convert to 0-based
            # Check multiple header rows for month information
            for header_row in [4, 5, 6]:
                if len(source_df) > header_row:
                    header_value = source_df.iloc[header_row, col_idx]
                    if pd.notna(header_value):
                        header_str = str(header_value).strip().lower()
                        
                        # Look for month/year patterns
                        month_patterns = [
                            r'\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\s*\d{4}',
                            r'\b(january|february|march|april|may|june|july|august|september|october|november|december)\s*\d{4}',
                            r'\d{1,2}[/-]\d{4}',  # MM/YYYY format
                            r'\d{4}-\d{1,2}',     # YYYY-MM format
                        ]
                        
                        for pattern in month_patterns:
                            if re.search(pattern, header_str):
                                month_columns.append((col_idx + 1, header_value, header_row))
                                break
        
        if not month_columns:
            # If no month patterns found, look for columns with "actual" or numeric data in entire spreadsheet
            for col_idx in range(search_end_col - 1, search_start_col - 2, -1):  # Search from right to left in entire sheet
                if col_idx < len(source_df.columns):

                    pass
                    for header_row in [4, 5, 6]:
                        if len(source_df) > header_row:
                            header_value = source_df.iloc[header_row, col_idx]
                            if pd.notna(header_value):
                                header_str = str(header_value).strip().lower()
                                if 'actual' in header_str or 'current' in header_str:
                                    # Cache the result
                                    if current_project:
                                        cache_key = f"{current_project.name}:{current_project.source_file_path}"
                                        self.target_month_cache[cache_key] = (col_idx + 1, header_value)
                                    return col_idx + 1
            
            # Last resort: use the rightmost column with data in entire spreadsheet
            for col_idx in range(search_end_col - 1, search_start_col - 2, -1):
                if col_idx < len(source_df.columns):
                    # Check if column has numeric data
                    has_data = False
                    for row_idx in range(7, min(len(source_df), 20)):  # Check a few data rows
                        value = source_df.iloc[row_idx, col_idx]
                        if pd.notna(value) and (isinstance(value, (int, float)) or 
                                              (isinstance(value, str) and value.strip() and value.strip() != '0')):
                            has_data = True
                            break
                    if has_data:
                        # Cache the result
                        if current_project:
                            cache_key = f"{current_project.name}:{current_project.source_file_path}"
                            self.target_month_cache[cache_key] = (col_idx + 1, "Data Column")
                        return col_idx + 1
        else:
            # Parse dates and find the most recent one
            parsed_months = []
            for col_idx, header_value, header_row in month_columns:
                try:
                    header_str = str(header_value).strip().lower()
                    
                    # Try to extract year from the header
                    year_match = re.search(r'(\d{4})', header_str)
                    month_match = re.search(r'(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)', header_str)
                    
                    if year_match and month_match:
                        year = int(year_match.group(1))
                        month_name = month_match.group(1)
                        
                        # Convert month name to number
                        month_map = {
                            'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
                            'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
                        }
                        month_num = month_map.get(month_name, 0)
                        
                        if month_num > 0:
                            # Create a sortable date tuple (year, month, col_idx)
                            parsed_months.append((year, month_num, col_idx, header_value))
                
                except Exception as e:
                    pass
            
            if parsed_months:
                # Sort by year and month (most recent first)
                parsed_months.sort(key=lambda x: (x[0], x[1]), reverse=True)
                target_col = parsed_months[0][2]
                target_header = parsed_months[0][3]
                # Cache the result
                if current_project:
                    cache_key = f"{current_project.name}:{current_project.source_file_path}"
                    self.target_month_cache[cache_key] = (target_col, target_header)
                return target_col
            else:
                # Fallback to rightmost column by position
                month_columns.sort(key=lambda x: x[0], reverse=True)  # Sort by column index, descending
                target_col = month_columns[0][0]
                # Cache the result
                if current_project:
                    cache_key = f"{current_project.name}:{current_project.source_file_path}"
                    self.target_month_cache[cache_key] = (target_col, month_columns[0][1])
                return target_col
        
        return None

    def extract_monthly_amounts(self, source_df, current_project):
        """Extract monthly amounts from the most recent month column in the source range"""
        import re
        from datetime import datetime
        monthly_amounts = {}
        
        # Parse source range to determine the actual data range and columns
        source_range = current_project.source_range
        
        if not source_range:

        
            pass
            return {}
            
        start_row, end_row, start_col, end_col = self.parse_excel_range(source_range)
        
        # Check if project already has a cached target month and use that column
        cached_target_month = getattr(current_project, 'target_month', '')
        target_col_index = None
        
        if cached_target_month and cached_target_month.strip():
            # Try to find the column that matches this cached month
            for col_idx in range(start_col - 1, min(end_col, len(source_df.columns))):
                for header_row in [4, 5, 6]:
                    if len(source_df) > header_row:
                        header_value = source_df.iloc[header_row, col_idx]
                        if pd.notna(header_value):
                            header_str = str(header_value).strip()
                            if header_str == cached_target_month:
                                target_col_index = col_idx + 1
                                break
                if target_col_index:

                    pass
                    break
        
        # If no cached column found, check the cache first before expensive search
        if not target_col_index:
            # Check cache first
            cache_key = f"{current_project.name}:{current_project.source_file_path}"
            if cache_key in self.target_month_cache:
                cached_col, _ = self.target_month_cache[cache_key]
                if cached_col <= len(source_df.columns):
                    target_col_index = cached_col
            else:
                # If not in cache, do the expensive search
                target_col_index = self.find_target_month_column(source_df, start_col, end_col)
            
            if not target_col_index:

            
                pass
                return {}
            
        target_col_index_0based = target_col_index - 1  # Convert to 0-based for pandas
        
        
        # Check if target column exists in the DataFrame
        if target_col_index_0based >= len(source_df.columns):

            pass
            return {}
        
        # Verify target column has a valid header
        target_month_header = "Unknown"
        if len(source_df) > 4:  # Check if we have enough rows
            for header_row in [4, 5, 6]:  # Check rows 5, 6, 7 for header
                if len(source_df) > header_row:
                    header_value = source_df.iloc[header_row, target_col_index_0based]
                    if pd.notna(header_value) and str(header_value).strip():
                        target_month_header = str(header_value).strip()
                        break
        
        # Also check what's in a few sample columns around target to help debug
        start_sample = max(0, target_col_index_0based - 2)
        end_sample = min(len(source_df.columns), target_col_index_0based + 3)
        for col_idx in range(start_sample, end_sample):
            sample_val = source_df.iloc[10, col_idx] if len(source_df) > 10 else "N/A"
            header_val = source_df.iloc[5, col_idx] if len(source_df) > 5 else "N/A"
        
        # Get source accounts - extract fresh from the range to include new accounts
        # Don't rely on mappings.keys() as it might not include new accounts like 8540
        source_accounts_from_range = self.extract_account_data("source", current_project.source_range)
        
        # Combine with existing mappings to ensure we get all accounts
        source_accounts = list(set(source_accounts_from_range) | set(current_project.mappings.keys() if current_project.mappings else []))
        
        # Extract account descriptions and amounts using the source range
        for account_desc in source_accounts:
            try:
                # Find the account within the specified source range
                found_row = None
                for row_idx in range(start_row - 1, min(end_row, len(source_df))):  # Convert to 0-based
                    # Check if this row contains our account description within the range
                    for col_idx in range(start_col - 1, min(end_col, len(source_df.columns))):  # Convert to 0-based
                        cell_value = source_df.iloc[row_idx, col_idx]
                        if pd.notna(cell_value) and str(cell_value).strip() == account_desc:
                            found_row = row_idx
                            break
                    if found_row is not None:

                        pass
                        break
                
                if found_row is not None:
                    # Extract amount from column Q
                    if target_col_index_0based >= 0 and target_col_index_0based < len(source_df.columns):
                        amount_value = source_df.iloc[found_row, target_col_index_0based]
                        
                        # Include amounts for all accounts, including subtotals
                        # Subtotals will be displayed but not used in aggregation
                        if pd.notna(amount_value) and str(amount_value).strip():
                            try:
                                # Convert to float, handling various formats
                                amount = float(str(amount_value).replace(',', '').replace('$', '').replace('(', '-').replace(')', '').strip())
                                monthly_amounts[account_desc] = amount
                            except ValueError:
                                pass  # Debug output removed
                        else:
                            pass
                else:
                    pass
                
            except Exception as e:
                pass  # Debug output removed
        
        # Cache the source amounts for future use
        cache_key = f"{current_project.name}:{current_project.source_file_path}:{current_project.source_range}"
        self.source_amounts_cache[cache_key] = monthly_amounts
        
        return monthly_amounts
    
    def aggregate_by_mappings(self, monthly_data, current_project):
        """Aggregate monthly amounts by their mapped categories"""
        aggregated = {}
        
        for account_desc, amount in monthly_data.items():
            if account_desc in current_project.mappings:
                mapped_category = current_project.mappings[account_desc]['rolling_account']
                
                if mapped_category in aggregated:
                    aggregated[mapped_category] += amount
                else:
                    aggregated[mapped_category] = amount
        
        
        return aggregated
    
    def has_account_number(self, account_desc):
        """Check if account description contains account numbers"""
        import re
        # Look for patterns like account numbers (digits, possibly with dashes or dots)
        account_patterns = [
            r'\b\d{3,}\b',      # 3+ consecutive digits
            r'\b\d+-\d+\b',     # Pattern like 1000-1500
            r'\b\d+\.\d+\b',    # Pattern like 1000.01
            r'#\d+',            # Pattern like #1001
            r'GL\s*\d+',        # Pattern like GL 1001
        ]
        
        account_desc_str = str(account_desc).upper()
        for pattern in account_patterns:
            if re.search(pattern, account_desc_str):

                pass
                return True
        return False
    
    def is_total_or_heading(self, account_desc):
        """Check if account description is a total, subtotal, or heading"""
        desc_lower = str(account_desc).lower()
        skip_keywords = [
            'total', 'subtotal', 'sum', 'grand total', 'sub total',
            'heading', 'header', '---', '___', 'section',
            'category', 'group', 'division', 'department total'
        ]
        
        for keyword in skip_keywords:
            if keyword in desc_lower:

                pass
                return True
        
        # Also skip if it's all caps (likely a heading)
        if str(account_desc).isupper() and len(str(account_desc)) > 3:

            pass
            return True
            
        return False
    
    def is_total_account(self, account_desc):
        """Check if account description contains 'TOTAL' and should be excluded from mapping options"""
        desc = str(account_desc).strip().upper()
        return 'TOTAL' in desc
    
    def is_income_account(self, account_name):
        """Check if account is likely an income/revenue account based on keywords"""
        account_lower = str(account_name).lower()
        income_keywords = [
            'income', 'revenue', 'rent', 'rental', 'lease', 'gross potential',
            'sales', 'fees', 'reimbursement', 'recovery', 'cam', 'tenant'
        ]
        return any(keyword in account_lower for keyword in income_keywords)
    
    def get_source_month_header(self, source_df):
        """Get the month header from source P&L - dynamically find target column"""
        try:
            current_project = self.project_manager.get_current_project()
            if not current_project:

                pass
                return "Unknown Month"
            
            # Parse source range to find target column dynamically
            source_range = current_project.source_range
            if not source_range:

                pass
                return "Unknown Month"
                
            start_row, end_row, start_col, end_col = self.parse_excel_range(source_range)
            
            # Check cache first before expensive search
            cache_key = f"{current_project.name}:{current_project.source_file_path}"
            if cache_key in self.target_month_cache:
                target_col_index, _ = self.target_month_cache[cache_key]
                if target_col_index > len(source_df.columns):
                    # Cache is invalid, do the search
                    target_col_index = self.find_target_month_column(source_df, start_col, end_col)
            else:
                target_col_index = self.find_target_month_column(source_df, start_col, end_col)
            
            if not target_col_index:

            
                pass
                return "Unknown Month"
                
            target_col_index_0based = target_col_index - 1  # Convert to 0-based for pandas
            
            
            if target_col_index_0based < len(source_df.columns):

            
            
                pass
                try:
                    # Try multiple header row positions (5, 6, 7) in target column
                    for header_row_idx in [4, 5, 6]:  # Rows 5, 6, 7 (0-based: 4, 5, 6)
                        if len(source_df) > header_row_idx:
                            header_value = source_df.iloc[header_row_idx, target_col_index_0based]
                            
                            if pd.notna(header_value) and str(header_value).strip():
                                header_str = str(header_value).strip()
                                # Look for any date pattern
                                if self.looks_like_date_header(header_str):

                                    pass
                                    return header_str
                            else:
                                pass
                        else:
                            pass
                except Exception as e:
                    pass  # Debug output removed
            else:

                pass
            
            # Second approach: Search row 6 for any date-like values
            if len(source_df) > 5:
                row_6_data = source_df.iloc[5]  # Row 6 (index 5)
                for col_idx, cell_value in enumerate(row_6_data):
                    if pd.notna(cell_value) and str(cell_value).strip():
                        cell_str = str(cell_value).strip()
                        # Check if it looks like a date
                        if self.looks_like_date_header(cell_str):

                            pass
                            return cell_str
            
            # Third approach: Search multiple rows (5, 6, 7) for date headers
            for row_idx in [4, 5, 6]:  # Rows 5, 6, 7 (0-based: 4, 5, 6)
                if len(source_df) > row_idx:
                    row_data = source_df.iloc[row_idx]
                    for col_idx, cell_value in enumerate(row_data):
                        if pd.notna(cell_value) and str(cell_value).strip():
                            cell_str = str(cell_value).strip()
                            if self.looks_like_date_header(cell_str):

                                pass
                                return cell_str
                
        except Exception as e:
            import traceback
            traceback.print_exc()
        
        return "Unknown Month"
    
    def looks_like_date_header(self, header_str):
        """Check if a header looks like a date/month header - enhanced version"""
        import re
        from datetime import datetime
        
        # If it's already a datetime object, it's definitely a date
        if isinstance(header_str, datetime):

            pass
            return True
            
        header_str = str(header_str).strip().lower()
        
        # Skip obvious non-dates
        if header_str in ['nan', '', 'none', 'null']:

            pass
            return False
        
        date_patterns = [
            r'\d{4}-\d{1,2}-\d{1,2}',  # YYYY-MM-DD
            r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}',  # MM/DD/YYYY or MM-DD-YY
            r'\d{4}[/-]\d{1,2}[/-]\d{1,2}',    # YYYY/MM/DD
            r'(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\s+\d{4}',  # Month names with year
            r'\d{1,2}/\d{4}',  # MM/YYYY
            r'(january|february|march|april|may|june|july|august|september|october|november|december)\s+\d{4}',
            r'(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\s+\d{4}\s+actual',  # Month Year Actual
            r'(january|february|march|april|may|june|july|august|september|october|november|december)\s+\d{4}\s+actual',
            r'\d{4}',  # Just a year (4 digits)
        ]
        
        for pattern in date_patterns:
            if re.search(pattern, header_str):

                pass
                return True
        
        # Try to parse as date
        try:
            datetime.strptime(header_str, '%Y-%m-%d')
            return True
        except:
            pass
            
        try:
            datetime.strptime(header_str, '%m/%d/%Y')
            return True
        except:
            pass
        
        return False
    
    def parse_excel_range(self, range_str):
        """Parse Excel range string like 'A8:F170' to get start/end rows and columns"""
        import re
        try:

            pass
            if ':' in range_str:
                start_cell, end_cell = range_str.split(':')
                
                # Parse start cell (e.g., 'A8')
                start_match = re.match(r'([A-Z]+)(\d+)', start_cell.strip())
                if start_match:
                    start_col_letters, start_row_str = start_match.groups()
                    start_row = int(start_row_str)
                    start_col = self.excel_col_to_num(start_col_letters)
                
                # Parse end cell (e.g., 'F170')
                end_match = re.match(r'([A-Z]+)(\d+)', end_cell.strip())
                if end_match:
                    end_col_letters, end_row_str = end_match.groups()
                    end_row = int(end_row_str)
                    end_col = self.excel_col_to_num(end_col_letters)
                
                return start_row, end_row, start_col, end_col
            else:
                # Single cell reference
                match = re.match(r'([A-Z]+)(\d+)', range_str.strip())
                if match:
                    col_letters, row_str = match.groups()
                    row = int(row_str)
                    col = self.excel_col_to_num(col_letters)
                    return row, row, col, col
        except Exception as e:
            pass  # Debug output removed
        return 1, 1000, 1, 10
    
    def excel_col_to_num(self, col_str):
        """Convert Excel column letters to number (A=1, B=2, ..., Z=26, AA=27, etc.)"""
        num = 0
        for char in col_str:
            num = num * 26 + (ord(char.upper()) - ord('A') + 1)
        return num
    
    def num_to_excel_col(self, col_num):
        """Convert column number to Excel column letters (1=A, 2=B, ..., 26=Z, 27=AA, etc.)"""
        result = ""
        while col_num > 0:
            col_num -= 1  # Adjust for 0-based
            result = chr(col_num % 26 + ord('A')) + result
            col_num //= 26
        return result
    
    def prepare_preview_data(self, rolling_df, target_month, aggregated_data):
        """Prepare preview data in rolling P&L sequential order, showing mapped data and preserving non-income accounts"""
        preview_data = []
        
        # Update preview tree column header with target month
        self.preview_tree.heading("Target Month", text=target_month, anchor=tk.W)
        
        # Get current target month values from rolling P&L
        current_target_values = self.get_current_target_month_values(target_month)
        
        # Get historical data for previous months
        current_project = self.project_manager.get_current_project()
        historical_data = self.get_historical_data_for_preview(current_project) if current_project else None
        
        # Debug: Print historical data status
        if historical_data:

            pass
            if 'month_headers' in historical_data:
                pass
            else:
                pass
        else:
            pass
        
        # Update column headers with actual month names
        if historical_data and 'month_headers' in historical_data:
            headers = historical_data['month_headers']
            if len(headers) >= 3:
                self.preview_tree.heading("Previous Month -2", text=headers[0])
                self.preview_tree.heading("Previous Month -1", text=headers[1])
                self.preview_tree.heading("Target Month", text=headers[2])
            else:

                pass
        
        # Get the set of rolling accounts that have mapped data
        mapped_rolling_accounts = set(aggregated_data.keys())
        
        # Use rolling P&L order by iterating through rolling_df
        # This preserves the sequential order as it appears in the rolling P&L
        if rolling_df is not None and not rolling_df.empty:
            import pandas as pd
            # Get the account description column (usually first column)
            desc_col = rolling_df.columns[0]
            
            for idx, row in rolling_df.iterrows():
                account_desc = row[desc_col]
                if pd.notna(account_desc) and str(account_desc).strip():
                    account_name = str(account_desc).strip()
                    
                    # Determine which amount to show:
                    # For income accounts that could receive mapped data:
                    #   - If mapped: use the aggregated amount
                    #   - If not mapped: show 0 (don't carry forward old values)
                    # For other accounts: preserve existing values
                    
                    if account_name in aggregated_data:
                        # This account has new mapped data from source P&L
                        amount = aggregated_data[account_name]
                    elif self.is_income_account(account_name) and len(mapped_rolling_accounts) > 0:
                        # This is an income account that could have received mapped data but didn't
                        # Set to 0 to avoid showing stale/duplicate values
                        amount = 0.0
                    else:
                        # Non-income account or no mappings exist - preserve existing value
                        amount = current_target_values.get(account_name, 0.0)
                    
                    # Get historical amounts for this account
                    prev_month_2 = 0.0
                    prev_month_1 = 0.0
                    
                    if historical_data and 'account_data' in historical_data:
                        # Try to find matching account in historical data
                        account_history = None
                        
                        # First try exact match
                        if account_name in historical_data['account_data']:
                            account_history = historical_data['account_data'][account_name]
                        else:
                            # Try fuzzy matching for similar account names
                            for hist_account, hist_data in historical_data['account_data'].items():
                                if account_name.lower() in hist_account.lower() or hist_account.lower() in account_name.lower():
                                    account_history = hist_data
                                    break
                        
                        if account_history and isinstance(account_history, dict):
                            prev_month_2 = account_history.get('prev_2', 0.0)
                            prev_month_1 = account_history.get('prev_1', 0.0)
                    
                    row_data = {
                        'account': account_name,
                        'prev_month_2': prev_month_2,
                        'prev_month_1': prev_month_1,
                        'target_amount': amount,  # Keep as numeric for formatting later
                        'rolling_order': idx  # Store original order for reference
                    }
                    preview_data.append(row_data)
        else:
            # Fallback: if no rolling_df, use aggregated_data order (original behavior)
            for account, amount in aggregated_data.items():
                row_data = {
                    'account': account,
                    'prev_month_2': 0.0,
                    'prev_month_1': 0.0,
                    'target_amount': amount,
                    'rolling_order': 999999  # Put at end if no order available
                }
                preview_data.append(row_data)
        
        # Ensure data is sorted by rolling order (even though it should already be in order)
        preview_data.sort(key=lambda x: x.get('rolling_order', 999999))
        
        return preview_data
    
    def get_current_target_month_values(self, target_month):
        """Extract current values from the target month column in rolling P&L"""
        try:
            current_project = self.project_manager.get_current_project()
            if not current_project or not self.project_manager.rolling_workbook_path or not current_project.rolling_sheet:

                pass
                return {}
            
                
            # Load rolling P&L workbook
            from openpyxl import load_workbook
            wb = load_workbook(self.project_manager.rolling_workbook_path, data_only=True)
            
            if current_project.rolling_sheet not in wb.sheetnames:

            
                pass
                return {}
                
            ws = wb[current_project.rolling_sheet]
            
            # Use the existing find_matching_column function for proper month matching
            target_col = self.find_matching_column(ws, target_month)
            
            if not target_col:

            
                pass
                return {}
            
            
            # Extract account data for the range
            current_values = {}
            
            # Parse rolling range to get row numbers
            rolling_rows = self.parse_range_for_rows(current_project.rolling_range)
            
            for row in range(rolling_rows[0], rolling_rows[1] + 1):
                # Get account description from column A (or first column in range)
                account_cell = ws.cell(row=row, column=1).value
                if account_cell and str(account_cell).strip():
                    account_name = str(account_cell).strip()
                    
                    # Get value from target month column
                    cell_value = ws.cell(row=row, column=target_col).value
                    
                    # Debug: Show all values for first few rows
                    if row <= rolling_rows[0] + 10:  # First 10 rows
                        pass
                        
                    # Check if account contains "electricity" (case insensitive)
                    if "electricity" in account_name.lower() or "electric" in account_name.lower():
                        # Show values in nearby columns too
                        for nearby_col in range(max(1, target_col-2), min(ws.max_column+1, target_col+3)):
                            nearby_value = ws.cell(row=row, column=nearby_col).value
                    
                    # Handle both numeric values and Excel formulas
                    if cell_value is not None and isinstance(cell_value, (int, float)):
                        current_values[account_name] = cell_value
                        if "electric" in account_name.lower():
                            pass
                        else:
                            pass
                    elif cell_value is not None and isinstance(cell_value, str) and cell_value.startswith('='):
                        # Try to evaluate simple Excel formulas
                        try:
                            # Handle simple arithmetic formulas like "=93+640"
                            formula = cell_value[1:]  # Remove the '=' sign
                            # Replace Excel operators with Python equivalents
                            formula = formula.replace('^', '**')  # Power operator
                            # Evaluate simple arithmetic expressions
                            # Allow numbers, operators, parentheses, and spaces
                            allowed_chars = '0123456789+-*/()^. '
                            if all(c in allowed_chars for c in formula):
                                result = eval(formula)
                                if "electric" in account_name.lower():
                                    pass
                            else:
                                result = 0.0
                                if "electric" in account_name.lower():
                                    pass
                            current_values[account_name] = float(result)
                            if "electric" in account_name.lower():
                                pass
                            else:
                                pass
                        except:
                            current_values[account_name] = 0.0
                            if "electric" in account_name.lower():
                                pass
                    else:
                        current_values[account_name] = 0.0
                        if "electric" in account_name.lower():
                            pass
            
            return current_values
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return {}
    
    def populate_preview_table(self, preview_data):
        """Populate the preview table with historical data and target amounts"""
        # Clear existing items
        for item in self.preview_tree.get_children():
            self.preview_tree.delete(item)
        
        # Get current project for historical data extraction
        current_project = self.project_manager.get_current_project()
        if not current_project:

            pass
            return
        
        # Always fetch fresh historical data when loading a project to ensure current data
        # This ensures that when switching projects, we get the correct historical data
        historical_data = self.get_historical_data_for_preview(current_project)
        
        # Always use fresh historical data to ensure current data when switching projects
        
        # Update column headers with actual month names from fresh historical data
        if historical_data and 'month_headers' in historical_data:
            headers = historical_data['month_headers']
            if len(headers) >= 3:
                self.preview_tree.heading("Previous Month -2", text=headers[0])
                self.preview_tree.heading("Previous Month -1", text=headers[1])
                self.preview_tree.heading("Target Month", text=headers[2])
        
        # Add preview data with historical context
        for row_data in preview_data:
            account_name = row_data['account']
            target_amount = row_data['target_amount']
            
            # Get historical amounts for this account
            # Use fresh historical data for all cases to ensure current data
            prev_month_2 = ""
            prev_month_1 = ""
            
            if historical_data and 'account_data' in historical_data:
                # Try to find matching account in historical data
                # The account_name here is the rolling account name from mappings
                account_history = None
                
                # First try exact match
                if account_name in historical_data['account_data']:
                    account_history = historical_data['account_data'][account_name]
                else:
                    # Try fuzzy matching for similar account names
                    for hist_account, hist_data in historical_data['account_data'].items():
                        if account_name.lower() in hist_account.lower() or hist_account.lower() in account_name.lower():
                            account_history = hist_data
                            break
                
                if account_history:
                    # Show blank for total accounts in historical data too
                    if "total" in account_name.lower():
                        prev_month_2 = ""
                        prev_month_1 = ""
                    else:
                        prev_month_2 = self.format_currency(account_history.get('prev_2', 0))
                        prev_month_1 = self.format_currency(account_history.get('prev_1', 0))
            
            # Format target amount - show blank for total accounts to avoid confusion
            if "total" in account_name.lower():

                pass
                formatted_target = ""
            else:

                pass
                formatted_target = self.format_currency(target_amount)
            
            self.preview_tree.insert("", "end", values=(
                account_name,
                prev_month_2,
                prev_month_1,
                formatted_target
            ))
    
    def get_historical_data_for_preview(self, current_project):
        """Extract historical data from rolling P&L for the previous 2 months"""
        try:

            pass
            if not self.project_manager.rolling_workbook_path or not current_project.rolling_sheet:

                pass
                return None
            
            # Load rolling P&L workbook
            from openpyxl import load_workbook
            wb = load_workbook(self.project_manager.rolling_workbook_path, data_only=True)
            
            if current_project.rolling_sheet not in wb.sheetnames:

            
                pass
                return None
                
            ws = wb[current_project.rolling_sheet]
            
            # Find target month column
            target_col = self.find_matching_column(ws, current_project.target_month)
            if not target_col:

                pass
                return None
            
            
            # Find previous 2 months (columns before target)
            prev_col_1 = target_col - 1  # Previous month
            prev_col_2 = target_col - 2  # Month before previous
            
            # Get month headers
            month_headers = []
            for col in [prev_col_2, prev_col_1, target_col]:
                if col > 0:
                    header_value = ws.cell(row=1, column=col).value
                    if header_value:
                        # Format header for display
                        header_str = str(header_value).strip()
                        # If it's a datetime, format it nicely
                        try:
                            import datetime
                            if isinstance(header_value, datetime.datetime):
                                header_str = header_value.strftime("%b %Y")
                            elif len(header_str) > 15:  # Truncate very long headers
                                header_str = header_str[:12] + "..."
                        except:
                            pass
                        month_headers.append(header_str)
                    else:
                        month_headers.append(f"Column {col}")
                else:
                    month_headers.append("N/A")
            
            # Extract account data for the range
            account_data = {}
            
            # Parse rolling range to get row numbers
            rolling_rows = self.parse_range_for_rows(current_project.rolling_range)
            
            for row in range(rolling_rows[0], rolling_rows[1] + 1):
                # Get account description from column A (or first column in range)
                account_cell = ws.cell(row=row, column=1).value
                if account_cell and str(account_cell).strip():
                    account_name = str(account_cell).strip()
                    
                    # Get values from previous 2 months and target month
                    prev_2_value = 0
                    prev_1_value = 0
                    
                    if prev_col_2 > 0:
                        cell_value = ws.cell(row=row, column=prev_col_2).value
                        if cell_value is not None and isinstance(cell_value, (int, float)):
                            prev_2_value = cell_value
                        elif cell_value is not None and isinstance(cell_value, str) and cell_value.startswith('='):
                            # Evaluate Excel formulas for historical data
                            try:

                                pass
                                formula = cell_value[1:]  # Remove the '=' sign
                                formula = formula.replace('^', '**')  # Power operator
                                allowed_chars = '0123456789+-*/()^. '
                                if all(c in allowed_chars for c in formula):
                                    prev_2_value = float(eval(formula))
                                else:
                                    prev_2_value = 0.0
                            except:
                                prev_2_value = 0.0
                    
                    if prev_col_1 > 0:
                        cell_value = ws.cell(row=row, column=prev_col_1).value
                        if cell_value is not None and isinstance(cell_value, (int, float)):
                            prev_1_value = cell_value
                        elif cell_value is not None and isinstance(cell_value, str) and cell_value.startswith('='):
                            # Evaluate Excel formulas for historical data
                            try:

                                pass
                                formula = cell_value[1:]  # Remove the '=' sign
                                formula = formula.replace('^', '**')  # Power operator
                                allowed_chars = '0123456789+-*/()^. '
                                if all(c in allowed_chars for c in formula):
                                    prev_1_value = float(eval(formula))
                                    if "electric" in account_name.lower():
                                        pass
                                else:
                                    prev_1_value = 0.0
                                    if "electric" in account_name.lower():
                                        pass
                            except Exception as e:
                                prev_1_value = 0.0
                                if "electric" in account_name.lower():
                                    pass
                    
                    account_data[account_name] = {
                        'prev_2': prev_2_value,
                        'prev_1': prev_1_value
                    }
            
            return {
                'month_headers': month_headers,
                'account_data': account_data
            }
            
        except Exception as e:

            
            pass
            return None
    
    def format_currency(self, value):
        """Format numeric value as currency"""
        try:
            num_value = float(str(value).replace(',', '').replace('$', ''))
            if num_value < 0:

                pass
                return f"(${abs(num_value):,.2f})"
            else:

                pass
                return f"${num_value:,.2f}"
        except:

            pass
            return str(value)
    
    def find_matching_column(self, worksheet, target_month):
        """Find matching column in rolling P&L based on target month header"""
        # Check row 1 for headers
        for col in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=1, column=col).value
            
            if cell_value and self.headers_match(cell_value, target_month):

            
                pass
                return col
        
        return None
    
    def headers_match(self, header1, header2):
        """Check if two headers match, ignoring formatting differences"""
        import re
        from datetime import datetime
        
        
        # Normalize headers - remove extra spaces, convert to lowercase
        h1 = str(header1).strip()
        h2 = str(header2).strip()
        
        # Direct match
        if h1.lower() == h2.lower():

            pass
            return True
        
        # Try to parse both as dates and compare month/year
        def extract_month_year(header_value):
            """Extract month and year from various date formats"""
            try:
                # First check if it's already a datetime object
                if isinstance(header_value, datetime):

                    pass
                    return header_value.month, header_value.year
                
                # Convert to string for parsing
                header_str = str(header_value).strip()
                
                # Try common date formats
                date_formats = [
                    '%m/%d/%Y',    # 6/22/2025
                    '%m/%Y',       # 6/2025
                    '%B %Y',       # June 2025
                    '%b %Y',       # Jun 2025
                    '%B %Y Actual',       # June 2025 Actual
                    '%b %Y Actual',       # Jun 2025 Actual
                    '%Y-%m-%d',    # 2025-06-22
                    '%Y/%m/%d',    # 2025/6/22
                ]
                
                for fmt in date_formats:
                    try:
                        parsed_date = datetime.strptime(header_str, fmt)
                        return parsed_date.month, parsed_date.year
                    except ValueError:

                        pass
                        continue
                
                # Handle special formats like "Jun 2025 Actual"
                if 'actual' in header_str.lower():
                    # Remove "Actual" and try again
                    clean_header = header_str.replace('Actual', '').replace('actual', '').strip()
                    for fmt in ['%B %Y', '%b %Y']:
                        try:
                            parsed_date = datetime.strptime(clean_header, fmt)
                            return parsed_date.month, parsed_date.year
                        except ValueError:

                            pass
                            continue
                
                # Try partial matching for formats like "June 2025"
                month_map = {
                    'january': 1, 'jan': 1, 'february': 2, 'feb': 2, 'march': 3, 'mar': 3,
                    'april': 4, 'apr': 4, 'may': 5, 'june': 6, 'jun': 6,
                    'july': 7, 'jul': 7, 'august': 8, 'aug': 8, 'september': 9, 'sep': 9,
                    'october': 10, 'oct': 10, 'november': 11, 'nov': 11, 'december': 12, 'dec': 12
                }
                
                header_lower = header_str.lower()
                year_match = re.search(r'\b(20\d{2})\b', header_str)
                
                if year_match:
                    year = int(year_match.group(1))
                    for month_name, month_num in month_map.items():
                        if month_name in header_lower:

                            pass
                            return month_num, year
                            
            except Exception as e:
                pass
            
            return None, None
        
        # Extract month/year from both headers
        month1, year1 = extract_month_year(header1)
        month2, year2 = extract_month_year(header2)
        
        # Compare month and year
        if month1 and year1 and month2 and year2:
            match = (month1 == month2 and year1 == year2)
            return match
        
        return False
    
    def create_sum_formula(self, existing_value, new_value):
        """Create a SUM formula that combines existing value/formula with new value
        
        Args:
            existing_value: The current cell value (can be None, number, or formula string)
            new_value: The new value to add
            
        Returns:
            The value to write to the cell (either a plain value or a formula string)
        """
        # Handle None or empty cells
        if existing_value is None or existing_value == "" or existing_value == 0:
            # Empty cell or zero - just write the new value directly
            return new_value
        
        # Format new value for formula (handle negatives properly)
        if new_value < 0:
            new_value_str = f"({new_value})"
        else:
            new_value_str = str(new_value)
        
        # Handle existing formulas
        if isinstance(existing_value, str) and existing_value.startswith('='):
            # Existing formula - wrap it and add new value
            # Remove the leading '=' from existing formula
            existing_formula = existing_value[1:]
            
            # Check formula length limit (Excel has 8192 character limit)
            new_formula = f"=({existing_formula})+{new_value_str}"
            if len(new_formula) > 8000:  # Leave some buffer
                # Formula too long, just return the new value
                print(f"Warning: Formula would be too long, writing value directly instead")
                return new_value
            
            return new_formula
        
        # Handle existing plain numbers
        elif isinstance(existing_value, (int, float)):
            # Format existing value for formula
            if existing_value < 0:
                existing_value_str = f"({existing_value})"
            else:
                existing_value_str = str(existing_value)
            
            # Create SUM formula
            return f"={existing_value_str}+{new_value_str}"
        
        # Handle any other type (text, dates, etc.) - just write new value
        else:
            return new_value
    
    def write_data_to_rolling_pnl(self, worksheet, target_col, current_project):
        """Write aggregated data to the rolling P&L worksheet, creating SUM formulas to preserve existing values"""
        rolling_range = current_project.rolling_range
        
        # Parse rolling range to determine rows and column to check
        start_row, end_row, account_col = self.parse_rolling_range_for_account_column(rolling_range)
        
        # For each row in the rolling range, check if account matches aggregated data
        for row in range(start_row, end_row + 1):
            # Get account description from the specific rolling range column (not search all columns)
            cell_value = worksheet.cell(row=row, column=account_col).value
            account_desc = None
            if cell_value and str(cell_value).strip():
                account_desc = str(cell_value).strip()
            
            # If we found an account description, check if we have data for it
            if account_desc and account_desc in current_project.aggregated_data:
                new_value = current_project.aggregated_data[account_desc]
                
                # Get the existing cell and its value
                existing_cell = worksheet.cell(row=row, column=target_col)
                existing_value = existing_cell.value
                
                # Create SUM formula or value to write
                value_to_write = self.create_sum_formula(existing_value, new_value)
                
                # Write the value/formula to the cell
                worksheet.cell(row=row, column=target_col).value = value_to_write
                
                # Log what we did for debugging
                if isinstance(value_to_write, str) and value_to_write.startswith('='):
                    print(f"Row {row}: Created formula for '{account_desc}': {value_to_write[:50]}...")
                else:
                    print(f"Row {row}: Wrote value for '{account_desc}': {value_to_write}")
    
    def parse_rolling_range_for_account_column(self, range_str):
        """Parse rolling range string to get start row, end row, and account column number"""
        try:

            pass
            if ':' in range_str:
                parts = range_str.split(':')
                start_cell = parts[0].strip().upper()
                end_cell = parts[1].strip().upper()
                
                # Extract row numbers
                start_row = int(''.join(filter(str.isdigit, start_cell))) if any(c.isdigit() for c in start_cell) else 1
                end_row = int(''.join(filter(str.isdigit, end_cell))) if any(c.isdigit() for c in end_cell) else 1000
                
                # Extract column from start cell (the account column)
                start_col_letters = ''.join(filter(str.isalpha, start_cell))
                account_col = self.column_number_from_letters(start_col_letters) if start_col_letters else 1
                
                return start_row, end_row, account_col
            else:
                # Single column reference, assume all rows
                col_letters = ''.join(filter(str.isalpha, range_str.upper()))
                account_col = self.column_number_from_letters(col_letters) if col_letters else 1
                return 1, 1000, account_col
        except Exception as e:

            pass
            return 1, 1000, 1
    
    def column_number_from_letters(self, letters):
        """Convert Excel column letters to number (A->1, B->2, AA->27, etc.)"""
        result = 0
        for char in letters.upper():
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result
    
    def column_letter_from_number(self, col_num):
        """Convert column number to Excel letter (1->A, 2->B, etc.)"""
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(ord('A') + col_num % 26) + result
            col_num //= 26
        return result
    
    def parse_range_for_rows(self, range_str):
        """Parse range string to get start and end row numbers"""
        try:

            pass
            if ':' in range_str:
                parts = range_str.split(':')
                start_cell = parts[0].strip().upper()
                end_cell = parts[1].strip().upper()
                
                # Extract row numbers
                start_row = int(''.join(filter(str.isdigit, start_cell))) if any(c.isdigit() for c in start_cell) else 1
                end_row = int(''.join(filter(str.isdigit, end_cell))) if any(c.isdigit() for c in end_cell) else 1000
                
                return start_row, end_row
            else:
                # Single column reference, assume all rows
                return 1, 1000
        except:

            pass
            return 1, 1000
    
    # UI Helper Methods
    def setup_treeview_styles(self):
        """Configure styles for left-aligned column headers"""
        try:
            style = ttk.Style()
            style.configure("Treeview.Heading", anchor=tk.W)
        except Exception as e:
            pass  # Debug output removed
    
    def apply_header_alignment(self):
        """Apply left alignment to all treeview headers after widgets are created"""
        try:
            # Re-apply header configurations with left alignment
            if hasattr(self, 'mapping_tree'):
                self.mapping_tree.heading("Select", anchor=tk.W)
                self.mapping_tree.heading("Account Description", anchor=tk.W)
                self.mapping_tree.heading("Mapped Account", anchor=tk.W)
                self.mapping_tree.heading("Confidence", anchor=tk.W)
        except Exception as e:
            pass  # Debug output removed
    
    
    def setup_keyboard_shortcuts(self):
        """Setup keyboard shortcuts"""
        # Focus on the root window to ensure keyboard events are captured
        self.root.focus_set()
        
        # Bind zoom shortcuts - use KeyPress events for better compatibility
        self.root.bind_all("<Control-KeyPress-equal>", lambda e: self.zoom_in())
        self.root.bind_all("<Control-KeyPress-plus>", lambda e: self.zoom_in())
        self.root.bind_all("<Control-KeyPress-minus>", lambda e: self.zoom_out())
        
        # Mac shortcuts
        if platform.system() == "Darwin":
            self.root.bind_all("<Command-KeyPress-equal>", lambda e: self.zoom_in())
            self.root.bind_all("<Command-KeyPress-plus>", lambda e: self.zoom_in())
            self.root.bind_all("<Command-KeyPress-minus>", lambda e: self.zoom_out())
    
    def zoom_in(self):
        """Increase zoom level"""
        if self.current_zoom_index < len(self.zoom_levels) - 1:
            self.current_zoom_index += 1
            self.current_zoom_level = self.zoom_levels[self.current_zoom_index]
            self.apply_zoom()
    
    def zoom_out(self):
        """Decrease zoom level"""
        if self.current_zoom_index > 0:
            self.current_zoom_index -= 1
            self.current_zoom_level = self.zoom_levels[self.current_zoom_index]
            self.apply_zoom()
    
    def apply_zoom(self):
        """Apply current zoom level to the interface"""
        # Simple implementation - just update fonts
        try:
            # Update treeview font
            style = ttk.Style()
            new_size = int(14 * self.current_zoom_level)
            style.configure("Large.Treeview", font=("Arial", new_size))
            
            # Update tag fonts
            heading_size = int(16 * self.current_zoom_level)
            normal_size = int(14 * self.current_zoom_level)
            
            if hasattr(self, 'mapping_tree'):
                self.mapping_tree.tag_configure("heading", font=("Arial", heading_size, "bold"))
                self.mapping_tree.tag_configure("normal", font=("Arial", normal_size))
                
        except Exception as e:
            pass  # Debug output removed
    
    def on_tree_click(self, event):
        """Handle tree click events for checkbox toggling and header toggle all"""
        # Ensure the tree has focus for keyboard events
        self.mapping_tree.focus_set()
        
        region = self.mapping_tree.identify_region(event.x, event.y)
        
        if region == "heading":
            # Handle header clicks
            column = self.mapping_tree.identify_column(event.x)
            if column == "#1":  # Clicked on Select column header
                self.toggle_all_selection()
        
        elif region == "cell":
            # Get the item and column clicked
            item = self.mapping_tree.identify_row(event.y)
            column = self.mapping_tree.identify_column(event.x)
            
            # Only toggle if checkbox column (column #1) and it's not a header row
            if column == "#1" and item:
                # Get item tags to check if it's a heading
                tags = self.mapping_tree.item(item, "tags")
                if tags and "heading" in tags:

                    pass
                    return  # Skip headings
                
                # Initialize checkbox state if not exists
                if item not in self.checkbox_states:
                    self.checkbox_states[item] = False
                
                # Toggle checkbox state
                current_state = self.checkbox_states.get(item, False)
                new_state = not current_state
                self.checkbox_states[item] = new_state
                
                # Update checkbox display
                values = self.mapping_tree.item(item, "values")
                new_values = list(values)
                new_values[0] = "☑" if new_state else "☐"
                self.mapping_tree.item(item, values=new_values)
                
                # Update selection counter
                self.update_selection_counter()
    
    def toggle_all_selection(self):
        """Toggle all checkbox selections"""
        # Count currently selected items
        selected_count = sum(1 for selected in self.checkbox_states.values() if selected)
        total_count = len(self.checkbox_states)
        
        # If more than half are selected, deselect all; otherwise select all
        select_all = selected_count < (total_count / 2)
        
        if select_all:
            self.select_all_items()
        else:
            self.deselect_all_items()
    
    def cleanup_checkbox_states(self):
        """Remove checkbox states for items that no longer exist in any tree"""
        # Get all valid items from both trees
        valid_items = set()
        
        # Add items from main tree
        if hasattr(self, 'mapping_tree'):

            pass
            for item in self.mapping_tree.get_children():
                valid_items.add(item)
        
        # Add items from popup tree if it exists
        if hasattr(self, 'popup_tree') and self.popup_tree:

            pass
            for item in self.popup_tree.get_children():
                valid_items.add(item)
        
        # Remove orphaned checkbox states
        orphaned = [item for item in self.checkbox_states.keys() if item not in valid_items]
        for item in orphaned:
            del self.checkbox_states[item]
    
    def update_selection_counter(self):
        """Update the selection counter label"""
        # Clean up orphaned states first
        self.cleanup_checkbox_states()
        
        selected_count = sum(1 for selected in self.checkbox_states.values() if selected)
        self.selection_label.config(text=f"{selected_count} items selected")
    
    def select_all_items(self):
        """Select all items in the mapping tree"""
        for item in self.mapping_tree.get_children():
            values = self.mapping_tree.item(item, "values")
            if len(values) >= 2 and values[1]:  # Has account description
                account_desc = values[1]
                # Skip headers and totals
                if not any(keyword in account_desc.lower() for keyword in 
                          ['total', 'income', 'expense', 'revenue', 'cost']):
                    self.checkbox_states[item] = True
                    new_values = list(values)
                    new_values[0] = "☑"
                    self.mapping_tree.item(item, values=new_values)
        
        self.update_selection_counter()
    
    def deselect_all_items(self):
        """Deselect all items in the mapping tree"""
        for item in self.mapping_tree.get_children():
            self.checkbox_states[item] = False
            values = self.mapping_tree.item(item, "values")
            if values:
                new_values = list(values)
                new_values[0] = "☐"
                self.mapping_tree.item(item, values=new_values)
        
        self.update_selection_counter()
    
    @profile_performance("edit_mapping")
    def edit_mapping(self, event):
        """Edit single mapping on double-click - Adapted from main.py"""
        item = self.mapping_tree.selection()[0] if self.mapping_tree.selection() else None
        if not item:

            pass
            return
            
        values = self.mapping_tree.item(item, "values")
        if len(values) < 2:

            pass
            return
            
        account_desc = values[1]
        current_mapping = values[3] if len(values) > 3 else ""  # Column 3 = "Mapped Account"
        
        # Skip headers and totals using proper detection logic
        if self.is_total_or_heading(account_desc):

            pass
            return
        
        # Create edit dialog
        dialog = tk.Toplevel(self.root)
        dialog.title("Edit Mapping")
        dialog.resizable(True, True)
        
        # Set geometry and center immediately before showing
        width, height = 600, 400
        x = (dialog.winfo_screenwidth() // 2) - (width // 2)
        y = (dialog.winfo_screenheight() // 2) - (height // 2)
        dialog.geometry(f"{width}x{height}+{x}+{y}")
        
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Main container with padding
        main_frame = ttk.Frame(dialog, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Account description (read-only)
        desc_frame = ttk.LabelFrame(main_frame, text="Account Description", padding="10")
        desc_frame.pack(fill=tk.X, pady=(0, 15))
        
        ttk.Label(desc_frame, text=account_desc, font=("Arial", 11), 
                 wraplength=550, justify="left").pack(anchor="w")
        
        # Mapping selection frame
        mapping_frame = ttk.LabelFrame(main_frame, text="Map to Rolling Account", padding="10")
        mapping_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
        
        # Search bar
        search_frame = ttk.Frame(mapping_frame)
        search_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(search_frame, text="🔍 Search:", font=("Arial", 10)).pack(side=tk.LEFT, padx=(0, 5))
        search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=search_var, width=40)
        search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        ttk.Button(search_frame, text="Clear", 
                  command=lambda: search_var.set(""), width=8).pack(side=tk.LEFT)
        
        # Get rolling accounts for dropdown - use cache if available
        current_project = self.project_manager.get_current_project()
        rolling_accounts = []
        if current_project:

            pass
            try:
                # Check cache first
                cache_key = f"{current_project.name}:{self.rolling_range_var.get()}"
                if cache_key in self.rolling_accounts_cache:
                    rolling_accounts = self.rolling_accounts_cache[cache_key]
                else:
                    # Fall back to extracting if not cached
                    rolling_data = self.extract_account_data("rolling", self.rolling_range_var.get())
                    rolling_accounts = sorted(rolling_data) if rolling_data else []
                    # Cache for future use
                    if rolling_accounts:
                        self.rolling_accounts_cache[cache_key] = rolling_accounts
            except:
                pass
        
        # Selected mapping display
        mapping_var = tk.StringVar(value=current_mapping)
        selected_frame = ttk.Frame(mapping_frame)
        selected_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(selected_frame, text="Selected Mapping:", font=("Arial", 10, "bold")).pack(anchor="w")
        selected_label = ttk.Label(selected_frame, text=current_mapping or "None selected", 
                                  font=("Arial", 10), foreground="blue")
        selected_label.pack(anchor="w", pady=(2, 0))
        
        # Rolling accounts listbox with scrollbar
        list_frame = ttk.Frame(mapping_frame)
        list_frame.pack(fill=tk.BOTH, expand=True)
        
        accounts_listbox = tk.Listbox(list_frame, font=("Arial", 10), height=8)
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=accounts_listbox.yview)
        accounts_listbox.configure(yscrollcommand=scrollbar.set)
        
        accounts_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Populate listbox
        def populate_listbox(filter_text=""):
            accounts_listbox.delete(0, tk.END)
            filtered_accounts = [acc for acc in rolling_accounts 
                               if filter_text.lower() in acc.lower()] if filter_text else rolling_accounts
            for account in filtered_accounts:
                accounts_listbox.insert(tk.END, account)
            
            # Re-select current mapping if it exists in filtered list
            if current_mapping and current_mapping in filtered_accounts:

                pass
                try:
                    index = filtered_accounts.index(current_mapping)
                    accounts_listbox.selection_set(index)
                    accounts_listbox.see(index)
                    mapping_var.set(current_mapping)
                    selected_label.config(text=current_mapping)
                except (ValueError, IndexError):
                    pass
        
        # Search functionality
        def on_search_change(*args):
            populate_listbox(search_var.get())
        
        search_var.trace_add('write', on_search_change)
        populate_listbox()
        
        # Ensure current mapping is selected after dialog is fully initialized
        def ensure_selection():
            if current_mapping and current_mapping in rolling_accounts:

                pass
                try:
                    # Get current list content
                    all_items = list(accounts_listbox.get(0, tk.END))
                    if current_mapping in all_items:
                        index = all_items.index(current_mapping)
                        accounts_listbox.selection_clear(0, tk.END)
                        accounts_listbox.selection_set(index)
                        accounts_listbox.see(index)
                        accounts_listbox.activate(index)
                        mapping_var.set(current_mapping)
                        selected_label.config(text=current_mapping)
                except (ValueError, IndexError):
                    pass
        
        # Schedule selection after dialog is shown
        dialog.after(100, ensure_selection)
        
        # Select account from listbox
        def on_account_select(event):
            selection = accounts_listbox.curselection()
            if selection:
                selected_account = accounts_listbox.get(selection[0])
                mapping_var.set(selected_account)
                selected_label.config(text=selected_account)
        
        accounts_listbox.bind("<<ListboxSelect>>", on_account_select)
        accounts_listbox.bind("<Double-Button-1>", lambda e: save_mapping())
        accounts_listbox.bind("<Return>", lambda e: save_mapping())  # Enter key to save mapping
        
        # Enable focus for keyboard navigation
        accounts_listbox.focus_set()
        
        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(10, 0))
        
        def save_mapping():
            new_mapping = mapping_var.get()
            # Update tree view
            new_values = list(values)
            new_values[3] = new_mapping  # Column 3 = "Mapped Account"
            new_values[4] = "Manual"  # Column 4 = "Confidence"
            self.mapping_tree.item(item, values=new_values)
            
            # Update project mappings
            current_project = self.project_manager.get_current_project()
            if current_project and hasattr(current_project, 'mappings'):
                current_project.mappings[account_desc] = {
                    'rolling_account': new_mapping,
                    'confidence': 'Manual',
                    'similarity': 100.0,
                    'user_edited': True
                }
                # Clear source amounts cache when mappings change
                cache_key = f"{current_project.name}:{current_project.source_file_path}:{current_project.source_range}"
                if cache_key in self.source_amounts_cache:
                    del self.source_amounts_cache[cache_key]
                self.project_manager.save_settings()
                
                # Mark mappings as modified for Step 3 indicator
                self.mark_mappings_modified()
            
            dialog.destroy()
        
        ttk.Button(button_frame, text="Save", command=save_mapping).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side=tk.LEFT)
    
    def bulk_edit_mappings(self):
        """Edit multiple mappings at once with search functionality"""
        # Get selected items that exist in the main tree
        selected_items = []
        for item in self.mapping_tree.get_children():
            if item in self.checkbox_states and self.checkbox_states[item]:
                selected_items.append(item)
        
        if not selected_items:
            messagebox.showwarning("Warning", "Please select items to edit.")
            return
        
        # Create bulk edit dialog
        dialog = tk.Toplevel(self.root)
        dialog.title("Bulk Edit Mappings")
        dialog.resizable(True, True)
        
        # Set geometry and center immediately before showing
        width, height = 700, 500
        x = (dialog.winfo_screenwidth() // 2) - (width // 2)
        y = (dialog.winfo_screenheight() // 2) - (height // 2)
        dialog.geometry(f"{width}x{height}+{x}+{y}")
        
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Main container with padding
        main_frame = ttk.Frame(dialog, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Selected accounts section
        accounts_frame = ttk.LabelFrame(main_frame, text=f"Editing {len(selected_items)} selected accounts", padding="10")
        accounts_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
        
        # Scrollable list of selected accounts
        list_frame = ttk.Frame(accounts_frame)
        list_frame.pack(fill=tk.BOTH, expand=True)
        
        accounts_listbox = tk.Listbox(list_frame, font=("Arial", 10), height=8)
        accounts_scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=accounts_listbox.yview)
        accounts_listbox.configure(yscrollcommand=accounts_scrollbar.set)
        
        for item in selected_items:
            values = self.mapping_tree.item(item, "values")
            if len(values) > 1:
                accounts_listbox.insert(tk.END, values[1])
        
        accounts_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        accounts_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Mapping selection frame
        mapping_frame = ttk.LabelFrame(main_frame, text="Map all selected accounts to", padding="10")
        mapping_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Get rolling accounts for selection
        current_project = self.project_manager.get_current_project()
        rolling_accounts = []
        if current_project:

            pass
            try:
                rolling_data = self.extract_account_data("rolling", self.rolling_range_var.get())
                rolling_accounts = sorted(rolling_data) if rolling_data else []
            except:
                pass
        
        # Search bar
        search_frame = ttk.Frame(mapping_frame)
        search_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(search_frame, text="Search rolling accounts:", font=("Arial", 10, "bold")).pack(anchor="w")
        search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=search_var, font=("Arial", 10))
        search_entry.pack(fill=tk.X, pady=(5, 0))
        
        # Rolling accounts listbox
        rolling_frame = ttk.Frame(mapping_frame)
        rolling_frame.pack(fill=tk.BOTH, expand=True)
        
        rolling_listbox = tk.Listbox(rolling_frame, font=("Arial", 10), height=6)
        rolling_scrollbar = ttk.Scrollbar(rolling_frame, orient="vertical", command=rolling_listbox.yview)
        rolling_listbox.configure(yscrollcommand=rolling_scrollbar.set)
        
        # Populate rolling accounts listbox
        all_rolling_accounts = rolling_accounts[:]
        for account in all_rolling_accounts:
            rolling_listbox.insert(tk.END, account)
        
        rolling_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        rolling_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Search functionality
        def on_search_change(*args):
            search_text = search_var.get().lower()
            rolling_listbox.delete(0, tk.END)
            
            for account in all_rolling_accounts:
                if search_text in account.lower():
                    rolling_listbox.insert(tk.END, account)
        
        search_var.trace_add('write', on_search_change)
        
        # Double-click to select
        def on_rolling_select(event):
            selection = rolling_listbox.curselection()
            if selection:
                selected_account = rolling_listbox.get(selection[0])
                save_bulk_mapping(selected_account)
        
        rolling_listbox.bind('<Double-1>', on_rolling_select)
        
        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=(10, 0))
        
        def save_bulk_mapping(selected_mapping=None):
            if selected_mapping is None:
                selection = rolling_listbox.curselection()
                if not selection:
                    messagebox.showwarning("Warning", "Please select a rolling account.")
                    return
                selected_mapping = rolling_listbox.get(selection[0])
            
            # Update all selected items
            current_project = self.project_manager.get_current_project()
            for item in selected_items:
                values = self.mapping_tree.item(item, "values")
                if len(values) > 1:
                    account_desc = values[1]
                    new_values = list(values)
                    new_values[3] = selected_mapping  # Column 3 = "Mapped Account"
                    new_values[4] = "Manual"  # Column 4 = "Confidence"
                    self.mapping_tree.item(item, values=new_values)
                    
                    # Update project mappings
                    if current_project and hasattr(current_project, 'mappings'):
                        current_project.mappings[account_desc] = {
                            'rolling_account': selected_mapping,
                            'confidence': 'Manual',
                            'similarity': 100.0,
                            'user_edited': True
                        }
                        # Clear source amounts cache when mappings change
                        cache_key = f"{current_project.name}:{current_project.source_file_path}:{current_project.source_range}"
                        if cache_key in self.source_amounts_cache:
                            del self.source_amounts_cache[cache_key]
            
            if current_project:
                self.project_manager.save_settings()
            
            # Clear checkboxes only for the items that were edited
            for item in selected_items:
                if item in self.checkbox_states:
                    self.checkbox_states[item] = False
                    # Update visual checkbox
                    values = self.mapping_tree.item(item, "values")
                    if values:
                        new_values = list(values)
                        new_values[0] = "☐"
                        self.mapping_tree.item(item, values=new_values)
            
            # Update selection counter
            self.update_selection_counter()
            dialog.destroy()
            
            # Success message removed - no confirmation popup needed
        
        def save_selected():
            save_bulk_mapping()
        
        ttk.Button(button_frame, text="Save", command=save_selected).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side=tk.LEFT)
        
        # Focus on search entry
        search_entry.focus()
    
    def apply_filter(self, *args):
        """Apply filter to the mapping tree with delay"""
        # Cancel any existing timer
        if self.filter_timer:
            self.root.after_cancel(self.filter_timer)
        
        # Set a new timer to apply filter after 300ms delay
        self.filter_timer = self.root.after(300, self._do_apply_filter)
    
    def _do_apply_filter(self):
        """Actually apply the filter to the mapping tree"""
        current_project = self.project_manager.get_current_project()
        if not current_project or not hasattr(current_project, 'mappings'):

            pass
            return
            
        # Save the current focus widget and cursor position
        current_focus = self.root.focus_get()
        cursor_position = None
        if current_focus == self.filter_entry:
            cursor_position = self.filter_entry.index('insert')
        
        filter_text = self.filter_var.get().lower()
        if not filter_text:
            # Show all mappings, then apply current sort
            self.apply_sort(None)
        else:
            # Filter mappings
            filtered_mappings = {}
            for account, mapping_info in current_project.mappings.items():
                # Search in account description and mapped account
                if (filter_text in account.lower() or 
                    filter_text in mapping_info.get('rolling_account', '').lower()):
                    filtered_mappings[account] = mapping_info
            
            # Apply current sort to filtered results
            self.apply_sort_to_mappings(filtered_mappings)
        
        # Restore focus and cursor position to the filter entry if it had focus
        if current_focus == self.filter_entry:
            self.filter_entry.focus_set()
            if cursor_position is not None:
                self.filter_entry.icursor(cursor_position)
    
    def apply_sort_to_mappings(self, mappings):
        """Apply current sort option to given mappings"""
        sort_option = self.sort_var.get()
        
        if sort_option == "Account Description A-Z":
            sorted_items = sorted(mappings.items(), key=lambda x: x[0].lower())
            sorted_mappings = dict(sorted_items)
        elif sort_option == "Account Description Z-A":
            sorted_items = sorted(mappings.items(), key=lambda x: x[0].lower(), reverse=True)
            sorted_mappings = dict(sorted_items)
        elif sort_option == "Mapped Account A-Z":
            sorted_items = sorted(mappings.items(), 
                                key=lambda x: x[1].get('rolling_account', '').lower())
            sorted_mappings = dict(sorted_items)
        elif sort_option == "Mapped Account Z-A":
            sorted_items = sorted(mappings.items(), 
                                key=lambda x: x[1].get('rolling_account', '').lower(), reverse=True)
            sorted_mappings = dict(sorted_items)
        elif sort_option == "Confidence Level":
            confidence_order = {"High": 1, "Medium": 2, "Manual": 3, "Low": 4, "Error": 5, "N/A": 6}
            sorted_items = sorted(mappings.items(), 
                                key=lambda x: confidence_order.get(x[1].get('confidence', 'N/A'), 99))
            sorted_mappings = dict(sorted_items)
        else:  # "Original Order" or default
            sorted_mappings = mappings
        
        self.populate_mapping_tree(sorted_mappings)
    
    def clear_filter(self):
        """Clear the current filter"""
        self.filter_var.set("")
        self.apply_filter()  # Refresh to show all items
    
    def apply_sort(self, event):
        """Apply sorting to the mapping tree"""
        current_project = self.project_manager.get_current_project()
        if not current_project or not hasattr(current_project, 'mappings'):

            pass
            return
        
        # Check if there's an active filter
        filter_text = self.filter_var.get().lower()
        if filter_text:
            # Apply filter first, then sort
            self.apply_filter()
        else:
            # No filter, sort all mappings
            self.apply_sort_to_mappings(current_project.mappings)
    
    def apply_popup_filter(self, *args):
        """Apply filter to the popup mapping tree with delay"""
        # Cancel any existing timer
        if hasattr(self, 'popup_filter_timer') and self.popup_filter_timer:
            self.root.after_cancel(self.popup_filter_timer)
        
        # Set a new timer to apply filter after 300ms delay
        self.popup_filter_timer = self.root.after(300, self._do_apply_popup_filter)
    
    def _do_apply_popup_filter(self):
        """Actually apply the filter to the popup mapping tree"""
        if not hasattr(self, 'popup_tree') or not self.popup_tree:

            pass
            return
            
        current_project = self.project_manager.get_current_project()
        if not current_project or not hasattr(current_project, 'mappings'):

            pass
            return
            
        # Save the current focus widget and cursor position
        current_focus = self.root.focus_get()
        cursor_position = None
        if hasattr(self, 'popup_filter_entry') and current_focus == self.popup_filter_entry:
            cursor_position = self.popup_filter_entry.index('insert')
        
        filter_text = self.popup_filter_var.get().lower()
        if not filter_text:
            # Show all mappings, then apply current sort
            self.apply_popup_sort_to_mappings(current_project.mappings)
        else:
            # Filter mappings
            filtered_mappings = {}
            for account, mapping_info in current_project.mappings.items():
                # Search in account description and mapped account
                if (filter_text in account.lower() or 
                    filter_text in mapping_info.get('rolling_account', '').lower()):
                    filtered_mappings[account] = mapping_info
            
            # Apply current sort to filtered results
            self.apply_popup_sort_to_mappings(filtered_mappings)
        
        # Restore focus and cursor position to the filter entry if it had focus
        if hasattr(self, 'popup_filter_entry') and current_focus == self.popup_filter_entry:
            self.popup_filter_entry.focus_set()
            if cursor_position is not None:
                self.popup_filter_entry.icursor(cursor_position)
    
    def clear_popup_filter(self):
        """Clear the popup filter"""
        self.popup_filter_var.set("")
        self.apply_popup_filter()  # Refresh to show all items
    
    def apply_popup_sort(self, event):
        """Apply sorting to the popup mapping tree"""
        if not hasattr(self, 'popup_tree') or not self.popup_tree:

            pass
            return
            
        current_project = self.project_manager.get_current_project()
        if not current_project or not hasattr(current_project, 'mappings'):

            pass
            return
        
        # Check if there's an active filter
        filter_text = self.popup_filter_var.get().lower()
        if filter_text:
            # Apply filter first, then sort
            self.apply_popup_filter()
        else:
            # No filter, sort all mappings
            self.apply_popup_sort_to_mappings(current_project.mappings)
    
    def apply_popup_sort_to_mappings(self, mappings):
        """Apply current sort option to given mappings for popup tree"""
        if not hasattr(self, 'popup_tree') or not self.popup_tree:

            pass
            return
            
        sort_option = self.popup_sort_var.get()
        
        if sort_option == "Account Description A-Z":
            sorted_items = sorted(mappings.items(), key=lambda x: x[0].lower())
            sorted_mappings = dict(sorted_items)
        elif sort_option == "Account Description Z-A":
            sorted_items = sorted(mappings.items(), key=lambda x: x[0].lower(), reverse=True)
            sorted_mappings = dict(sorted_items)
        elif sort_option == "Mapped Account A-Z":
            sorted_items = sorted(mappings.items(), 
                                key=lambda x: x[1].get('rolling_account', '').lower())
            sorted_mappings = dict(sorted_items)
        elif sort_option == "Mapped Account Z-A":
            sorted_items = sorted(mappings.items(), 
                                key=lambda x: x[1].get('rolling_account', '').lower(), reverse=True)
            sorted_mappings = dict(sorted_items)
        elif sort_option == "Confidence Level":
            confidence_order = {"High": 1, "Medium": 2, "Manual": 3, "Low": 4, "Error": 5, "N/A": 6}
            sorted_items = sorted(mappings.items(), 
                                key=lambda x: confidence_order.get(x[1].get('confidence', 'N/A'), 99))
            sorted_mappings = dict(sorted_items)
        else:  # "Original Order" or default
            sorted_mappings = mappings
        
        self.populate_popup_mapping_tree(sorted_mappings)
    
    def populate_popup_mapping_tree(self, mappings):
        """Populate the popup mapping tree with given mappings"""
        if not hasattr(self, 'popup_tree') or not self.popup_tree:

            pass
            return
            
        # Clear existing items
        for item in self.popup_tree.get_children():
            self.popup_tree.delete(item)
        
        # Store the current checkbox states before clearing
        current_checkboxes = {}
        if hasattr(self, 'checkbox_states'):
            current_checkboxes = self.checkbox_states.copy()
        
        # Get source amounts for all accounts (same as main tree)
        source_amounts = self.get_source_amounts_for_mappings()
        
        # Add mappings to popup tree with proper formatting like main tree
        first_heading = True
        for source_account, mapping_info in mappings.items():
            rolling_account = mapping_info.get('rolling_account', '')
            confidence = mapping_info.get('confidence', 'Low')
            similarity = mapping_info.get('similarity', 0)
            
            # Get source amount for this account
            source_amount = source_amounts.get(source_account, 0)
            formatted_amount = self.format_currency(source_amount) if source_amount else ""
            
            # Check if this is a heading/category or total (same logic as main tree)
            is_heading = self.is_total_or_heading(source_account)
            
            if is_heading:
                # Add spacing above headings (except for the first one)
                if not first_heading:
                    self.popup_tree.insert("", "end", values=(
                        "", "", "", "", ""  # Empty spacer row
                    ), tags=("spacer",))
                first_heading = False
                
                # For headings: no checkbox, bold text, show amount, no mapped account, no confidence
                item = self.popup_tree.insert("", "end", values=(
                    "",  # No checkbox for headings
                    source_account,
                    formatted_amount,  # Show amount for headings/totals
                    "",  # No mapped account for headings/totals
                    ""  # No confidence for headings/totals
                ), tags=("heading",))
                # Don't add to checkbox states for headings
            else:
                # For normal accounts: checkbox, normal text, show amount, mapped account, confidence
                # Preserve checkbox state from current selection
                checkbox_state = "☑" if current_checkboxes.get(source_account, False) else "☐"
                
                # Format confidence display - handle unmapped accounts
                if confidence == "None" or not rolling_account:
                    confidence_display = "Not Mapped"
                else:
                    confidence_display = f"{confidence} ({similarity}%)"
                tag = "normal"
                
                item = self.popup_tree.insert("", "end", values=(
                    checkbox_state,  # Preserve checkbox state
                    source_account,
                    formatted_amount,  # Show source amount
                    rolling_account if rolling_account else "",  # Empty string for unmapped
                    confidence_display
                ), tags=(tag,))
        
        # Configure tag styles to match main tree
        self.popup_tree.tag_configure("heading", font=("Arial", 16, "bold"))
        self.popup_tree.tag_configure("normal", font=("Arial", 14))
        self.popup_tree.tag_configure("spacer", font=("Arial", 8))  # Smaller font for spacer rows
        
        # Update popup selection counter
        self.update_popup_selection_counter()
    
    def update_popup_selection_counter(self):
        """Update the selection counter for popup window"""
        if not hasattr(self, 'popup_selection_label') or not self.popup_selection_label:

            pass
            return
            
        selected_count = 0
        if hasattr(self, 'popup_tree') and self.popup_tree:

            pass
            for item in self.popup_tree.get_children():
                values = self.popup_tree.item(item, "values")
                if values and len(values) > 0 and values[0] == "☑":
                    selected_count += 1
        
        self.popup_selection_label.config(text=f"{selected_count} items selected")
    
    def show_context_menu(self, event):
        """Show context menu on right-click"""
        try:
            self.context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.context_menu.grab_release()
    
    def add_placeholder_text(self):
        """Add placeholder text when no mappings are loaded"""
        self.mapping_tree.insert("", "end", values=("", "No mappings generated yet. Use 'Generate Mappings' button.", "", ""))
    
    def has_account_number(self, account_desc):
        """Check if account description contains account numbers"""
        import re
        account_desc_str = str(account_desc).strip()
        
        # Look for patterns like account numbers at the beginning of the line
        account_patterns = [
            r'^\d{3,}',         # Starts with 3+ consecutive digits (e.g., "7350 Domain / Website")
            r'^\d+-\d+',        # Starts with pattern like 1000-1500
            r'^\d+\.\d+',       # Starts with pattern like 1000.01
            r'^#\d+',           # Starts with pattern like #1001
            r'^GL\s*\d+',       # Starts with pattern like GL 1001
        ]
        
        for pattern in account_patterns:
            if re.search(pattern, account_desc_str):

                pass
                return True
        return False
    
    def is_total_or_heading(self, account_desc):
        """Check if account description is a total, subtotal, or heading"""
        desc_lower = str(account_desc).lower().strip()
        desc_original = str(account_desc).strip()
        
        # If it has an account number, it's NOT a heading/total
        if self.has_account_number(account_desc):

            pass
            return False
        
        # Keywords that indicate totals, headings, or categories
        skip_keywords = [
            'total', 'subtotal', 'sum', 'grand total', 'sub total',
            'heading', 'header', '---', '___', 'section',
            'category', 'group', 'division', 'department total'
        ]
        
        # Check for exact matches or keywords that clearly indicate totals/headings
        for keyword in skip_keywords:
            if keyword in desc_lower:

                pass
                return True
        
        # Common real estate and accounting category headers
        category_headers = [
            # Income categories
            'income', 'revenue', 'gross possible rent', 'gross potential rent',
            'rental income', 'other income', 'miscellaneous income',
            
            # Expense categories  
            'expense', 'expenses', 'operating expense', 'operating expenses',
            'administrative', 'maintenance', 'repairs', 'maintenance & repairs',
            'maintenance and repairs', 'utilities', 'insurance', 'taxes',
            'property taxes', 'management fees', 'advertising', 'legal',
            'professional fees', 'supplies', 'landscaping', 'security',
            'deductions', 'vacancy', 'vacancy allowance', 'credit loss',
            'bad debt', 'concessions',
            
            # Balance sheet categories
            'assets', 'liabilities', 'equity', 'current assets', 'fixed assets',
            'current liabilities', 'long term debt', 'owner equity',
            
            # General categories
            'cost', 'costs', 'fees', 'charges', 'allowances', 'reserves'
        ]
        
        # Check for category headers (exact match or partial match)
        for header in category_headers:
            if desc_lower == header or header in desc_lower:

                pass
                return True
        
        # Check if it's title case or all caps (typical for headers)
        # Title case: Most words start with capital letter
        words = desc_original.split()
        if len(words) >= 2:
            capitalized_words = sum(1 for word in words if word and word[0].isupper())
            # If most words are capitalized, it's likely a header
            if capitalized_words >= len(words) * 0.7:  # 70% or more words capitalized
                return True
        
        # All caps and reasonable length (likely a heading)
        if desc_original.isupper() and 3 < len(desc_original) < 50:

            pass
            return True
        
        # If it doesn't have specific account details and is short, treat as header
        # Account descriptions usually have more specific details
        if len(words) <= 4 and not any(char.isdigit() for char in desc_original):
            # Simple phrases without numbers are likely category headers
            return True
            
        return False
    
    def save_user_mappings(self):
        """Save user mappings for current project - Adapted from main.py"""
        current_project = self.project_manager.get_current_project()
        if not current_project or not hasattr(current_project, 'mappings') or not current_project.mappings:
            messagebox.showwarning("Warning", "No mappings to save.")
            return
        
        filename = filedialog.asksaveasfilename(
            title="Save Mapping File",
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("Mapping files", "*.mapping"), ("All files", "*.*")]
        )
        
        if filename:

        
            pass
            try:
                # Create mapping data structure
                mapping_data = {
                    "version": "2.0",
                    "created_date": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "project_name": current_project.name,
                    "source_file": os.path.basename(self.project_manager.source_workbook_path) if self.project_manager.source_workbook_path else "",
                    "rolling_file": os.path.basename(self.project_manager.rolling_workbook_path) if self.project_manager.rolling_workbook_path else "",
                    "mappings": current_project.mappings
                }
                
                with open(filename, 'w') as f:
                    json.dump(mapping_data, f, indent=2)
                
                # Update the project with the saved mapping file path
                current_project.mapping_file_path = filename
                self.mapping_file_var.set(filename)
                self.project_manager.save_settings()  # Persist the change
                
                self.status_var.set(f"Mappings saved to {os.path.basename(filename)}")
                messagebox.showinfo("Success", f"Mappings saved successfully to:\n{filename}")
                
            except Exception as e:
                messagebox.showerror("Error", f"Error saving mappings: {str(e)}")
    
    def reset_mappings(self):
        """Reset mappings for current project - Adapted from main.py"""
        current_project = self.project_manager.get_current_project()
        if not current_project:

            pass
            return
        
        result = messagebox.askyesno(
            "Reset Mappings",
            "Are you sure you want to reset all mappings for this project?\n\n"
            "This will clear all current mappings and return to Step 1.\n"
            "This action cannot be undone."
        )
        if result:

            pass
            if hasattr(current_project, 'mappings'):
                current_project.mappings.clear()
            self.populate_mapping_tree({})
            self.update_ui_state()  # Disable Step 3 buttons when mappings reset
            self.project_manager.save_settings()
            self.status_var.set("Mappings reset successfully.")
            messagebox.showinfo("Reset Complete", "All mappings have been reset.")
    
    def load_range_settings(self):
        """Load saved range settings"""
        try:
            settings_file = "range_settings.json"
            if os.path.exists(settings_file):

                pass
                with open(settings_file, 'r') as f:
                    settings = json.load(f)
                    self.source_range_var.set(settings.get('source_range', 'A8:F200'))
                    self.rolling_range_var.set(settings.get('rolling_range', 'A1:A100'))
                    # Note: rolling_sheet not loaded (removed per user request)
            else:
                # Set default values
                self.source_range_var.set('A8:F200')
                self.rolling_range_var.set('A1:A100')
        except Exception as e:
            self.source_range_var.set('A8:F200')
            self.rolling_range_var.set('A1:A100')
    
    def get_default_source_range(self):
        """Get default source range from global settings or fallback"""
        try:
            settings_file = "range_settings.json"
            if os.path.exists(settings_file):

                pass
                with open(settings_file, 'r') as f:
                    settings = json.load(f)
                    return settings.get('source_range', 'A8:F200')
            else:

                pass
                return 'A8:F200'
        except Exception as e:

            pass
            return 'A8:F200'
    
    def get_default_rolling_range(self):
        """Get default rolling range from global settings or fallback"""
        try:
            settings_file = "range_settings.json"
            if os.path.exists(settings_file):

                pass
                with open(settings_file, 'r') as f:
                    settings = json.load(f)
                    return settings.get('rolling_range', 'A1:A100')
            else:

                pass
                return 'A1:A100'
        except Exception as e:

            pass
            return 'A1:A100'
    
    def load_ui_state(self, project):
        """Load UI state from project data"""
        if not project or not hasattr(project, 'ui_state'):

            pass
            return
        
        ui_state = project.ui_state
        
        # Load filter state
        if hasattr(self, 'filter_var') and ui_state.get('filter_value'):
            self.filter_var.set(ui_state['filter_value'])
        
        # Load sort state  
        if hasattr(self, 'sort_var') and ui_state.get('sort_value'):
            self.sort_var.set(ui_state['sort_value'])
        
        # Load zoom level
        if ui_state.get('zoom_level') and ui_state['zoom_level'] != 1.0:
            zoom_level = ui_state['zoom_level']
            # Find closest zoom level index
            closest_idx = min(range(len(self.zoom_levels)), 
                             key=lambda i: abs(self.zoom_levels[i] - zoom_level))
            self.current_zoom_index = closest_idx
            self.current_zoom_level = self.zoom_levels[closest_idx]
            self.apply_zoom()
        
        # Load checkbox states
        if ui_state.get('checkbox_states'):
            self.checkbox_states = ui_state['checkbox_states'].copy()
    
    def save_ui_state(self, project):
        """Save current UI state to project data"""
        if not project:

            pass
            return
        
        if not hasattr(project, 'ui_state'):
            project.ui_state = {}
        
        # Save filter state
        if hasattr(self, 'filter_var'):
            project.ui_state['filter_value'] = self.filter_var.get()
        
        # Save sort state
        if hasattr(self, 'sort_var'):
            project.ui_state['sort_value'] = self.sort_var.get()
        
        # Save zoom level
        project.ui_state['zoom_level'] = self.current_zoom_level
        
        # Save checkbox states
        project.ui_state['checkbox_states'] = self.checkbox_states.copy()
        
        # Save to persistence
        self.project_manager.save_settings()
    
    def load_step4_data(self, project):
        """Load Step 3 data from project and restore preview table"""
        if not project:
            # Clear Step 3 UI if no project
            self.clear_step4_ui()
            return
        
        # Load Step 3 data if it exists
        if (hasattr(project, 'aggregated_data') and project.aggregated_data and
            hasattr(project, 'preview_data') and project.preview_data):

            pass
            
            # Restore the Step 3 preview table
            self.populate_preview_table(project.preview_data)
            
            # Enable Step 3 buttons based on completion state
            if hasattr(project, 'step4_completed') and project.step4_completed:

                pass
                if hasattr(self, 'finalize_button'):
                    self.finalize_button.config(state="normal")
            else:
                # Enable generate button if mappings exist but Step 3 not completed
                if project.mappings and hasattr(self, 'generate_monthly_button'):
                    self.generate_monthly_button.config(state="normal")
            
            # Update Step 3 status if we have data
            if hasattr(self, 'step4_status_var') and project.target_month:
                num_categories = len(project.aggregated_data)
                if hasattr(project, 'step4_completed') and project.step4_completed:
                    status_msg = f"Step 3 completed. Generated preview for {num_categories} categories. Target month: {project.target_month}"
                else:
                    status_msg = f"Generated preview for {num_categories} categories. Target month: {project.target_month}"
                self.step4_status_var.set(status_msg)
        else:
            # Clear Step 3 UI if no data
            self.clear_step4_ui()
    
    def clear_step4_ui(self):
        """Clear Step 3 UI elements"""
        # Clear preview table
        if hasattr(self, 'preview_tree'):

            pass
            for item in self.preview_tree.get_children():
                self.preview_tree.delete(item)
        
        # Clear status
        if hasattr(self, 'step4_status_var'):
            self.step4_status_var.set("")
        
        # Reset button states - will be handled by update_ui_state()
    
    def save_step4_data(self, project):
        """Save Step 3 data to project"""
        if not project:

            pass
            return
        
        # Save Step 3 completion state
        if hasattr(project, 'step4_completed'):
            project.step4_completed = getattr(self, '_step4_completed', False)
        
        # Data is already saved during the generate_monthly_statement process
        # This method ensures workflow state is updated
        if hasattr(project, 'workflow_state'):
            project.workflow_state['step4_complete'] = project.step4_completed
            project.workflow_state['has_generated_monthly'] = bool(project.monthly_data)
        
        # Save to persistence
        self.project_manager.save_settings()
    
    def clear_current_project(self):
        """Clear all data for the current project"""
        current_project = self.project_manager.get_current_project()
        if not current_project:
            messagebox.showwarning("Warning", "No project selected to clear.")
            return
        
        # Confirm with user
        result = messagebox.askyesno(
            "Clear Project Data",
            f"Are you sure you want to clear all data for project '{current_project.name}'?\n\n"
            "This will remove:\n"
            "• All uploaded files and selections\n"
            "• Source and rolling ranges\n"
            "• Generated mappings and edits\n"
            "• Step 3 monthly data and preview\n"
            "• UI settings (filters, zoom, etc.)\n\n"
            "This action cannot be undone."
        )
        
        if result:
            # Clear all project data
            current_project.clear_all_project_data()
            
            # Clear UI
            self.clear_ui_for_project()
            
            # Update UI state
            self.update_ui_state()
            
            # Save changes
            self.project_manager.save_settings()
            
            messagebox.showinfo("Project Cleared", f"All data for project '{current_project.name}' has been cleared.")
    
    def clear_all_projects(self):
        """Clear all projects and start fresh"""
        if not self.project_manager.has_projects():
            messagebox.showinfo("No Projects", "No projects to clear.")
            return
        
        # Confirm with user
        result = messagebox.askyesno(
            "Clear All Projects",
            "Are you sure you want to clear ALL projects and start fresh?\n\n"
            "This will remove:\n"
            "• All projects and their data\n"
            "• All file uploads and mappings\n"
            "• All user settings and preferences\n\n"
            "This action cannot be undone."
        )
        
        if result:
            # Clear all projects and rolling workbook path
            self.project_manager.reset_all_projects(preserve_rolling_workbook=False)
            
            # Clear UI completely
            self.clear_ui_completely()
            
            # Update UI state
            self.update_ui_state()
            
            messagebox.showinfo("All Projects Cleared", "All projects and data have been cleared. You can now start fresh.")
    
    def clear_ui_for_project(self, preserve_rolling_file: bool = False):
        """Clear UI elements for current project"""
        # Get current project's preserved ranges or defaults
        current_project = self.project_manager.get_current_project()
        if current_project and current_project.source_range:
            self.source_range_var.set(current_project.source_range)  # Use preserved range
        else:
            self.source_range_var.set("A8:F200")  # Use default only if no preserved range
            
        if current_project and current_project.rolling_range:
            self.rolling_range_var.set(current_project.rolling_range)  # Use preserved range
        else:
            self.rolling_range_var.set("A1:A100")  # Use default only if no preserved range
        
        # Clear file selections
        if hasattr(self, 'source_file_var'):
            self.source_file_var.set("")
        if hasattr(self, 'rolling_file_var') and not preserve_rolling_file:
            self.rolling_file_var.set("")
        if hasattr(self, 'rolling_sheet_var'):
            self.rolling_sheet_var.set("")
        
        # Clear mappings
        self.populate_mapping_tree({})
        
        # Clear Step 3 preview
        if hasattr(self, 'preview_tree'):

            pass
            for item in self.preview_tree.get_children():
                self.preview_tree.delete(item)
        
        # Clear status messages
        self.status_var.set("Project cleared. Please upload files to begin.")
        if hasattr(self, 'step4_status_var'):
            self.step4_status_var.set("")
        
        # Reset filters and sorting
        if hasattr(self, 'filter_var'):
            self.filter_var.set("")
        if hasattr(self, 'sort_var'):
            self.sort_var.set("")
        
        # Clear checkboxes
        self.checkbox_states.clear()
    
    def clear_ui_completely(self):
        """Clear UI completely for fresh start"""
        # Clear project selection
        self.project_var.set("")
        
        # Clear everything including rolling file for Start Fresh workflow
        self.clear_ui_for_project(preserve_rolling_file=False)
        
        # Update project menu
        self.refresh_project_menu()
        
        # Reset zoom to default
        self.current_zoom_index = 2  # Default to 1.0
        self.current_zoom_level = 1.0
        self.apply_zoom()
    
    def save_range_settings(self, *args):
        """Save range settings when they change"""
        # Skip saving during project data loading to prevent interference
        if getattr(self, '_loading_project_data', False):

            pass
            return
            
        try:
            # Save to current project
            current_project = self.project_manager.get_current_project()
            if current_project:
                old_rolling_range = getattr(current_project, 'rolling_range', '')
                current_project.source_range = self.source_range_var.get().strip()
                current_project.rolling_range = self.rolling_range_var.get().strip()
                
                # Clear rolling accounts cache if rolling range changed
                if old_rolling_range != current_project.rolling_range:
                    cache_key = f"{current_project.name}:{old_rolling_range}"
                    if cache_key in self.rolling_accounts_cache:
                        del self.rolling_accounts_cache[cache_key]
                
                # Also memorize ranges for the current sheet
                if current_project.rolling_sheet:
                    # Initialize sheet_ranges if needed
                    if not hasattr(current_project, 'sheet_ranges'):
                        current_project.sheet_ranges = {}
                    
                    # Save ranges for current sheet
                    current_project.sheet_ranges[current_project.rolling_sheet] = {
                        'source': self.source_range_var.get().strip(),
                        'rolling': self.rolling_range_var.get().strip()
                    }
                
                # Store in persistent range memory
                self.project_manager.store_project_ranges(
                    current_project.name,
                    current_project.source_range,
                    current_project.rolling_range,
                    current_project.sheet_ranges
                )
                
                self.project_manager.save_settings()
            
            # Also save to local settings file
            settings = {
                'source_range': self.source_range_var.get(),
                'rolling_range': self.rolling_range_var.get()
                # Note: rolling_sheet not saved (removed per user request)
            }
            with open("range_settings.json", 'w') as f:
                json.dump(settings, f, indent=2)
        except Exception as e:
            pass  # Debug output removed
    
    def check_and_add_new_accounts(self, silent_mode=True):
        """Check for new accounts and add them without regenerating existing mappings"""
        current_project = self.project_manager.get_current_project()
        if not current_project or not current_project.source_range:
            return
        
        try:
            # Extract current source accounts from file
            source_data = self.extract_account_data("source", current_project.source_range)
            
            # Get existing mapped accounts
            existing_accounts = set(current_project.mappings.keys()) if current_project.mappings else set()
            
            # Find new accounts not in existing mappings
            new_accounts = [acc for acc in source_data if acc not in existing_accounts]
            
            if new_accounts:
                print(f"\n🔍 Found {len(new_accounts)} new accounts to add:")
                for acc in new_accounts:
                    if '8540' in str(acc):
                        print(f"  + {acc} (NEW ACCOUNT)")
                    else:
                        print(f"  + {acc}")
                
                # Get rolling data for matching
                rolling_data = self.extract_account_data("rolling", current_project.rolling_range)
                
                # Create mappings only for new accounts
                new_mappings = self.create_intelligent_mappings(new_accounts, rolling_data, {})
                
                # Merge with existing mappings while preserving source file order
                if current_project.mappings:
                    # Create ordered dict with all accounts in source file order
                    from collections import OrderedDict
                    ordered_mappings = OrderedDict()
                    
                    # Add all accounts in the order they appear in source_data
                    for account in source_data:
                        if account in current_project.mappings:
                            # Use existing mapping
                            ordered_mappings[account] = current_project.mappings[account]
                        elif account in new_mappings:
                            # Use new mapping
                            ordered_mappings[account] = new_mappings[account]
                    
                    current_project.mappings = ordered_mappings
                else:
                    current_project.mappings = new_mappings
                
                # Clear source amounts cache to get fresh amounts
                cache_key = f"{current_project.name}:{current_project.source_file_path}:{current_project.source_range}"
                if cache_key in self.source_amounts_cache:
                    del self.source_amounts_cache[cache_key]
                
                # Save updated mappings
                self.project_manager.save_settings()
                
                # Update UI
                self.populate_mapping_tree(current_project.mappings)
                self.update_ui_state()
                
                print(f"  ✅ Added {len(new_accounts)} new accounts, total mappings: {len(current_project.mappings)}")
            else:
                if not silent_mode:
                    print("  ✅ No new accounts found - all accounts already mapped")
                    
        except Exception as e:
            print(f"  ⚠️ Error checking for new accounts: {str(e)}")
    
    def generate_mappings(self, silent_mode=False):
        """Generate automatic mappings between source and rolling accounts with confidence levels"""
        # IMPORTANT: Get the actual project from project manager, not a copy
        current_project = self.project_manager.get_current_project()
        if not current_project:

            pass
            if not silent_mode:
                messagebox.showerror("Error", "Please select a project first.")
            return
        
        source_range = self.source_range_var.get().strip()
        rolling_range = self.rolling_range_var.get().strip()
        
        # Clear source amounts cache to ensure fresh extraction
        # This ensures amounts for new accounts like 8540 are extracted
        if current_project and current_project.source_file_path:
            cache_key = f"{current_project.name}:{current_project.source_file_path}:{source_range}"
            if cache_key in self.source_amounts_cache:
                del self.source_amounts_cache[cache_key]
                print(f"  🔄 Cleared source amounts cache for fresh extraction")
        
        # Also clear alternative cache key format
        if current_project and self.project_manager.source_workbook_path:
            cache_key2 = f"{current_project.name}:{self.project_manager.source_workbook_path}:{source_range}"
            if cache_key2 in self.source_amounts_cache:
                del self.source_amounts_cache[cache_key2]
        
        if not source_range or not rolling_range:

        
            pass
            if not silent_mode:
                messagebox.showerror("Error", "Please specify both source and rolling ranges.")
            return
        
        try:
            # Save ranges to current project first
            current_project.source_range = source_range
            current_project.rolling_range = rolling_range
            
            # Ensure project has source file path (fixes issue after project clearing)
            if not current_project.source_file_path and self.project_manager.source_workbook_path:
                current_project.source_file_path = self.project_manager.source_workbook_path
            
            # Traditional mapping approach
            source_data = self.extract_account_data("source", source_range)
            rolling_data = self.extract_account_data("rolling", rolling_range)
            
            # Debug: Print source accounts to check if 8540 is included
            print(f"\n📋 Extracted {len(source_data)} source accounts from range {source_range}")
            for account in source_data:
                if '8540' in str(account):
                    print(f"  ✅ Found account: {account}")
            if not any('8540' in str(account) for account in source_data):
                print(f"  ❌ Account 8540 NOT found in extracted source data!")
            
            # Cache rolling accounts for performance improvement in edit dialogs
            if rolling_data:
                cache_key = f"{current_project.name}:{rolling_range}"
                self.rolling_accounts_cache[cache_key] = sorted(rolling_data)
            
            if not source_data or not rolling_data:

            
                pass
                if not silent_mode:
                    messagebox.showerror("Error", "No account data found in specified ranges.")
                return
            
            # Check for existing mappings to preserve user edits and apply to new accounts
            existing_mappings = current_project.mappings if hasattr(current_project, 'mappings') else {}
            
            # Create mappings for ALL source accounts (including new ones without matches)
            mappings = self.create_intelligent_mappings(source_data, rolling_data, existing_mappings)
            
            # Check if current_project is the same object as what's in project manager
            proj_in_manager = self.project_manager.projects.get(current_project.name)
            if proj_in_manager:
                is_same_object = (current_project is proj_in_manager)
                print(f"  🔍 current_project is same object as manager's project: {is_same_object}")
                if not is_same_object:
                    print(f"  ⚠️ WARNING: Objects are different!")
                    print(f"  🔍 current_project id: {id(current_project)}")
                    print(f"  🔍 manager project id: {id(proj_in_manager)}")
                    # If objects are different, we must update the one in the manager
                    # because that's what gets saved
                    print(f"  🔄 Updating the project in manager instead of current_project")
                    self.project_manager.projects[current_project.name].mappings = mappings
                    # Also ensure current_project points to the right object
                    self.project_manager.select_project(current_project.name)
                    current_project = self.project_manager.get_current_project()
                    print(f"  ✅ Realigned current_project to manager's object")
                else:
                    # Objects are the same, update once
                    current_project.mappings = mappings
                    print(f"  ✅ Updated mappings ({len(mappings)} total)")
            else:
                print(f"  ⚠️ Project '{current_project.name}' not found in manager!")
                # This shouldn't happen but if it does, add it
                self.project_manager.projects[current_project.name] = current_project
                current_project.mappings = mappings
            
            # Update UI
            self.populate_mapping_tree(mappings)
            self.update_ui_state()  # Enable Step 3 buttons when mappings generated
            
            # Update status
            total_mappings = len(mappings)
            high_confidence = len([m for m in mappings.values() if m.get('confidence', '') == 'High'])
            medium_confidence = len([m for m in mappings.values() if m.get('confidence', '') == 'Medium'])
            low_confidence = len([m for m in mappings.values() if m.get('confidence', '') == 'Low'])
            
            status_msg = f"Generated {total_mappings} mappings: {high_confidence} High, {medium_confidence} Medium, {low_confidence} Low confidence"
            self.status_var.set(status_msg)
            
            # Save settings
            print(f"  💾 Saving {len(current_project.mappings)} mappings for {current_project.name}")
            self.project_manager.save_settings()
            
            # Verify save was successful
            import json
            with open('project_settings.json', 'r') as f:
                saved_data = json.load(f)
            saved_count = len(saved_data['projects'][current_project.name]['mappings'])
            if saved_count != len(current_project.mappings):
                print(f"  ⚠️ WARNING: Mismatch! Tried to save {len(current_project.mappings)} but file has {saved_count}")
            else:
                print(f"  ✅ Verified: {saved_count} mappings successfully saved to disk")
            
            if not silent_mode:
                messagebox.showinfo("Success", f"Generated {total_mappings} account mappings!\n\n"
                                  f"High confidence: {high_confidence}\n"
                                  f"Medium confidence: {medium_confidence}\n"
                                  f"Low confidence: {low_confidence}\n\n"
                                  f"Please review and edit mappings as needed.")
            else:
                # Silent mode - used for auto-refresh on project load
                print(f"  ✅ Auto-refreshed {total_mappings} mappings (including new accounts)")
                              
        except Exception as e:

                              
            pass
            if not silent_mode:
                messagebox.showerror("Error", f"Error generating mappings: {str(e)}")
                import traceback
                traceback.print_exc()
            else:

                pass
    
    def extract_account_data(self, file_type, range_str):
        """Extract account descriptions from specified file and range"""
        current_project = self.project_manager.get_current_project()
        if not current_project:

            pass
            return []
        
        if file_type == "source":
            file_path = self.project_manager.source_workbook_path
            sheet_name = current_project.source_sheet
        else:
            file_path = self.project_manager.rolling_workbook_path
            sheet_name = current_project.rolling_sheet
        
        if not file_path or not sheet_name:

        
            pass
            return []
        
        try:
            import pandas as pd
            df = self._load_excel_with_cache(file_path, sheet_name)
            return self.extract_range_data(df, range_str, include_amounts=False)
        except Exception as e:

            pass
            return []
    
    def create_intelligent_mappings(self, source_accounts, rolling_accounts, existing_mappings=None):
        """Create intelligent mappings with confidence levels based on similarity
        Now includes ALL source accounts, even those without matches"""
        import difflib
        from collections import OrderedDict
        
        mappings = OrderedDict()
        
        for source_account in source_accounts:
            # Check if we have an existing mapping for this account
            if existing_mappings and source_account in existing_mappings:
                # Preserve existing mapping (user may have edited it)
                mappings[source_account] = existing_mappings[source_account]
            else:
                # Create new mapping (either find match or leave empty)
                best_match = None
                best_ratio = 0.0
                confidence = "Low"
                
                # Find best match using fuzzy string matching
                for rolling_account in rolling_accounts:
                    # Calculate similarity ratio
                    ratio = difflib.SequenceMatcher(None, 
                        source_account.lower().strip(), 
                        rolling_account.lower().strip()).ratio()
                    
                    if ratio > best_ratio:
                        best_ratio = ratio
                        best_match = rolling_account
                
                # Determine confidence level based on similarity
                if best_ratio >= 0.8:
                    confidence = "High"
                elif best_ratio >= 0.6:
                    confidence = "Medium"
                elif best_ratio >= 0.4:  # Lower threshold for Low confidence
                    confidence = "Low"
                else:
                    # No good match found - create empty mapping for user to fill
                    confidence = "None"
                    best_match = ""  # Empty string for unmapped accounts
                
                # Store new mapping
                mappings[source_account] = {
                    'rolling_account': best_match if best_match else "",
                    'confidence': confidence,
                    'similarity': round(best_ratio * 100, 1),
                    'user_edited': False
                }
        
        return mappings
    
    @profile_performance("get_source_amounts_for_mappings")
    def get_source_amounts_for_mappings(self):
        """Extract source amounts for each account from the target month column"""
        current_project = self.project_manager.get_current_project()
        if not current_project or not current_project.source_file_path:

            pass
            return {}
        
        # Check cache first
        cache_key = f"{current_project.name}:{current_project.source_file_path}:{current_project.source_range}"
        if cache_key in self.source_amounts_cache:

            pass
            return self.source_amounts_cache[cache_key]
        
        try:
            # Get the current project's monthly data if it exists and is fresh
            if hasattr(current_project, 'monthly_data') and current_project.monthly_data:
                # Use the already extracted monthly data and cache it
                self.source_amounts_cache[cache_key] = current_project.monthly_data
                return current_project.monthly_data
            
            # Extract source amounts for Step 2 display
            # This is needed to show the Source Amount column in the mapping tree
            source_df = self._load_excel_with_cache(current_project.source_file_path, current_project.source_sheet)
            source_amounts = self.extract_monthly_amounts(source_df, current_project)
            
            # Cache the result
            if source_amounts:
                self.source_amounts_cache[cache_key] = source_amounts
            
            return source_amounts
            
        except Exception as e:

            
            pass
            return {}
    
    @profile_performance("populate_mapping_tree")
    def populate_mapping_tree(self, mappings):
        """Populate the mapping tree view with generated mappings with source amounts"""
        # Clear existing items and checkbox states
        for item in self.mapping_tree.get_children():
            self.mapping_tree.delete(item)
        self.checkbox_states.clear()
        
        # Get source amounts for all accounts
        source_amounts = self.get_source_amounts_for_mappings()
        
        # Debug: Check if 8540 has an amount
        for account_key in source_amounts.keys():
            if '8540' in str(account_key):
                print(f"  💰 Found amount for {account_key}: {source_amounts[account_key]}")
        
        # Check mappings for 8540
        accounts_with_8540 = [k for k in mappings.keys() if '8540' in str(k)]
        if accounts_with_8540:
            print(f"  📌 Accounts with 8540 in mappings: {accounts_with_8540}")
            for acc in accounts_with_8540:
                amount = source_amounts.get(acc, 'NO AMOUNT FOUND')
                print(f"    - {acc}: Amount = {amount}")
        
        # Add mappings to tree with spacing above category titles
        first_heading = True
        for source_account, mapping_info in mappings.items():
            rolling_account = mapping_info.get('rolling_account', '')
            confidence = mapping_info.get('confidence', 'Low')
            similarity = mapping_info.get('similarity', 0)
            
            # Get source amount for this account
            source_amount = source_amounts.get(source_account, 0)
            formatted_amount = self.format_currency(source_amount) if source_amount else ""
            
            # Check if this is a heading/category or total
            # Items WITH account numbers are regular accounts (even if they contain keywords)
            # Items WITHOUT account numbers are checked for heading/total keywords
            is_heading = self.is_total_or_heading(source_account)
            
            if is_heading:
                # Add spacing above headings (except for the first one)
                if not first_heading:
                    self.mapping_tree.insert("", "end", values=(
                        "", "", "", "", ""  # Empty spacer row
                    ), tags=("spacer",))
                first_heading = False
                
                # For headings: no checkbox, bold text, show amount, no mapped account, no confidence
                item = self.mapping_tree.insert("", "end", values=(
                    "",  # No checkbox for headings
                    source_account,
                    formatted_amount,  # Show amount for headings/totals
                    "",  # No mapped account for headings/totals
                    ""  # No confidence for headings/totals
                ), tags=("heading",))
                # Don't add to checkbox states for headings
            else:
                # For normal accounts: checkbox, normal text, show amount, mapped account, confidence
                # Format confidence display - handle unmapped accounts
                if confidence == "None" or not rolling_account:
                    confidence_display = "Not Mapped"
                else:
                    confidence_display = f"{confidence} ({similarity}%)"
                tag = "normal"
                
                item = self.mapping_tree.insert("", "end", values=(
                    "☐",  # Unchecked checkbox
                    source_account,
                    formatted_amount,  # Show source amount
                    rolling_account if rolling_account else "",  # Empty string for unmapped
                    confidence_display
                ), tags=(tag,))
                
                # Initialize checkbox state as unchecked
                self.checkbox_states[item] = False
        
        # Configure tag styles (no color coding)
        self.mapping_tree.tag_configure("heading", font=("Arial", 16, "bold"))
        self.mapping_tree.tag_configure("normal", font=("Arial", 14))
        self.mapping_tree.tag_configure("spacer", font=("Arial", 8))  # Smaller font for spacer rows
        
        # Update selection counter
        self.update_selection_counter()
        
        # Set focus to the tree to enable keyboard shortcuts
        self.mapping_tree.focus_set()
    
    def pop_out_mapping_window(self):
        """Pop out the mapping tree into a separate window"""
        if self.is_popped_out:

            pass
            return
            
        # Create popup window
        self.popup_window = tk.Toplevel(self.root)
        
        # Set title with current project name
        current_project = self.project_manager.get_current_project()
        project_name = current_project.name if current_project else "No Project"
        target_month = self.ensure_consistent_target_month(current_project) if current_project else None
        if target_month:
            clean_target_month = self.clean_target_month_text(target_month)
            target_month_text = f" - {clean_target_month}"
        else:
            target_month_text = ""
        self.popup_window.title(f"Step 2: Review & Edit Mappings - {project_name}{target_month_text} - Pop Out")
        self.popup_window.geometry("1200x800")
        self.popup_window.minsize(800, 600)
        
        # Center the popup window
        self.center_popup_window()
        
        # Make it resizable
        self.popup_window.resizable(True, True)
        
        # Create main frame in popup window
        popup_main_frame = ttk.Frame(self.popup_window)
        popup_main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Create header with target month and pop-in button
        header_frame = ttk.Frame(popup_main_frame)
        header_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Add target month label if available
        if target_month:
            clean_target_month = self.clean_target_month_text(target_month)
            target_month_label = ttk.Label(header_frame, text=f"Target Month: {clean_target_month}", 
                                         font=("Arial", 14, "bold"), foreground="darkblue")
            target_month_label.pack(side=tk.TOP, anchor=tk.W, pady=(0, 5))
        
        # Create instruction and button frame
        instruction_frame = ttk.Frame(header_frame)
        instruction_frame.pack(fill=tk.X)
        
        header_label = ttk.Label(instruction_frame, text="📝 Double-click or press Enter to edit mapping • Use checkboxes for bulk editing", 
                               font=("Arial", 12, "bold"))
        header_label.pack(side=tk.LEFT, padx=5)
        
        pop_in_button = ttk.Button(instruction_frame, text="🔳 Pop In", command=self.pop_in_mapping_window, width=12)
        pop_in_button.pack(side=tk.RIGHT, padx=5)
        
        # Add filter and sort controls to popup
        popup_filter_frame = ttk.Frame(popup_main_frame)
        popup_filter_frame.pack(fill=tk.X, pady=(5, 5))
        
        # Filter controls
        ttk.Label(popup_filter_frame, text="🔍 Filter:").pack(side=tk.LEFT, padx=(0, 5))
        self.popup_filter_var = tk.StringVar()
        self.popup_filter_var.trace_add('write', self.apply_popup_filter)
        self.popup_filter_entry = ttk.Entry(popup_filter_frame, textvariable=self.popup_filter_var, width=30)
        self.popup_filter_entry.pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Button(popup_filter_frame, text="Clear", 
                  command=self.clear_popup_filter).pack(side=tk.LEFT, padx=(0, 10))
        
        # Sort controls
        ttk.Label(popup_filter_frame, text="📊 Sort by:").pack(side=tk.LEFT, padx=(10, 5))
        self.popup_sort_var = tk.StringVar(value="Original Order")
        popup_sort_combo = ttk.Combobox(popup_filter_frame, textvariable=self.popup_sort_var, 
                                       values=["Original Order", "Account Description A-Z", "Account Description Z-A", 
                                              "Mapped Account A-Z", "Mapped Account Z-A", "Confidence Level"], 
                                       width=20, state="readonly")
        popup_sort_combo.pack(side=tk.LEFT, padx=(0, 5))
        popup_sort_combo.bind('<<ComboboxSelected>>', self.apply_popup_sort)
        
        # Selection counter for popup
        popup_selection_frame = ttk.Frame(popup_main_frame)
        popup_selection_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.popup_selection_label = ttk.Label(popup_selection_frame, text="0 items selected", font=("Arial", 9), foreground="gray")
        self.popup_selection_label.pack(side=tk.LEFT)
        
        # Create content frame for tree and scrollbar
        content_frame = ttk.Frame(popup_main_frame)
        content_frame.pack(fill=tk.BOTH, expand=True)
        content_frame.columnconfigure(0, weight=1)
        content_frame.rowconfigure(0, weight=1)
        
        # Create new tree view for popup (clone of original)
        columns = ("Select", "Account Description", "Source Amount", "Mapped Account", "Confidence")
        self.popup_tree = ttk.Treeview(content_frame, columns=columns, show="headings")
        
        # Configure same style as original
        style = ttk.Style()
        style.configure("Large.Treeview", font=("Arial", 14))
        self.popup_tree.configure(style="Large.Treeview")
        
        # Configure columns with same settings
        self.popup_tree.heading("Select", text="☐", anchor=tk.W)
        self.popup_tree.heading("Account Description", text="Account Description", anchor=tk.W)
        self.popup_tree.heading("Source Amount", text="Source Amount", anchor=tk.W)
        self.popup_tree.heading("Mapped Account", text="Mapped Account", anchor=tk.W)
        self.popup_tree.heading("Confidence", text="Confidence", anchor=tk.W)
        
        self.popup_tree.column("Select", width=80, anchor="center")
        self.popup_tree.column("Account Description", width=500)
        self.popup_tree.column("Source Amount", width=150, anchor="w")
        self.popup_tree.column("Mapped Account", width=400)
        self.popup_tree.column("Confidence", width=120)
        
        # Configure same tags
        self.popup_tree.tag_configure("heading", font=("Arial", 16, "bold"))
        self.popup_tree.tag_configure("normal", font=("Arial", 14))
        
        # Create scrollbar for popup tree
        popup_scrollbar = ttk.Scrollbar(content_frame, orient="vertical", command=self.popup_tree.yview)
        self.popup_tree.configure(yscrollcommand=popup_scrollbar.set)
        
        # Position tree and scrollbar
        self.popup_tree.grid(row=0, column=0, sticky="nsew", padx=(0, 5))
        popup_scrollbar.grid(row=0, column=1, sticky="ns")
        
        # Copy all data from original tree to popup tree
        self.sync_tree_data(self.mapping_tree, self.popup_tree, copy_checkboxes=True)
        
        # Bind events to popup tree (same as original)
        self.popup_tree.bind("<Double-1>", self.edit_mapping_popup)
        self.popup_tree.bind("<Return>", self.edit_mapping_popup)  # Enter key for editing
        self.popup_tree.bind("<Button-1>", self.on_tree_click_popup)
        self.popup_tree.bind("<Up>", self.on_arrow_key_navigation_popup)  # Up arrow key navigation
        self.popup_tree.bind("<Down>", self.on_arrow_key_navigation_popup)  # Down arrow key navigation
        self.popup_tree.bind("<space>", self.on_space_key_toggle_popup)  # Space bar to toggle checkbox
        self.popup_tree.bind("<Key-space>", self.on_space_key_toggle_popup)  # Alternative space binding
        self.popup_tree.bind("<FocusIn>", lambda e: None)  # Enable keyboard focus
        
        # Platform-specific right-click binding for popup
        if platform.system() == "Darwin":  # macOS
            self.popup_tree.bind("<Button-2>", self.show_context_menu_popup)
            self.popup_tree.bind("<Control-Button-1>", self.show_context_menu_popup)
        else:  # Windows/Linux
            self.popup_tree.bind("<Button-3>", self.show_context_menu_popup)
        
        # Hide original tree
        self.mapping_tree.grid_remove()
        
        # Update button state
        self.popup_button.config(text="🔲 Popped Out", state="disabled")
        self.is_popped_out = True
        
        # Handle window close
        self.popup_window.protocol("WM_DELETE_WINDOW", self.pop_in_mapping_window)
        
        # Focus on popup window and tree
        self.popup_window.focus_set()
        self.popup_tree.focus_set()
        
    def pop_in_mapping_window(self):
        """Pop the mapping tree back into the main window"""
        if not self.is_popped_out or not self.popup_window:

            pass
            return
            
        # Sync data back from popup tree to main tree
        if hasattr(self, 'popup_tree'):
            self.sync_tree_data(self.popup_tree, self.mapping_tree, copy_checkboxes=True)
        
        # Show original tree
        self.mapping_tree.grid()
        
        # Restore focus to main tree
        self.mapping_tree.focus_set()
        
        # Update button state
        self.popup_button.config(text="🔲 Pop Out", state="normal")
        self.is_popped_out = False
        
        # Close popup window
        if self.popup_window:
            self.popup_window.destroy()
            self.popup_window = None
        
        # Force update of Step 3 button style after popup changes
        self.update_step4_button_style()
            
        # Clear popup tree reference
        if hasattr(self, 'popup_tree'):
            delattr(self, 'popup_tree')
    
    def clear_popup_checkboxes(self):
        """Clear all checkboxes in the popup window after bulk edit save"""
        if not self.popup_tree:

            pass
            return
            
        # Clear all checkbox states for popup items
        for item in self.popup_tree.get_children():
            values = list(self.popup_tree.item(item, 'values'))
            if len(values) > 0:
                # Change checkbox from ☑ to ☐
                values[0] = "☐"
                self.popup_tree.item(item, values=values)
                
                # Update checkbox state tracking
                if item in self.checkbox_states:
                    self.checkbox_states[item] = False
        
        # Update selection counter if it exists
        if hasattr(self, 'update_selection_counter'):
            self.update_selection_counter()
    
    def get_account_description_from_item(self, tree, item):
        """Get account description from a tree item"""
        values = tree.item(item, "values")
        if len(values) > 1:

            pass
            return values[1]
        return None
        
    def center_popup_window(self):
        """Center the popup window on screen"""
        if not self.popup_window:

            pass
            return
            
        # Update geometry to ensure correct size calculation
        self.popup_window.update_idletasks()
        
        # Get screen dimensions
        screen_width = self.popup_window.winfo_screenwidth()
        screen_height = self.popup_window.winfo_screenheight()
        
        # Get window dimensions
        window_width = 1200
        window_height = 800
        
        # Calculate position
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        
        # Set geometry
        self.popup_window.geometry(f"{window_width}x{window_height}+{x}+{y}")
    
    def sync_tree_data(self, source_tree, target_tree, copy_checkboxes=False):
        """Synchronize data between two tree views"""
        # Clear target tree
        for item in target_tree.get_children():
            target_tree.delete(item)
        
        # Clear checkbox states if we're syncing back to main tree
        if target_tree == self.mapping_tree and source_tree == self.popup_tree:
            # Clean up all old checkbox states from popup tree
            old_items = [item for item in self.checkbox_states.keys() if self.popup_tree.exists(item)]
            for item in old_items:
                del self.checkbox_states[item]
        
        # Copy all items from source to target
        for item in source_tree.get_children():
            values = source_tree.item(item, "values")
            tags = source_tree.item(item, "tags")
            new_item = target_tree.insert("", "end", values=values, tags=tags)
            
            # If copying checkboxes, update checkbox states with new item IDs
            if copy_checkboxes and item in self.checkbox_states:
                # Transfer the checkbox state to the new item
                self.checkbox_states[new_item] = self.checkbox_states[item]
            elif tags and "heading" not in tags and "spacer" not in tags:
                # Initialize checkbox state for non-heading items if not already set
                if values and len(values) > 0 and values[0] in ["☑", "☐"]:
                    self.checkbox_states[new_item] = (values[0] == "☑")
    
    def edit_mapping_popup(self, event):
        """Handle double-click edit in popup window"""
        item = self.popup_tree.selection()[0] if self.popup_tree.selection() else None
        if not item:

            pass
            return
            
        values = self.popup_tree.item(item, "values")
        if len(values) < 2:

            pass
            return
            
        account_desc = values[1]
        current_mapping = values[3] if len(values) > 3 else ""  # Column 3 = "Mapped Account"
        
        # Skip headers and totals using proper detection logic
        if self.is_total_or_heading(account_desc):

            pass
            return
        
        # Create edit dialog (same as original but update popup tree)
        dialog = tk.Toplevel(self.popup_window)  # Parent to popup window
        dialog.title("Edit Mapping")
        dialog.resizable(True, True)
        
        # Set geometry and center immediately before showing
        width, height = 600, 400
        x = (dialog.winfo_screenwidth() // 2) - (width // 2)
        y = (dialog.winfo_screenheight() // 2) - (height // 2)
        dialog.geometry(f"{width}x{height}+{x}+{y}")
        
        dialog.transient(self.popup_window)
        dialog.grab_set()
        
        # Main container with padding
        main_frame = ttk.Frame(dialog, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Account description (read-only)
        desc_frame = ttk.LabelFrame(main_frame, text="Account Description", padding="10")
        desc_frame.pack(fill=tk.X, pady=(0, 15))
        
        ttk.Label(desc_frame, text=account_desc, font=("Arial", 11), 
                 wraplength=550, justify="left").pack(anchor="w")
        
        # Mapping selection frame
        mapping_frame = ttk.LabelFrame(main_frame, text="Map to Rolling Account", padding="10")
        mapping_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
        
        # Search bar
        search_frame = ttk.Frame(mapping_frame)
        search_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(search_frame, text="🔍 Search:", font=("Arial", 10)).pack(side=tk.LEFT, padx=(0, 5))
        search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=search_var, width=40)
        search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        ttk.Button(search_frame, text="Clear", 
                  command=lambda: search_var.set(""), width=8).pack(side=tk.LEFT)
        
        # Get rolling accounts for dropdown - use cache if available
        current_project = self.project_manager.get_current_project()
        rolling_accounts = []
        if current_project:

            pass
            try:
                # Check cache first
                cache_key = f"{current_project.name}:{self.rolling_range_var.get()}"
                if cache_key in self.rolling_accounts_cache:
                    rolling_accounts = self.rolling_accounts_cache[cache_key]
                else:
                    # Fall back to extracting if not cached
                    rolling_data = self.extract_account_data("rolling", self.rolling_range_var.get())
                    rolling_accounts = sorted(rolling_data) if rolling_data else []
                    # Cache for future use
                    if rolling_accounts:
                        self.rolling_accounts_cache[cache_key] = rolling_accounts
            except:
                pass
        
        # Selected mapping display
        mapping_var = tk.StringVar(value=current_mapping)
        selected_frame = ttk.Frame(mapping_frame)
        selected_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(selected_frame, text="Selected Mapping:", font=("Arial", 10, "bold")).pack(anchor="w")
        selected_label = ttk.Label(selected_frame, text=current_mapping or "None selected", 
                                  font=("Arial", 10), foreground="blue")
        selected_label.pack(anchor="w", pady=(2, 0))
        
        # Rolling accounts listbox with scrollbar
        list_frame = ttk.Frame(mapping_frame)
        list_frame.pack(fill=tk.BOTH, expand=True)
        
        accounts_listbox = tk.Listbox(list_frame, font=("Arial", 10), height=10)
        list_scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=accounts_listbox.yview)
        accounts_listbox.configure(yscrollcommand=list_scrollbar.set)
        
        accounts_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        list_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Populate and filter accounts
        def update_accounts_list():
            search_text = search_var.get().lower()
            accounts_listbox.delete(0, tk.END)
            
            for account in rolling_accounts:
                if not search_text or search_text in account.lower():
                    accounts_listbox.insert(tk.END, account)
                    
            # Highlight current mapping if visible
            if current_mapping:

                pass
                try:
                    idx = list(accounts_listbox.get(0, tk.END)).index(current_mapping)
                    accounts_listbox.selection_set(idx)
                    accounts_listbox.see(idx)
                except ValueError:
                    pass
        
        search_var.trace_add('write', lambda *args: update_accounts_list())
        update_accounts_list()
        
        # Ensure current mapping is selected after dialog is fully initialized
        def ensure_selection():
            if current_mapping and current_mapping in rolling_accounts:

                pass
                try:
                    # Get current list content
                    all_items = list(accounts_listbox.get(0, tk.END))
                    if current_mapping in all_items:
                        index = all_items.index(current_mapping)
                        accounts_listbox.selection_clear(0, tk.END)
                        accounts_listbox.selection_set(index)
                        accounts_listbox.see(index)
                        accounts_listbox.activate(index)
                        mapping_var.set(current_mapping)
                        selected_label.config(text=current_mapping)
                except (ValueError, IndexError):
                    pass
        
        # Schedule selection after dialog is shown
        dialog.after(100, ensure_selection)
        
        def on_account_select(event):
            selection = accounts_listbox.curselection()
            if selection:
                selected_account = accounts_listbox.get(selection[0])
                mapping_var.set(selected_account)
                selected_label.config(text=selected_account)
        
        accounts_listbox.bind('<<ListboxSelect>>', on_account_select)
        accounts_listbox.bind("<Double-Button-1>", lambda e: save_mapping())
        accounts_listbox.bind("<Return>", lambda e: save_mapping())  # Enter key to save mapping
        
        # Enable focus for keyboard navigation
        accounts_listbox.focus_set()
        
        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(10, 0))
        
        def save_mapping():
            new_mapping = mapping_var.get()
            # Update popup tree
            new_values = list(values)
            new_values[3] = new_mapping  # Column 3 = "Mapped Account"
            new_values[4] = "Manual"  # Column 4 = "Confidence"
            self.popup_tree.item(item, values=new_values)
            
            # Update project mappings
            current_project = self.project_manager.get_current_project()
            if current_project and hasattr(current_project, 'mappings'):
                current_project.mappings[account_desc] = {
                    'rolling_account': new_mapping,
                    'confidence': 'Manual',
                    'similarity': 100.0,
                    'user_edited': True
                }
                # Clear source amounts cache when mappings change
                cache_key = f"{current_project.name}:{current_project.source_file_path}:{current_project.source_range}"
                if cache_key in self.source_amounts_cache:
                    del self.source_amounts_cache[cache_key]
                self.project_manager.save_settings()
                
                # Mark mappings as modified for Step 3 indicator
                self.mark_mappings_modified()
            
            dialog.destroy()
        
        ttk.Button(button_frame, text="Save", command=save_mapping).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side=tk.LEFT)
    
    def on_tree_click_popup(self, event):
        """Handle single click events in popup window"""
        # Ensure the popup tree has focus for keyboard events
        self.popup_tree.focus_set()
        
        region = self.popup_tree.identify_region(event.x, event.y)
        
        if region == "heading":
            # Handle header clicks
            column = self.popup_tree.identify_column(event.x)
            if column == "#1":  # Clicked on Select column header
                self.toggle_all_selection_popup()
        
        elif region == "cell":
            # Get the item and column clicked
            item = self.popup_tree.identify_row(event.y)
            column = self.popup_tree.identify_column(event.x)
            
            # Only toggle if checkbox column (column #1) and it's not a header row
            if column == "#1" and item:
                # Get item tags to check if it's a heading
                tags = self.popup_tree.item(item, "tags")
                if tags and "heading" in tags:

                    pass
                    return  # Skip headings
                
                # Initialize checkbox state if not exists
                if item not in self.checkbox_states:
                    self.checkbox_states[item] = False
                
                # Toggle checkbox state
                current_state = self.checkbox_states.get(item, False)
                new_state = not current_state
                self.checkbox_states[item] = new_state
                
                # Update checkbox display
                values = self.popup_tree.item(item, "values")
                new_values = list(values)
                new_values[0] = "☑" if new_state else "☐"
                self.popup_tree.item(item, values=new_values)
                
                # Update selection counter
                self.update_selection_counter()
    
    def toggle_all_selection_popup(self):
        """Toggle all checkboxes in popup window"""
        # Determine if we should select all or deselect all
        selected_count = sum(1 for selected in self.checkbox_states.values() if selected)
        select_all = selected_count < len(self.checkbox_states) / 2
        
        # Update all checkbox states
        for item in self.checkbox_states:
            self.checkbox_states[item] = select_all
            
            # Update checkbox display
            values = self.popup_tree.item(item, "values")
            new_values = list(values)
            new_values[0] = "☑" if select_all else "☐"
            self.popup_tree.item(item, values=new_values)
        
        # Update selection counter
        self.update_selection_counter()
    
    def show_context_menu_popup(self, event):
        """Show context menu for popup window"""
        # Create popup-specific context menu
        popup_context_menu = tk.Menu(self.popup_window, tearoff=0)
        popup_context_menu.add_command(label="Edit Selected Categories", command=self.bulk_edit_mappings_popup)
        popup_context_menu.add_separator()
        popup_context_menu.add_command(label="Select All", command=self.select_all_items_popup)
        popup_context_menu.add_command(label="Deselect All", command=self.deselect_all_items_popup)
        popup_context_menu.post(event.x_root, event.y_root)
    
    def bulk_edit_mappings_popup(self):
        """Edit multiple mappings at once in popup window"""
        # Get selected items that exist in popup tree
        selected_items = []
        for item in self.popup_tree.get_children():
            if item in self.checkbox_states and self.checkbox_states[item]:
                selected_items.append(item)
        
        if not selected_items:
            messagebox.showwarning("Warning", "Please select items to edit.")
            return
            
        # Use the same bulk edit dialog but work with popup tree
        self._bulk_edit_dialog_for_tree(selected_items, self.popup_tree)
    
    def _bulk_edit_dialog_for_tree(self, selected_items, target_tree):
        """Generic bulk edit dialog that works with any tree"""
        # Create bulk edit dialog
        dialog = tk.Toplevel(self.popup_window if target_tree == self.popup_tree else self.root)
        dialog.title("Bulk Edit Mappings")
        dialog.resizable(True, True)
        dialog.transient(self.popup_window if target_tree == self.popup_tree else self.root)
        dialog.grab_set()
        
        # Calculate center position immediately
        window_width = 800
        window_height = 600
        x = (dialog.winfo_screenwidth() // 2) - (window_width // 2)
        y = (dialog.winfo_screenheight() // 2) - (window_height // 2)
        
        # Set geometry with center position from the start
        dialog.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # Main container
        main_frame = ttk.Frame(dialog, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Title
        title_label = ttk.Label(main_frame, text=f"Bulk Edit {len(selected_items)} Selected Mappings", 
                               font=("Arial", 14, "bold"))
        title_label.pack(pady=(0, 15))
        
        # Selected accounts section
        accounts_frame = ttk.LabelFrame(main_frame, text=f"Selected accounts ({len(selected_items)})", padding="10")
        accounts_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Scrollable list of selected accounts
        list_frame = ttk.Frame(accounts_frame)
        list_frame.pack(fill=tk.X)
        
        accounts_listbox = tk.Listbox(list_frame, font=("Arial", 10), height=6)
        accounts_scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=accounts_listbox.yview)
        accounts_listbox.configure(yscrollcommand=accounts_scrollbar.set)
        
        for item in selected_items:
            values = target_tree.item(item, "values")
            if len(values) > 1:
                accounts_listbox.insert(tk.END, values[1])
        
        accounts_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        accounts_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Account selection frame
        mapping_frame = ttk.LabelFrame(main_frame, text="Select Target Rolling Account", padding="10")
        mapping_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
        
        # Search functionality
        search_frame = ttk.Frame(mapping_frame)
        search_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(search_frame, text="🔍 Search:", font=("Arial", 10)).pack(side=tk.LEFT, padx=(0, 5))
        search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=search_var, width=40)
        search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        ttk.Button(search_frame, text="Clear", 
                  command=lambda: search_var.set(""), width=8).pack(side=tk.LEFT)
        
        # Get rolling accounts - use cache if available
        current_project = self.project_manager.get_current_project()
        rolling_accounts = []
        if current_project:

            pass
            try:
                # Check cache first
                cache_key = f"{current_project.name}:{self.rolling_range_var.get()}"
                if cache_key in self.rolling_accounts_cache:
                    rolling_accounts = self.rolling_accounts_cache[cache_key]
                else:
                    # Fall back to extracting if not cached
                    rolling_data = self.extract_account_data("rolling", self.rolling_range_var.get())
                    rolling_accounts = sorted(rolling_data) if rolling_data else []
                    # Cache for future use
                    if rolling_accounts:
                        self.rolling_accounts_cache[cache_key] = rolling_accounts
            except:
                pass
        
        # Listbox for account selection
        list_frame = ttk.Frame(mapping_frame)
        list_frame.pack(fill=tk.BOTH, expand=True)
        
        accounts_listbox = tk.Listbox(list_frame, font=("Arial", 10), height=15)
        list_scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=accounts_listbox.yview)
        accounts_listbox.configure(yscrollcommand=list_scrollbar.set)
        
        accounts_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        list_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Populate accounts with search filter
        def update_accounts_list():
            search_text = search_var.get().lower()
            accounts_listbox.delete(0, tk.END)
            
            for account in rolling_accounts:
                if not search_text or search_text in account.lower():
                    accounts_listbox.insert(tk.END, account)
        
        search_var.trace_add('write', lambda *args: update_accounts_list())
        update_accounts_list()
        
        # Selected mapping display
        selected_mapping_var = tk.StringVar()
        selected_frame = ttk.Frame(mapping_frame)
        selected_frame.pack(fill=tk.X, pady=(10, 0))
        
        ttk.Label(selected_frame, text="Selected Mapping:", font=("Arial", 10, "bold")).pack(anchor="w")
        selected_label = ttk.Label(selected_frame, text="None selected", 
                                  font=("Arial", 10), foreground="blue")
        selected_label.pack(anchor="w", pady=(2, 0))
        
        def on_account_select(event):
            selection = accounts_listbox.curselection()
            if selection:
                selected_account = accounts_listbox.get(selection[0])
                selected_mapping_var.set(selected_account)
                selected_label.config(text=selected_account)
        
        accounts_listbox.bind('<<ListboxSelect>>', on_account_select)
        accounts_listbox.bind("<Return>", lambda e: apply_bulk_edit())  # Enter key to apply bulk edit
        
        # Enable focus for keyboard navigation
        accounts_listbox.focus_set()
        
        # Buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(10, 0))
        
        def apply_bulk_edit():
            selected_mapping = selected_mapping_var.get()
            if not selected_mapping:
                messagebox.showwarning("Warning", "Please select a rolling account.")
                return
            
            # Apply to all selected items
            current_project = self.project_manager.get_current_project()
            for item in selected_items:
                if target_tree.exists(item):
                    values = target_tree.item(item, "values")
                    if len(values) > 1:
                        account_desc = values[1]
                        new_values = list(values)
                        new_values[3] = selected_mapping  # Column 3 = "Mapped Account"
                        new_values[4] = "Manual"  # Column 4 = "Confidence"
                        target_tree.item(item, values=new_values)
                        
                        # Update project mappings
                        if current_project and hasattr(current_project, 'mappings'):
                            current_project.mappings[account_desc] = {
                                'rolling_account': selected_mapping,
                                'confidence': 'Manual',
                                'similarity': 100.0,
                                'user_edited': True
                            }
            
            # Save settings
            if current_project:
                self.project_manager.save_settings()
                
                # Mark mappings as modified for Step 3 indicator
                self.mark_mappings_modified()
            
            # Clear checkboxes only for the items that were edited
            for item in selected_items:
                if item in self.checkbox_states:
                    self.checkbox_states[item] = False
                    # Update visual checkbox in target tree
                    if target_tree.exists(item):
                        values = target_tree.item(item, "values")
                        if values:
                            new_values = list(values)
                            new_values[0] = "☐"
                            target_tree.item(item, values=new_values)
            
            # If this was a popup edit, sync the changes back to main tree immediately
            if target_tree == self.popup_tree and hasattr(self, 'mapping_tree'):
                self.sync_tree_data(self.popup_tree, self.mapping_tree, copy_checkboxes=True)
            
            # Update selection counter
            self.update_selection_counter()
            
            # Ensure tree focus is maintained
            target_tree.focus_set()
            
            dialog.destroy()
            # Success message removed - no confirmation popup needed
        
        ttk.Button(button_frame, text="Apply to All Selected", 
                  command=apply_bulk_edit).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side=tk.LEFT)
    
    def select_all_items_popup(self):
        """Select all items in popup window"""
        for item in self.popup_tree.get_children():
            # Skip headings and spacers
            tags = self.popup_tree.item(item, "tags")
            if tags and ("heading" in tags or "spacer" in tags):

                pass
                continue
            
            # Initialize checkbox state if not exists
            if item not in self.checkbox_states:
                self.checkbox_states[item] = False
            
            self.checkbox_states[item] = True
            values = self.popup_tree.item(item, "values")
            new_values = list(values)
            new_values[0] = "☑"
            self.popup_tree.item(item, values=new_values)
        self.update_selection_counter()
    
    def deselect_all_items_popup(self):
        """Deselect all items in popup window"""
        for item in self.popup_tree.get_children():
            # Skip headings and spacers
            tags = self.popup_tree.item(item, "tags")
            if tags and ("heading" in tags or "spacer" in tags):

                pass
                continue
            
            if item in self.checkbox_states:
                self.checkbox_states[item] = False
                values = self.popup_tree.item(item, "values")
                new_values = list(values)
                new_values[0] = "☐"
                self.popup_tree.item(item, values=new_values)
        self.update_selection_counter()
    
    def on_arrow_key_navigation(self, event):
        """Handle up/down arrow key navigation in main mapping tree"""
        current_selection = self.mapping_tree.selection()
        all_items = self.mapping_tree.get_children()
        
        if not all_items:

        
            pass
            return "break"
        
        if not current_selection:
            # No selection, select first item
            self.mapping_tree.selection_set(all_items[0])
            self.mapping_tree.focus(all_items[0])
            self.mapping_tree.see(all_items[0])
            return "break"
        
        current_item = current_selection[0]
        current_index = all_items.index(current_item)
        
        if event.keysym == "Up":
            # Move up
            new_index = max(0, current_index - 1)
        elif event.keysym == "Down":
            # Move down
            new_index = min(len(all_items) - 1, current_index + 1)
        else:

            pass
            return "break"
        
        new_item = all_items[new_index]
        self.mapping_tree.selection_set(new_item)
        self.mapping_tree.focus(new_item)
        self.mapping_tree.see(new_item)
        return "break"  # Prevent default TreeView arrow key behavior
    
    def on_arrow_key_navigation_popup(self, event):
        """Handle up/down arrow key navigation in popup mapping tree"""
        current_selection = self.popup_tree.selection()
        all_items = self.popup_tree.get_children()
        
        if not all_items:

        
            pass
            return "break"
        
        if not current_selection:
            # No selection, select first item
            self.popup_tree.selection_set(all_items[0])
            self.popup_tree.focus(all_items[0])
            self.popup_tree.see(all_items[0])
            return "break"
        
        current_item = current_selection[0]
        current_index = all_items.index(current_item)
        
        if event.keysym == "Up":
            # Move up
            new_index = max(0, current_index - 1)
        elif event.keysym == "Down":
            # Move down
            new_index = min(len(all_items) - 1, current_index + 1)
        else:

            pass
            return "break"
        
        new_item = all_items[new_index]
        self.popup_tree.selection_set(new_item)
        self.popup_tree.focus(new_item)
        self.popup_tree.see(new_item)
        return "break"  # Prevent default TreeView arrow key behavior
    
    def on_space_key_toggle(self, event):
        """Handle space bar key to toggle checkbox of selected item in main tree"""
        current_selection = self.mapping_tree.selection()
        if not current_selection:

            pass
            return "break"
        
        item = current_selection[0]
        
        # Skip if it's a heading row (no checkbox to toggle)
        values = self.mapping_tree.item(item, "values")
        if not values or len(values) < 2:

            pass
            return "break"
        
        # Check if this is a heading row by looking at the tag
        tags = self.mapping_tree.item(item, "tags")
        if "heading" in tags:

            pass
            return "break"
        
        # Initialize checkbox state if not exists
        if item not in self.checkbox_states:
            self.checkbox_states[item] = False
        
        # Toggle the checkbox state
        current_state = self.checkbox_states[item]
        new_state = not current_state
        self.checkbox_states[item] = new_state
        
        # Update visual checkbox
        new_values = list(values)
        new_values[0] = "☑" if new_state else "☐"
        self.mapping_tree.item(item, values=new_values)
        
        # Update selection counter
        self.update_selection_counter()
        
        return "break"  # Prevent default space key behavior
    
    def on_space_key_toggle_popup(self, event):
        """Handle space bar key to toggle checkbox of selected item in popup tree"""
        current_selection = self.popup_tree.selection()
        if not current_selection:

            pass
            return "break"
        
        item = current_selection[0]
        
        # Skip if it's a heading row (no checkbox to toggle)
        values = self.popup_tree.item(item, "values")
        if not values or len(values) < 2:

            pass
            return "break"
        
        # Check if this is a heading row by looking at the tag
        tags = self.popup_tree.item(item, "tags")
        if "heading" in tags:

            pass
            return "break"
        
        # Initialize checkbox state if not exists
        if item not in self.checkbox_states:
            self.checkbox_states[item] = False
        
        # Toggle the checkbox state
        current_state = self.checkbox_states[item]
        new_state = not current_state
        self.checkbox_states[item] = new_state
        
        # Update visual checkbox
        new_values = list(values)
        new_values[0] = "☑" if new_state else "☐"
        self.popup_tree.item(item, values=new_values)
        
        # Update selection counter
        self.update_selection_counter()
        
        return "break"  # Prevent default space key behavior
    
    def mark_mappings_modified(self):
        """Mark that mappings have been modified after monthly statement generation"""
        current_project = self.project_manager.get_current_project()
        
        # Check if monthly statement exists (either from session or cached from previous session)
        # Note: monthly_data can be an empty dict {} if no amounts were mapped, but it still counts as "generated"
        has_monthly_data = (current_project and 
                           hasattr(current_project, 'aggregated_data') and
                           current_project.aggregated_data)
        
        if current_project:

        
            pass
            if hasattr(current_project, 'monthly_data'):
                pass
            if hasattr(current_project, 'aggregated_data'):
                pass

                pass
        
        if has_monthly_data and current_project:
            self.mappings_modified_after_generation = True
            # Mark this project as having modified mappings for auto-workflow
            self.projects_with_modified_mappings.add(current_project.name)
            self.update_step4_button_style()
            
            # Update UI to show regeneration indication (without automatic regeneration)
            self.update_ui_state()
        else:
            # No monthly data exists, so no regeneration indication needed
            pass
    
    def update_step4_button_style(self):
        """Update Step 3 button style based on modification state"""
        style = ttk.Style()
        
        if self.mappings_modified_after_generation:
            # Configure a custom style with orange background and bold font
            style.configure("OrangeModified.TButton",
                          background="#FF8C00",
                          foreground="black",
                          borderwidth=3,
                          relief="raised",
                          font=("Arial", 12, "bold"))
            
            # Map all states to ensure text is always visible
            style.map("OrangeModified.TButton",
                     background=[("active", "#FF4500"), ("pressed", "#FF6347"), ("!active", "#FF8C00")],
                     foreground=[("active", "black"), ("pressed", "black"), ("!active", "black")])
            
            # Apply style and change text
            self.generate_monthly_button.configure(
                text="🔄 Regenerate Statement",
                style="OrangeModified.TButton"
            )
            
            # Debug logging
        else:
            # Configure normal style
            style.configure("Normal.TButton",
                          font=("Arial", 10, "normal"))
            
            # Reset to default
            self.generate_monthly_button.configure(
                text="📊 Generate Monthly Statement",
                style="Normal.TButton"
            )
            
        # Force UI update
        self.root.update_idletasks()
        
        # Also update popup button if it exists
        self.update_popup_button_style()
    
    def reset_step4_modification_flag(self):
        """Reset the modification flag when monthly statement is regenerated"""
        self.mappings_modified_after_generation = False
        self.update_step4_button_style()
        self.update_popup_button_style()
    
    def update_popup_button_style(self):
        """Update Step 3 popup generate button style based on modification state"""
        if not hasattr(self, 'step3_popup_generate_button') or not self.step3_popup_generate_button:

            pass
            return
            
        style = ttk.Style()
        
        if self.mappings_modified_after_generation:
            # Configure a custom style with orange background and bold font
            style.configure("PopupOrangeModified.TButton",
                          background="#FF8C00",
                          foreground="black",
                          borderwidth=3,
                          relief="raised",
                          font=("Arial", 12, "bold"))
            
            # Map all states to ensure text is always visible
            style.map("PopupOrangeModified.TButton",
                     background=[("active", "#FF4500"), ("pressed", "#FF6347"), ("!active", "#FF8C00")],
                     foreground=[("active", "black"), ("pressed", "black"), ("!active", "black")])
            
            # Apply style and change text
            self.step3_popup_generate_button.configure(
                text="🔄 Regenerate Statement",
                style="PopupOrangeModified.TButton"
            )
        else:
            # Configure normal style
            style.configure("PopupNormal.TButton",
                          font=("Arial", 10, "normal"))
            
            # Reset to default
            self.step3_popup_generate_button.configure(
                text="📊 Generate Monthly Statement",
                style="PopupNormal.TButton"
            )
    
    def update_popup_finalize_button_state(self):
        """Update Step 3 popup finalize button state based on project data"""
        if not hasattr(self, 'step3_popup_finalize_button') or not self.step3_popup_finalize_button:

            pass
            return
            
        current_project = self.project_manager.get_current_project()
        if not current_project:
            self.step3_popup_finalize_button.config(state="disabled")
            return
        
        # Check if monthly statement has been generated
        has_monthly_data = (hasattr(current_project, 'aggregated_data') and 
                          current_project.aggregated_data)
        
        # Check if Step 3 was previously completed
        step4_completed = hasattr(current_project, 'step4_completed') and current_project.step4_completed
        can_finalize = has_monthly_data or step4_completed
        
        # Update button state
        self.step3_popup_finalize_button.config(state="normal" if can_finalize else "disabled")
    
    def ensure_consistent_target_month(self, current_project):
        """Ensure the project has a consistent target month detected and cached"""
        if not current_project or not current_project.source_file_path:

            pass
            return None
            
        # Check if we already have a cached target month
        cached_month = getattr(current_project, 'target_month', '')
        if cached_month and cached_month.strip():

            pass
            return cached_month
        
        # Check the target month cache before expensive operations
        cache_key = f"{current_project.name}:{current_project.source_file_path}"
        if cache_key in self.target_month_cache:
            _, header = self.target_month_cache[cache_key]
            if header and header != "Unknown Month":
                current_project.target_month = header
                return header
        
        # No cached month, need to detect and cache it
        try:
            import pandas as pd
            source_df = self._load_excel_with_cache(current_project.source_file_path, current_project.source_sheet)
            
            # Use the improved target month detection
            detected_month = self.get_source_month_header(source_df)
            
            if detected_month and detected_month != "Unknown Month":
                # Cache the detected month for consistency
                current_project.target_month = detected_month
                self.project_manager.save_settings()
                return detected_month
            else:

                pass
                return None
                
        except Exception as e:

                
            pass
            return None

    def clean_target_month_text(self, target_month):
        """Clean up target month text for display by removing 'Actual' and excess whitespace"""
        if not target_month:

            pass
            return target_month
            
        display_text = str(target_month).strip()
        
        # Remove "Actual" from the display text
        display_text = display_text.replace('Actual', '').replace('actual', '').strip()
        
        # If it's too long, try to extract just the month and year
        if len(display_text) > 15:
            # Try to extract month/year pattern
            import re
            month_year_match = re.search(r'(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s*(\d{4})', display_text, re.IGNORECASE)
            if month_year_match:
                display_text = f"{month_year_match.group(1)} {month_year_match.group(2)}"
        
        return display_text

    def update_target_month_display(self):
        """Update the target month display in the header"""
        try:
            current_project = self.project_manager.get_current_project()
            
            if not current_project:
                self.target_month_label.config(text="")
                return
            
            # Ensure consistent target month detection
            target_month = self.ensure_consistent_target_month(current_project)
            
            if target_month:
                # Clean up the target month text for better display
                display_text = self.clean_target_month_text(target_month)
                self.target_month_label.config(text=f"Target Month: {display_text}")
            else:
                self.target_month_label.config(text="Target Month: Not detected")
                    
        except Exception as e:
            self.target_month_label.config(text="")
    
    # Backup and Restore Functionality
    def create_backup(self):
        """Create a comprehensive backup of all projects and data"""
        try:
            import os
            import shutil
            import json
            from datetime import datetime
            
            # Create backup directory if it doesn't exist
            backup_dir = os.path.join(os.path.dirname(__file__), "backups")
            os.makedirs(backup_dir, exist_ok=True)
            
            # Generate backup filename with timestamp and target month
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # Get target month from current project or most recent project with target month
            target_month_str = ""
            current_project = self.project_manager.get_current_project()
            if current_project and hasattr(current_project, 'target_month') and current_project.target_month:
                target_month_str = f"_{current_project.target_month.replace(' ', '_').replace('/', '_')}"
            else:
                # If no current project target month, find any project with target month
                for project in self.project_manager.projects.values():
                    if hasattr(project, 'target_month') and project.target_month:
                        target_month_str = f"_{project.target_month.replace(' ', '_').replace('/', '_')}"
                        break
            
            backup_name = f"account_mapping_backup{target_month_str}_{timestamp}"
            backup_path = os.path.join(backup_dir, backup_name)
            os.makedirs(backup_path, exist_ok=True)
            
            # Create backup data structure
            backup_data = {
                "backup_timestamp": timestamp,
                "backup_date": datetime.now().isoformat(),
                "app_version": "2.0",
                "projects": {},
                "global_settings": {},
                "uploaded_files": {}
            }
            
            # Backup global settings (range_settings.json)
            range_settings_path = "range_settings.json"
            if os.path.exists(range_settings_path):
                backup_data["global_settings"]["range_settings"] = json.load(open(range_settings_path, 'r'))
                # Copy the file
                shutil.copy2(range_settings_path, os.path.join(backup_path, "range_settings.json"))
            
            # Backup project settings
            project_settings_path = self.project_manager.settings_file
            if os.path.exists(project_settings_path):
                backup_data["global_settings"]["project_settings"] = json.load(open(project_settings_path, 'r'))
                # Copy the file
                shutil.copy2(project_settings_path, os.path.join(backup_path, "project_settings.json"))
            
            # Create files directory in backup
            files_backup_dir = os.path.join(backup_path, "files")
            os.makedirs(files_backup_dir, exist_ok=True)
            
            # Backup each project's data and files
            for project_name, project in self.project_manager.projects.items():
                project_backup = {
                    "name": project.name,
                    "source_sheet": project.source_sheet,
                    "rolling_sheet": project.rolling_sheet,
                    "source_range": project.source_range,
                    "rolling_range": project.rolling_range,
                    "mappings": dict(project.mappings) if project.mappings else {},
                    "monthly_data": project.monthly_data if hasattr(project, 'monthly_data') else {},
                    "aggregated_data": project.aggregated_data if hasattr(project, 'aggregated_data') else {},
                    "preview_data": project.preview_data if hasattr(project, 'preview_data') else [],
                    "target_month": project.target_month if hasattr(project, 'target_month') else "",
                    "ui_state": project.ui_state if hasattr(project, 'ui_state') else {},
                    "workflow_state": project.workflow_state if hasattr(project, 'workflow_state') else {},
                    "sheet_ranges": project.sheet_ranges if hasattr(project, 'sheet_ranges') else {},
                    "files": {}
                }
                
                # Backup project files
                if project.source_file_path and os.path.exists(project.source_file_path):
                    source_filename = f"{project_name}_source_file.xlsx"
                    source_backup_path = os.path.join(files_backup_dir, source_filename)
                    shutil.copy2(project.source_file_path, source_backup_path)
                    project_backup["files"]["source_file"] = source_filename
                
                # Backup rolling workbook (shared across projects, but backup once per project that uses it)
                if (self.project_manager.rolling_workbook_path and 
                    os.path.exists(self.project_manager.rolling_workbook_path)):
                    rolling_filename = "rolling_workbook.xlsx"
                    rolling_backup_path = os.path.join(files_backup_dir, rolling_filename)
                    if not os.path.exists(rolling_backup_path):  # Only copy once
                        shutil.copy2(self.project_manager.rolling_workbook_path, rolling_backup_path)
                    project_backup["files"]["rolling_workbook"] = rolling_filename
                
                backup_data["projects"][project_name] = project_backup
            
            # Save backup metadata
            metadata_path = os.path.join(backup_path, "backup_metadata.json")
            with open(metadata_path, 'w', encoding='utf-8') as f:
                json.dump(backup_data, f, indent=2, ensure_ascii=False, default=str)
            
            # Show success message
            backup_size = self.get_folder_size(backup_path)
            messagebox.showinfo(
                "Backup Created", 
                f"Backup successfully created!\n\n"
                f"Location: {backup_path}\n"
                f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
                f"Size: {backup_size:.2f} MB\n"
                f"Projects: {len(self.project_manager.projects)}\n\n"
                f"Backup includes:\n"
                f"• All project data and mappings\n"
                f"• Uploaded Excel files\n"
                f"• Range settings\n"
                f"• Generated monthly statements"
            )
            
        except Exception as e:
            messagebox.showerror("Backup Error", f"Failed to create backup:\n{str(e)}")
    
    def get_folder_size(self, folder_path):
        """Calculate folder size in MB"""
        total_size = 0
        for dirpath, dirnames, filenames in os.walk(folder_path):
            for filename in filenames:
                filepath = os.path.join(dirpath, filename)
                if os.path.exists(filepath):
                    total_size += os.path.getsize(filepath)
        return total_size / (1024 * 1024)  # Convert to MB
    
    def show_backup_menu(self):
        """Show menu with available backup files"""
        try:
            import os
            from datetime import datetime
            
            backup_dir = os.path.join(os.path.dirname(__file__), "backups")
            
            if not os.path.exists(backup_dir):
                messagebox.showinfo("No Backups", "No backup files found. Create a backup first.")
                return
            
            # Find all backup folders
            backup_folders = []
            for item in os.listdir(backup_dir):
                item_path = os.path.join(backup_dir, item)
                if os.path.isdir(item_path) and item.startswith("account_mapping_backup"):
                    metadata_path = os.path.join(item_path, "backup_metadata.json")
                    if os.path.exists(metadata_path):
                        backup_folders.append((item, item_path, metadata_path))
            
            if not backup_folders:
                messagebox.showinfo("No Backups", "No valid backup files found.")
                return
            
            # Create backup selection window
            backup_window = tk.Toplevel(self.root)
            backup_window.title("Load from Backup")
            backup_window.transient(self.root)
            backup_window.grab_set()
            
            # Calculate center position immediately
            window_width = 900  # Much wider to ensure buttons are visible
            window_height = 500  # Increased to accommodate buttons
            x = (backup_window.winfo_screenwidth() - window_width) // 2
            y = (backup_window.winfo_screenheight() - window_height) // 2
            
            # Set geometry with center position from the start
            backup_window.geometry(f"{window_width}x{window_height}+{x}+{y}")
            
            # Create UI elements
            main_frame = ttk.Frame(backup_window)
            main_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=20)
            
            # Title
            title_label = ttk.Label(main_frame, text="Select Backup to Load", 
                                   font=("Arial", 14, "bold"))
            title_label.pack(pady=(0, 20))
            
            # Create treeview for backup list
            columns = ("Date", "Time", "Target Month", "Projects", "Size")
            tree = ttk.Treeview(main_frame, columns=columns, show="headings", height=10)
            
            # Configure columns
            tree.heading("Date", text="Date")
            tree.heading("Time", text="Time") 
            tree.heading("Target Month", text="Target Month")
            tree.heading("Projects", text="Projects")
            tree.heading("Size", text="Size (MB)")
            
            tree.column("Date", width=120, minwidth=100, stretch=tk.YES)
            tree.column("Time", width=100, minwidth=80, stretch=tk.YES)
            tree.column("Target Month", width=160, minwidth=140, stretch=tk.YES)
            tree.column("Projects", width=80, minwidth=60, stretch=tk.NO)
            tree.column("Size", width=100, minwidth=80, stretch=tk.NO)
            
            # Add scrollbar
            scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=tree.yview)
            tree.configure(yscrollcommand=scrollbar.set)
            
            # Pack treeview and scrollbar
            tree_frame = ttk.Frame(main_frame)
            tree_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 20))
            
            tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # Populate backup list
            backup_data_map = {}
            for folder_name, folder_path, metadata_path in sorted(backup_folders, reverse=True):
                try:
                    import json
                    with open(metadata_path, 'r', encoding='utf-8') as f:
                        metadata = json.load(f)
                    
                    # Parse timestamp
                    timestamp = metadata.get("backup_timestamp", "")
                    if timestamp:
                        dt = datetime.strptime(timestamp, "%Y%m%d_%H%M%S")
                        date_str = dt.strftime("%Y-%m-%d")
                        time_str = dt.strftime("%H:%M:%S")
                    else:
                        date_str = "Unknown"
                        time_str = "Unknown"
                    
                    projects_count = len(metadata.get("projects", {}))
                    folder_size = self.get_folder_size(folder_path)
                    
                    # Extract target month from folder name or project data
                    target_month = "N/A"
                    # First try to extract from folder name
                    import re
                    folder_parts = folder_name.split('_')
                    if len(folder_parts) >= 4:  # account_mapping_backup_[month]_[timestamp]
                        # Check if there's a month part before the timestamp
                        potential_month = '_'.join(folder_parts[3:-2]) if len(folder_parts) > 4 else folder_parts[3]
                        if potential_month and not re.match(r'^\d{8}$', potential_month):  # Not just a date
                            target_month = potential_month.replace('_', ' ')
                    
                    # If not found in folder name, try to get from project data
                    if target_month == "N/A":
                        projects = metadata.get("projects", {})
                        for project_data in projects.values():
                            if project_data.get("target_month"):
                                target_month = project_data["target_month"]
                                break
                    
                    item_id = tree.insert("", "end", values=(date_str, time_str, target_month, projects_count, f"{folder_size:.1f}"))
                    backup_data_map[item_id] = {
                        "folder_path": folder_path,
                        "metadata": metadata,
                        "folder_name": folder_name
                    }
                    
                except Exception as e:
                    pass  # Debug output removed
            button_frame = ttk.Frame(main_frame)
            button_frame.pack(fill=tk.X, pady=(20, 0))
            
            def load_selected_backup():
                selection = tree.selection()
                if not selection:
                    messagebox.showwarning("No Selection", "Please select a backup to load.")
                    return
                
                item_id = selection[0]
                backup_info = backup_data_map[item_id]
                
                # Confirm load
                result = messagebox.askyesno(
                    "Confirm Load Backup",
                    f"This will replace all current data with the backup from:\n"
                    f"{backup_info['metadata'].get('backup_date', 'Unknown date')}\n\n"
                    f"Projects in backup: {len(backup_info['metadata'].get('projects', {}))}\n\n"
                    f"Are you sure you want to continue?"
                )
                
                if result:
                    backup_window.destroy()
                    self.load_backup(backup_info['folder_path'], backup_info['metadata'])
            
            def close_window():
                backup_window.destroy()
            
            # Buttons - centered
            load_button = ttk.Button(button_frame, text="Load Selected", command=load_selected_backup)
            cancel_button = ttk.Button(button_frame, text="Cancel", command=close_window)
            
            # Center buttons using grid for better control
            button_frame.columnconfigure(0, weight=1)
            button_frame.columnconfigure(1, weight=0)
            button_frame.columnconfigure(2, weight=0)
            button_frame.columnconfigure(3, weight=1)
            
            load_button.grid(row=0, column=1, padx=(0, 10))
            cancel_button.grid(row=0, column=2)
            
            # Double-click to load
            tree.bind("<Double-1>", lambda e: load_selected_backup())
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to show backup menu:\n{str(e)}")
    
    def load_backup(self, backup_path, metadata):
        """Load backup data and restore all projects"""
        try:
            import os
            import shutil
            import json
            from datetime import datetime
            
            # Set flag to force loading from backup
            self.project_manager._force_backup_load = True
            
            # Clear current data
            self.project_manager.reset_all_projects()
            
            # Restore global settings
            if "global_settings" in metadata:
                # Restore range_settings.json
                if "range_settings" in metadata["global_settings"]:

                    pass
                    with open("range_settings.json", 'w', encoding='utf-8') as f:
                        json.dump(metadata["global_settings"]["range_settings"], f, indent=2)
            
            # Restore uploaded files
            files_backup_dir = os.path.join(backup_path, "files")
            app_dir = os.path.dirname(__file__)
            
            # Create a restored_files directory
            restored_files_dir = os.path.join(app_dir, "restored_files")
            os.makedirs(restored_files_dir, exist_ok=True)
            
            # Restore projects
            for project_name, project_data in metadata.get("projects", {}).items():
                # Create new project
                project = Project(
                    name=project_data["name"],
                    source_sheet=project_data["source_sheet"],
                    rolling_sheet=project_data.get("rolling_sheet")
                )
                
                # Restore project settings
                project.source_range = project_data.get("source_range", "")
                project.rolling_range = project_data.get("rolling_range", "")
                project.mappings = OrderedDict(project_data.get("mappings", {}))
                project.monthly_data = project_data.get("monthly_data", {})
                project.aggregated_data = project_data.get("aggregated_data", {})
                project.preview_data = project_data.get("preview_data", [])
                project.target_month = project_data.get("target_month", "")
                project.ui_state = project_data.get("ui_state", {})
                project.workflow_state = project_data.get("workflow_state", {})
                project.sheet_ranges = project_data.get("sheet_ranges", {})
                
                # Restore project files
                if "files" in project_data:

                    pass
                    if "source_file" in project_data["files"]:
                        source_file_backup = os.path.join(files_backup_dir, project_data["files"]["source_file"])
                        if os.path.exists(source_file_backup):
                            # Restore to new location
                            restored_source = os.path.join(restored_files_dir, f"{project_name}_source.xlsx")
                            shutil.copy2(source_file_backup, restored_source)
                            project.source_file_path = restored_source
                    
                    if "rolling_workbook" in project_data["files"]:
                        rolling_file_backup = os.path.join(files_backup_dir, project_data["files"]["rolling_workbook"])
                        if os.path.exists(rolling_file_backup):
                            # Restore rolling workbook to new location
                            restored_rolling = os.path.join(restored_files_dir, "rolling_workbook.xlsx")
                            if not os.path.exists(restored_rolling):  # Only copy once
                                shutil.copy2(rolling_file_backup, restored_rolling)
                            self.project_manager.rolling_workbook_path = restored_rolling
                
                # Add project to manager
                self.project_manager.add_project(project)
            
            # Save restored settings
            self.project_manager.save_settings()
            
            # Refresh UI
            self.refresh_project_menu()
            self.load_project_data()
            self.update_ui_state()
            
            # Show success message
            messagebox.showinfo(
                "Backup Loaded",
                f"Backup successfully loaded!\n\n"
                f"Backup date: {metadata.get('backup_date', 'Unknown')}\n"
                f"Projects restored: {len(metadata.get('projects', {}))}\n"
                f"Files restored to: {restored_files_dir}\n\n"
                f"All project data, mappings, and files have been restored."
            )
            
        except Exception as e:
            messagebox.showerror("Restore Error", f"Failed to load backup:\n{str(e)}")
            # Try to reload existing settings if restore failed
            try:
                self.project_manager.load_settings()
                self.refresh_project_menu()
                self.load_project_data()
                self.update_ui_state()
            except:
                pass
    
    def pop_out_step3_window(self):
        """Pop out the Step 3 preview tree into a separate window"""
        if self.step3_is_popped_out:

            pass
            return
            
        # Create popup window but don't show it yet
        self.step3_popup_window = tk.Toplevel(self.root)
        self.step3_popup_window.withdraw()  # Hide window initially
        
        # Set title with current project name
        current_project = self.project_manager.get_current_project()
        project_name = current_project.name if current_project else "No Project"
        target_month = self.ensure_consistent_target_month(current_project) if current_project else None
        if target_month:
            clean_target_month = self.clean_target_month_text(target_month)
            target_month_text = f" - {clean_target_month}"
        else:
            target_month_text = ""
        self.step3_popup_window.title(f"Step 3: Rolling P&L Statement Preview - {project_name}{target_month_text} - Pop Out")
        self.step3_popup_window.geometry("1600x900")  # Increased width for better column visibility
        self.step3_popup_window.minsize(800, 600)  # Match Step 2 popup resize limits for consistency
        
        # Make it resizable
        self.step3_popup_window.resizable(True, True)
        
        # Create main frame in popup window
        popup_main_frame = ttk.Frame(self.step3_popup_window)
        popup_main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Create header with target month and pop-in button
        header_frame = ttk.Frame(popup_main_frame)
        header_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Add target month label if available
        if target_month:
            clean_target_month = self.clean_target_month_text(target_month)
            target_month_label = ttk.Label(header_frame, text=f"Target Month: {clean_target_month}", 
                                         font=("Arial", 14, "bold"), foreground="darkblue")
            target_month_label.pack(side=tk.TOP, anchor=tk.W, pady=(0, 5))
        
        # Create instruction and button frame
        instruction_frame = ttk.Frame(header_frame)
        instruction_frame.pack(fill=tk.X)
        
        header_label = ttk.Label(instruction_frame, text="📊 Rolling P&L Statement Preview - Previous 2 Months + Target Month", 
                               font=("Arial", 12, "bold"))
        header_label.pack(side=tk.LEFT, padx=5)
        
        # Add Finalize & Export button for convenience
        project_name = current_project.name if current_project else "Project"
        self.step3_popup_finalize_button = ttk.Button(instruction_frame, text=f"✅ Finalize & Export Final Excel for {project_name}", 
                                                     command=self.finalize_and_export, width=40)
        self.step3_popup_finalize_button.pack(side=tk.RIGHT, padx=5)
        
        # Add Generate Monthly Statement button for convenience
        self.step3_popup_generate_button = ttk.Button(instruction_frame, text="📊 Generate Monthly Statement", 
                                                     command=self.generate_monthly_statement, width=25)
        self.step3_popup_generate_button.pack(side=tk.RIGHT, padx=5)
        
        pop_in_button = ttk.Button(instruction_frame, text="🔳 Pop In", command=self.pop_in_step3_window, width=12)
        pop_in_button.pack(side=tk.RIGHT, padx=5)
        
        # Create content frame for tree and scrollbar
        content_frame = ttk.Frame(popup_main_frame)
        content_frame.pack(fill=tk.BOTH, expand=True)
        content_frame.columnconfigure(0, weight=1)
        content_frame.rowconfigure(0, weight=1)
        
        # Create new tree view for popup (clone of original preview tree)
        preview_columns = ("Account Description", "Previous Month -2", "Previous Month -1", "Target Month")
        self.step3_popup_tree = ttk.Treeview(content_frame, columns=preview_columns, show="headings")
        
        # Configure same style as original
        style = ttk.Style()
        style.configure("Large.Treeview", font=("Arial", 14))
        self.step3_popup_tree.configure(style="Large.Treeview")
        
        # Configure columns with same settings as main preview tree
        self.step3_popup_tree.heading("Account Description", text="Account Description", anchor=tk.W)
        self.step3_popup_tree.heading("Previous Month -2", text="Previous Month -2", anchor=tk.W)
        self.step3_popup_tree.heading("Previous Month -1", text="Previous Month -1", anchor=tk.W)
        self.step3_popup_tree.heading("Target Month", text="Target Month", anchor=tk.W)
        
        self.step3_popup_tree.column("Account Description", width=600)  # Wider for account names
        self.step3_popup_tree.column("Previous Month -2", width=250, anchor="e")  # Wider for better visibility
        self.step3_popup_tree.column("Previous Month -1", width=250, anchor="e")  # Wider for better visibility
        self.step3_popup_tree.column("Target Month", width=250, anchor="e")  # Wider for better visibility
        
        # Configure same tags for formatting
        self.step3_popup_tree.tag_configure("heading", font=("Arial", 16, "bold"))
        self.step3_popup_tree.tag_configure("normal", font=("Arial", 14))
        
        # Create scrollbar for popup tree
        popup_scrollbar = ttk.Scrollbar(content_frame, orient="vertical", command=self.step3_popup_tree.yview)
        self.step3_popup_tree.configure(yscrollcommand=popup_scrollbar.set)
        
        # Position tree and scrollbar
        self.step3_popup_tree.grid(row=0, column=0, sticky="nsew", padx=(0, 5))
        popup_scrollbar.grid(row=0, column=1, sticky="ns")
        
        # Update popup tree column headings with actual month names (same as main tree) - do this first
        self.update_step3_popup_headings()
        
        # Copy all data from original preview tree to popup tree
        self.sync_preview_tree_data(self.preview_tree, self.step3_popup_tree)
        
        # Now that everything is ready, center and show the window
        self.center_step3_popup_window()
        self.step3_popup_window.deiconify()  # Show the window
        
        # Hide original preview tree
        self.preview_tree.grid_remove()
        
        # Update button text and state
        self.step3_popup_button.config(text="🔳 Popped Out", state="disabled")
        self.step3_is_popped_out = True
        
        # Update popup generate button style based on current modification state
        self.update_popup_button_style()
        
        # Set initial state for popup finalize button
        self.update_popup_finalize_button_state()
        
        # Set up window close protocol
        self.step3_popup_window.protocol("WM_DELETE_WINDOW", self.pop_in_step3_window)

    def pop_in_step3_window(self):
        """Pop the Step 3 preview back into the main window"""
        if not self.step3_is_popped_out:

            pass
            return
            
        # Sync data back from popup to main tree
        self.sync_preview_tree_data(self.step3_popup_tree, self.preview_tree)
        
        # Show the original tree
        self.preview_tree.grid()
        
        # Destroy popup window
        if self.step3_popup_window:
            self.step3_popup_window.destroy()
            self.step3_popup_window = None
            self.step3_popup_tree = None
            self.step3_popup_generate_button = None
            self.step3_popup_finalize_button = None
        
        # Update button state
        self.step3_popup_button.config(text="🔲 Pop Out", state="normal")
        self.step3_is_popped_out = False

    def center_step3_popup_window(self):
        """Center the Step 3 popup window on the screen"""
        if not self.step3_popup_window:

            pass
            return
            
        # Use the geometry dimensions we set (1600x900)
        width = 1600
        height = 900
        
        # Get screen dimensions
        screen_width = self.step3_popup_window.winfo_screenwidth()
        screen_height = self.step3_popup_window.winfo_screenheight()
        
        # Calculate center position
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        
        # Ensure window doesn't go off screen
        if x < 0:
            x = 0
        if y < 0:
            y = 0
            
        # Set window position
        self.step3_popup_window.geometry(f"{width}x{height}+{x}+{y}")

    def sync_preview_tree_data(self, source_tree, target_tree):
        """Sync data between preview trees - optimized version"""
        if not source_tree or not target_tree:

            pass
            return
            
        # Get all source items at once
        source_items = source_tree.get_children()
        if not source_items:

            pass
            return
            
        # Clear target tree
        target_tree.delete(*target_tree.get_children())
        
        # Batch insert all items
        for item in source_items:
            values = source_tree.item(item, "values")
            tags = source_tree.item(item, "tags")
            target_tree.insert("", "end", values=values, tags=tags)

    def update_step3_popup_headings(self):
        """Update Step 3 popup tree column headings with actual month names"""
        if not self.step3_popup_tree or not self.preview_tree:

            pass
            return
            
        # Get the headings from the main preview tree (already loaded)
        try:
            heading1 = self.preview_tree.heading("Previous Month -2")["text"]
            heading2 = self.preview_tree.heading("Previous Month -1")["text"] 
            heading3 = self.preview_tree.heading("Target Month")["text"]
            
            # Update popup tree headings with the same headings as main tree
            self.step3_popup_tree.heading("Previous Month -2", text=heading1)
            self.step3_popup_tree.heading("Previous Month -1", text=heading2)  
            self.step3_popup_tree.heading("Target Month", text=heading3)
        except Exception:
            # Fallback to generic headings if something goes wrong
            pass


def main():
    """Main application entry point"""
    root = tk.Tk()
    app = MultiProjectAccountMappingApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()