#!/usr/bin/env python3
"""
Project Manager for Multi-Project Account Mapping Application

Handles multiple projects, settings persistence, and project data isolation.
"""

import json
import os
from typing import Dict, List, Optional, Any
from collections import OrderedDict
import pandas as pd


class Project:
    """Individual project data container"""
    
    def __init__(self, name: str, source_sheet: str, rolling_sheet: Optional[str] = None):
        self.name = name
        self.source_sheet = source_sheet
        self.rolling_sheet = rolling_sheet
        
        # Project-specific settings
        self.source_range = ""
        self.rolling_range = ""
        self.mapping_file_path = ""
        
        # Project-specific data (persistent)
        self.source_accounts = []
        self.rolling_accounts = []
        self.mappings = OrderedDict()
        
        # File paths (persistent)
        self.source_file_path = ""
        
        # Step 4 data (persistent)
        self.monthly_data = {}
        self.aggregated_data = {}
        self.preview_data = []
        self.target_month = ""
        self.step4_completed = False
        self.last_export_file = ""
        
        # UI state (persistent)
        self.ui_state = {
            'filter_value': '',
            'sort_value': '',
            'zoom_level': 1.0,
            'checkbox_states': {},
            'last_active_step': 1,
            'window_geometry': {}
        }
        
        # Workflow state (persistent)
        self.workflow_state = {
            'step1_complete': False,
            'step2_complete': False,
            'step3_complete': False,
            'step4_complete': False,
            'last_status_message': '',
            'has_generated_mappings': False,
            'has_generated_monthly': False
        }
        
        # Runtime data (not persistent - cleared on startup)
        self.filtered_mappings = OrderedDict()
        self.checkbox_states = {}  # Kept for backward compatibility
        
        # Range memory per sheet (persistent)
        self.sheet_ranges = {}  # Format: {"sheet_name": {"source": "A2:A100", "rolling": "B2:B150"}}
        
    def to_dict(self) -> Dict[str, Any]:
        """Convert project to dictionary for JSON serialization"""
        return {
            # Basic project info
            'name': self.name,
            'source_sheet': self.source_sheet,
            'rolling_sheet': self.rolling_sheet,
            'source_range': self.source_range,
            'rolling_range': self.rolling_range,
            'mapping_file_path': self.mapping_file_path,
            'source_file_path': self.source_file_path,
            
            # Core project data
            'mappings': dict(self.mappings) if self.mappings else {},
            'source_accounts': self.source_accounts,
            'rolling_accounts': self.rolling_accounts,
            
            # Step 4 data (persistent)
            'monthly_data': self.monthly_data,
            'aggregated_data': self.aggregated_data,
            'preview_data': self.preview_data,
            'target_month': self.target_month,
            'step4_completed': self.step4_completed,
            'last_export_file': self.last_export_file,
            
            # UI state (persistent)
            'ui_state': self.ui_state,
            
            # Workflow state (persistent)
            'workflow_state': self.workflow_state,
            
            # Range memory (persistent)
            'sheet_ranges': self.sheet_ranges
        }
    
    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> 'Project':
        """Create project from dictionary"""
        project = cls(
            name=data['name'],
            source_sheet=data['source_sheet'],
            rolling_sheet=data.get('rolling_sheet')
        )
        
        # Basic project settings
        project.source_range = data.get('source_range', '')
        project.rolling_range = data.get('rolling_range', '')
        project.mapping_file_path = data.get('mapping_file_path', '')
        project.source_file_path = data.get('source_file_path', '')
        project.source_accounts = data.get('source_accounts', [])
        project.rolling_accounts = data.get('rolling_accounts', [])
        
        # Convert mappings back to OrderedDict
        mappings_data = data.get('mappings', {})
        if mappings_data:
            project.mappings = OrderedDict(mappings_data)
        
        # Step 4 data (persistent)
        project.monthly_data = data.get('monthly_data', {})
        project.aggregated_data = data.get('aggregated_data', {})
        project.preview_data = data.get('preview_data', [])
        project.target_month = data.get('target_month', '')
        project.step4_completed = data.get('step4_completed', False)
        project.last_export_file = data.get('last_export_file', '')
        
        # UI state (persistent)
        default_ui_state = {
            'filter_value': '',
            'sort_value': '',
            'zoom_level': 1.0,
            'checkbox_states': {},
            'last_active_step': 1,
            'window_geometry': {}
        }
        project.ui_state = data.get('ui_state', default_ui_state)
        
        # Workflow state (persistent)
        default_workflow_state = {
            'step1_complete': False,
            'step2_complete': False,
            'step3_complete': False,
            'step4_complete': False,
            'last_status_message': '',
            'has_generated_mappings': False,
            'has_generated_monthly': False
        }
        project.workflow_state = data.get('workflow_state', default_workflow_state)
        
        # Load sheet ranges
        project.sheet_ranges = data.get('sheet_ranges', {})
        
        return project
    
    def clear_runtime_data(self):
        """Clear runtime data that shouldn't be persisted"""
        # Only clear truly runtime data - Step 4 data is now persistent
        self.filtered_mappings.clear()
        self.checkbox_states.clear()  # This is kept for backward compatibility
    
    def clear_all_project_data(self):
        """Clear all project data - used for project reset"""
        # Preserve range memory before clearing
        preserved_source_range = self.source_range
        preserved_rolling_range = self.rolling_range
        preserved_sheet_ranges = self.sheet_ranges.copy() if self.sheet_ranges else {}
        
        # Clear all persistent data
        self.source_accounts.clear()
        self.rolling_accounts.clear()
        self.mappings.clear()
        self.monthly_data.clear()
        self.aggregated_data.clear()
        self.preview_data.clear()
        
        # Reset file paths and settings but preserve ranges
        self.source_range = preserved_source_range  # Keep range memory
        self.rolling_range = preserved_rolling_range  # Keep range memory
        self.sheet_ranges = preserved_sheet_ranges  # Keep sheet-specific ranges
        self.mapping_file_path = ""
        self.source_file_path = ""
        self.target_month = ""
        self.step4_completed = False
        self.last_export_file = ""
        
        # Reset UI state
        self.ui_state = {
            'filter_value': '',
            'sort_value': '',
            'zoom_level': 1.0,
            'checkbox_states': {},
            'last_active_step': 1,
            'window_geometry': {}
        }
        
        # Reset workflow state
        self.workflow_state = {
            'step1_complete': False,
            'step2_complete': False,
            'step3_complete': False,
            'step4_complete': False,
            'last_status_message': '',
            'has_generated_mappings': False,
            'has_generated_monthly': False
        }
        
        # Clear runtime data
        self.clear_runtime_data()


class ProjectManager:
    """Manages multiple projects and handles persistence"""
    
    def __init__(self, settings_file: str = "project_settings.json"):
        self.settings_file = settings_file
        self.projects: Dict[str, Project] = {}
        self.current_project: Optional[Project] = None
        
        # File paths
        self.source_workbook_path = ""
        self.rolling_workbook_path = ""
        
        # Persistent range memory that survives "start fresh"
        self.range_memory_file = "range_memory.json"
        self.persistent_range_memory = {}  # Format: {project_name: {source_range, rolling_range, sheet_ranges}}
        
        # Load existing settings
        self.load_settings()
        self.load_range_memory()
    
    def add_project(self, project: Project) -> None:
        """Add a project to the manager"""
        self.projects[project.name] = project
        
        # Set as current if it's the first project
        # IMPORTANT: Use the object from the dictionary to ensure consistency
        if self.current_project is None:
            self.current_project = self.projects[project.name]
            pass
    
    def select_project(self, project_name: str) -> bool:
        """Select a project as current"""
        if project_name in self.projects:
            # Always use the actual project object from the dictionary
            self.current_project = self.projects[project_name]
            pass
            return True
        return False
    
    def get_project_names(self) -> List[str]:
        """Get list of all project names"""
        return list(self.projects.keys())
    
    def get_current_project(self) -> Optional[Project]:
        """Get currently selected project"""
        return self.current_project
    
    def remove_project(self, project_name: str) -> bool:
        """Remove a project"""
        if project_name in self.projects:
            # If removing current project, select another or set to None
            if self.current_project and self.current_project.name == project_name:
                remaining_projects = [name for name in self.projects.keys() if name != project_name]
                if remaining_projects:
                    self.current_project = self.projects[remaining_projects[0]]
                else:
                    self.current_project = None
            
            del self.projects[project_name]
            return True
        return False
    
    def scan_source_workbook(self, workbook_path: str) -> List[Dict[str, str]]:
        """
        Scan source workbook for projects
        Returns list of {'sheet_name': str, 'project_name': str}
        """
        projects_found = []
        
        try:
            # Read all sheet names
            xl_file = pd.ExcelFile(workbook_path)
            
            for sheet_name in xl_file.sheet_names:
                try:
                    # Use openpyxl for more reliable cell A1 reading
                    import openpyxl
                    wb = openpyxl.load_workbook(workbook_path, data_only=True)
                    sheet = wb[sheet_name]
                    project_name = sheet['A1'].value
                    
                    if project_name and str(project_name).strip() and str(project_name).strip().lower() != 'nan':
                        project_name = str(project_name).strip()
                        projects_found.append({
                            'sheet_name': sheet_name,
                            'project_name': project_name
                        })
                
                except Exception as e:
                    pass
                    # Skip problematic sheets
                    continue
        
        except Exception as e:
            pass
            return []
        
        return projects_found
    
    def create_projects_from_workbook(self, workbook_path: str) -> int:
        """
        Create projects from source workbook or update existing projects
        Returns number of projects created/updated
        """
        # First, save current project ranges to persistent memory before any operations
        for proj_name, proj in self.projects.items():
            if proj.source_range or proj.rolling_range or proj.sheet_ranges:
                self.store_project_ranges(
                    proj_name, 
                    proj.source_range, 
                    proj.rolling_range, 
                    proj.sheet_ranges.copy() if proj.sheet_ranges else {}
                )
        
        projects_data = self.scan_source_workbook(workbook_path)
        
        if not projects_data:
            return 0
        
        # Set source workbook path
        self.source_workbook_path = workbook_path
        
        # Load saved project data - ALWAYS prioritize project_settings.json over backup
        saved_project_data = {}
        
        # Check if we should force backup load (only when user explicitly chooses "Load from Backup")
        force_backup = getattr(self, '_force_backup_load', False)
        
        # FIRST: Always try to load from project_settings.json (unless forcing backup)
        if os.path.exists(self.settings_file) and not force_backup:
            pass
            try:
                with open(self.settings_file, 'r', encoding='utf-8') as f:
                    saved_settings = json.load(f)
                saved_projects = saved_settings.get('projects', {})
                pass
                for proj_name, proj_data in saved_projects.items():
                    mappings_count = len(proj_data.get('mappings', {}))
                    pass
                    if proj_data.get('mappings'):  # Only include projects with mappings
                        saved_project_data[proj_name] = proj_data
            except Exception as e:
                pass
                saved_project_data = {}
        # SECOND: Only use backup if explicitly requested or no settings file exists
        elif hasattr(self, '_backup_settings') and self._backup_settings:
            pass
            saved_settings = self._backup_settings
            saved_projects = saved_settings.get('projects', {})
            pass
            for proj_name, proj_data in saved_projects.items():
                mappings_count = len(proj_data.get('mappings', {}))
                pass
                if proj_data.get('mappings'):  # Only include projects with mappings
                    saved_project_data[proj_name] = proj_data
            # Clear backup after use
            self._backup_settings = None
            self._force_backup_load = False  # Reset flag
        else:
            pass
        
        # Check if we should update existing projects or create new ones
        existing_projects = list(self.projects.keys())
        scanned_projects = [proj['project_name'] for proj in projects_data]
        
        # Always try to preserve existing projects when possible
        # Even if projects don't match exactly, preserve rolling sheet and other data
        projects_match = existing_projects and set(existing_projects) == set(scanned_projects)
        pass
        
        # If we have existing projects, always try to preserve them unless forced to recreate
        has_existing_projects = bool(existing_projects)
        pass
        
        if projects_match:
            # Update existing projects with new source file path
            for project_data in projects_data:
                project_name = project_data['project_name']
                if project_name in self.projects:
                    existing_project = self.projects[project_name]
                    # Update source file path and sheet (in case sheet name changed)
                    existing_project.source_file_path = workbook_path
                    existing_project.source_sheet = project_data['sheet_name']
            
            # Save settings immediately after updating
            self.save_settings()
            
            return len(projects_data)
        elif has_existing_projects:
            # Projects don't match exactly, but we have existing data - preserve what we can
            pass
            
            # Store existing project data before clearing, but prefer saved data if available
            existing_project_data = {}
            for proj_name, proj in self.projects.items():
                existing_project_data[proj_name] = proj.to_dict()
                pass
            
            # Merge with saved project data (saved data takes precedence)
            for proj_name, proj_data in saved_project_data.items():
                existing_project_data[proj_name] = proj_data
                pass
            
            # Clear existing projects and create new ones
            self.projects.clear()
            self.current_project = None
        else:
            # No existing projects - create fresh but can use saved project data
            pass
            existing_project_data = saved_project_data
        
        # Create project objects
        for project_data in projects_data:
            project_name = project_data['project_name']
            
            # Check if we have existing data for this project
            if project_name in existing_project_data:
                # Restore from existing data
                project = Project.from_dict(existing_project_data[project_name])
                # Update the source file path and sheet in case they changed
                project.source_file_path = workbook_path
                project.source_sheet = project_data['sheet_name']
                pass
            else:
                # Check if we have backed up data for this project (from existing_project_data)
                pass
                if project_name in existing_project_data:
                    # Restore from backed up data
                    project = Project.from_dict(existing_project_data[project_name])
                    # Update the source file path and sheet in case they changed
                    project.source_file_path = workbook_path
                    project.source_sheet = project_data['sheet_name']
                    pass
                else:
                    pass
                    # Create new project
                    project = Project(
                        name=project_name,
                        source_sheet=project_data['sheet_name']
                    )
                    # Set the source file path for each project
                    project.source_file_path = workbook_path
                    
                    # Restore range memory from persistent storage
                    stored_ranges = self.get_project_ranges(project_name)
                    if stored_ranges:
                        project.source_range = stored_ranges.get('source_range', '')
                        project.rolling_range = stored_ranges.get('rolling_range', '')
                        project.sheet_ranges = stored_ranges.get('sheet_ranges', {})
                        pass
                        
                        # Try to restore rolling sheet from sheet_ranges if available
                        if project.sheet_ranges:
                            # Use the first sheet name from sheet_ranges as the rolling sheet
                            first_sheet_name = next(iter(project.sheet_ranges.keys()), None)
                            if first_sheet_name:
                                project.rolling_sheet = first_sheet_name
                                pass
            
            self.add_project(project)
        
        return len(projects_data)
    
    def get_rolling_sheets(self, workbook_path: str) -> List[str]:
        """Get list of sheet names from rolling workbook"""
        try:
            # Try with openpyxl first (handles most Excel files)
            import openpyxl
            wb = openpyxl.load_workbook(workbook_path, data_only=True)
            return wb.sheetnames
        except Exception as e:
            try:
                # Fallback to pandas
                xl_file = pd.ExcelFile(workbook_path)
                return xl_file.sheet_names
            except Exception as e2:
                pass
                return []
    
    def set_rolling_workbook(self, workbook_path: str) -> None:
        """Set rolling workbook path"""
        self.rolling_workbook_path = workbook_path
    
    def save_settings(self) -> None:
        """Save current settings to JSON file"""
        settings_data = {
            'source_workbook_path': self.source_workbook_path,
            'rolling_workbook_path': self.rolling_workbook_path,
            'current_project': self.current_project.name if self.current_project else None,
            'projects': {}
        }
        
        # Save each project
        for project_name, project in self.projects.items():
            # Clear runtime data before saving
            project.clear_runtime_data()
            project_dict = project.to_dict()
            settings_data['projects'][project_name] = project_dict
            pass
        
        try:
            with open(self.settings_file, 'w', encoding='utf-8') as f:
                json.dump(settings_data, f, indent=2, ensure_ascii=False)
            
            pass
        except Exception as e:
            pass
    
    def load_settings(self) -> None:
        """Load settings from JSON file"""
        if not os.path.exists(self.settings_file):
            return
        
        try:
            with open(self.settings_file, 'r', encoding='utf-8') as f:
                settings_data = json.load(f)
            
            # Load file paths
            self.source_workbook_path = settings_data.get('source_workbook_path', '')
            self.rolling_workbook_path = settings_data.get('rolling_workbook_path', '')
            
            # Load projects
            projects_data = settings_data.get('projects', {})
            for project_name, project_data in projects_data.items():
                project = Project.from_dict(project_data)
                self.projects[project_name] = project
            
            # Set current project - ALWAYS use object from dictionary
            current_project_name = settings_data.get('current_project')
            if current_project_name and current_project_name in self.projects:
                self.current_project = self.projects[current_project_name]
                pass
            elif self.projects:
                # Default to first project if current not found
                first_project_name = list(self.projects.keys())[0]
                self.current_project = self.projects[first_project_name]
                pass
        
        except Exception as e:
            pass
    
    def load_range_memory(self) -> None:
        """Load persistent range memory from JSON file"""
        if not os.path.exists(self.range_memory_file):
            return
        
        try:
            with open(self.range_memory_file, 'r', encoding='utf-8') as f:
                self.persistent_range_memory = json.load(f)
        except Exception as e:
            pass
            self.persistent_range_memory = {}
    
    def save_range_memory(self) -> None:
        """Save persistent range memory to JSON file"""
        try:
            with open(self.range_memory_file, 'w', encoding='utf-8') as f:
                json.dump(self.persistent_range_memory, f, indent=2, ensure_ascii=False)
        except Exception as e:
            pass
    
    def store_project_ranges(self, project_name: str, source_range: str, rolling_range: str, sheet_ranges: dict = None) -> None:
        """Store range memory for a project"""
        self.persistent_range_memory[project_name] = {
            'source_range': source_range,
            'rolling_range': rolling_range,
            'sheet_ranges': sheet_ranges or {}
        }
        self.save_range_memory()
        pass
    
    def get_project_ranges(self, project_name: str) -> dict:
        """Get stored range memory for a project"""
        return self.persistent_range_memory.get(project_name, {})
    
    def has_projects(self) -> bool:
        """Check if any projects exist"""
        return len(self.projects) > 0
    
    def reset_all_projects(self, preserve_rolling_workbook: bool = True) -> None:
        """Reset all project data but keep a backup of settings for mapping restoration"""
        # Create a backup of current settings before clearing
        backup_settings = None
        if os.path.exists(self.settings_file):
            try:
                with open(self.settings_file, 'r', encoding='utf-8') as f:
                    backup_settings = json.load(f)
                pass
            except Exception as e:
                pass
        
        self.projects.clear()
        self.current_project = None
        self.source_workbook_path = ""
        # Preserve rolling workbook path for Start Fresh workflow so new projects can use it
        if not preserve_rolling_workbook:
            self.rolling_workbook_path = ""
        
        # Save the cleared state to persist the reset
        self.save_settings()
        
        # Store backup for later restoration
        self._backup_settings = backup_settings